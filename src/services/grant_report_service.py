"""Grant compliance report generation service."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Literal, Optional

from sqlalchemy import and_
from sqlalchemy.orm import Session, joinedload

from src.models import Grant, Expense, Student, Member
from src.models.grant_enrollment import GrantEnrollment
from src.services.grant_metrics_service import GrantMetricsService
from src.db.enums import GrantEnrollmentStatus


class GrantReportService:
    """Generate compliance reports for grants."""

    def __init__(self, db: Session):
        self.db = db
        self.metrics = GrantMetricsService(db)

    def generate_report(
        self,
        grant_id: int,
        report_type: Literal["summary", "detailed", "funder"],
        period_start: Optional[date] = None,
        period_end: Optional[date] = None,
    ) -> Optional[dict]:
        """Generate a compliance report."""
        summary = self.metrics.get_grant_summary(grant_id)

        if not summary:
            return None

        if report_type == "summary":
            return self._format_summary_report(summary)
        elif report_type == "detailed":
            return self._format_detailed_report(grant_id, summary)
        elif report_type == "funder":
            return self._format_funder_report(
                grant_id, summary, period_start, period_end
            )

        return None

    def _format_summary_report(self, summary: dict) -> dict:
        """Format executive summary report."""
        return {
            "report_type": "summary",
            "generated_at": datetime.utcnow().isoformat(),
            "grant": {
                "id": summary["grant_id"],
                "name": summary["grant_name"],
                "grant_number": summary["grant_number"],
                "funder": summary["funder"],
                "status": summary["status"].value if summary["status"] else None,
                "period": summary["period"],
            },
            "highlights": {
                "total_enrolled": summary["enrollment"]["total_enrolled"],
                "currently_active": summary["enrollment"]["currently_active"],
                "completed": summary["enrollment"]["completed"],
                "retention_rate": f"{summary['enrollment']['retention_rate']}%",
                "budget_utilization": f"{summary['financials']['utilization_rate']}%",
                "placements": summary["outcomes"]["placement_count"],
            },
            "targets": summary["targets"],
            "progress_toward_targets": summary["progress"],
        }

    def _format_detailed_report(
        self,
        grant_id: int,
        summary: dict
    ) -> dict:
        """Format detailed report with student-level data."""
        # Get all enrollments for this grant
        enrollments = self._get_enrollment_details(grant_id)

        # Get all expenses for this grant
        expenses = self._get_expense_details(grant_id)

        return {
            "report_type": "detailed",
            "generated_at": datetime.utcnow().isoformat(),
            "summary": {
                "grant_id": summary["grant_id"],
                "grant_name": summary["grant_name"],
                "grant_number": summary["grant_number"],
                "funder": summary["funder"],
                "status": summary["status"].value if summary["status"] else None,
                "period": summary["period"],
                "enrollment": summary["enrollment"],
                "financials": summary["financials"],
                "outcomes": summary["outcomes"],
                "targets": summary["targets"],
                "progress": summary["progress"],
            },
            "enrollments": enrollments,
            "expenses": expenses,
        }

    def _format_funder_report(
        self,
        grant_id: int,
        summary: dict,
        period_start: Optional[date],
        period_end: Optional[date],
    ) -> dict:
        """Format report for funder submission."""
        # Get grant record for grant number
        grant = self.db.query(Grant).filter(Grant.id == grant_id).first()

        # Filter enrollments by period if specified
        enrollments_query = self.db.query(GrantEnrollment).filter(
            GrantEnrollment.grant_id == grant_id,
            GrantEnrollment.is_deleted == False
        )

        if period_start:
            enrollments_query = enrollments_query.filter(
                GrantEnrollment.enrollment_date >= period_start
            )
        if period_end:
            enrollments_query = enrollments_query.filter(
                GrantEnrollment.enrollment_date <= period_end
            )

        period_enrollments = enrollments_query.count()

        # Period completions
        completions_query = self.db.query(GrantEnrollment).filter(
            GrantEnrollment.grant_id == grant_id,
            GrantEnrollment.is_deleted == False,
            GrantEnrollment.status == GrantEnrollmentStatus.COMPLETED
        )
        if period_start:
            completions_query = completions_query.filter(
                GrantEnrollment.completion_date >= period_start
            )
        if period_end:
            completions_query = completions_query.filter(
                GrantEnrollment.completion_date <= period_end
            )
        period_completions = completions_query.count()

        # Period expenses
        expenses_query = self.db.query(Expense).filter(
            Expense.grant_id == grant_id,
            Expense.is_deleted == False
        )
        if period_start:
            expenses_query = expenses_query.filter(
                Expense.purchased_at >= period_start
            )
        if period_end:
            expenses_query = expenses_query.filter(
                Expense.purchased_at <= period_end
            )

        from sqlalchemy import func
        period_expenses = expenses_query.with_entities(
            func.coalesce(func.sum(Expense.total_price), Decimal("0"))
        ).scalar() or Decimal("0")

        return {
            "report_type": "funder",
            "report_period": {
                "start": period_start.isoformat() if period_start else None,
                "end": period_end.isoformat() if period_end else None,
            },
            "generated_at": datetime.utcnow().isoformat(),
            "grant_number": grant.grant_number if grant else None,
            "grant_name": summary["grant_name"],
            "grantee": "IBEW Local 46",
            "program_name": "Inside to Pre-Apprenticeship (IP2A)",
            "cumulative_metrics": {
                "participants_served": summary["enrollment"]["total_enrolled"],
                "participants_completed": summary["enrollment"]["completed"],
                "job_placements": summary["outcomes"]["placement_count"],
                "retention_rate": f"{summary['enrollment']['retention_rate']}%",
                "funds_expended": summary["financials"]["total_spent"],
                "funds_remaining": summary["financials"]["remaining"],
            },
            "period_metrics": {
                "new_enrollments": period_enrollments,
                "completions": period_completions,
                "funds_expended": float(period_expenses),
            },
            "narrative": self._generate_narrative(summary),
        }

    def _get_enrollment_details(self, grant_id: int) -> list[dict]:
        """Get detailed enrollment records."""
        enrollments = self.db.query(GrantEnrollment).options(
            joinedload(GrantEnrollment.student).joinedload(Student.member)
        ).filter(
            GrantEnrollment.grant_id == grant_id,
            GrantEnrollment.is_deleted == False
        ).order_by(GrantEnrollment.enrollment_date.desc()).all()

        return [
            {
                "id": e.id,
                "student_id": e.student_id,
                "student_name": (
                    f"{e.student.member.first_name} {e.student.member.last_name}"
                    if e.student and e.student.member else "Unknown"
                ),
                "student_number": e.student.student_number if e.student else None,
                "enrollment_date": e.enrollment_date.isoformat() if e.enrollment_date else None,
                "status": e.status.value if e.status else None,
                "completion_date": e.completion_date.isoformat() if e.completion_date else None,
                "outcome": e.outcome.value if e.outcome else None,
                "outcome_date": e.outcome_date.isoformat() if e.outcome_date else None,
                "placement_employer": e.placement_employer,
                "placement_date": e.placement_date.isoformat() if e.placement_date else None,
                "placement_wage": e.placement_wage,
                "placement_job_title": e.placement_job_title,
            }
            for e in enrollments
        ]

    def _get_expense_details(self, grant_id: int) -> list[dict]:
        """Get detailed expense records."""
        expenses = self.db.query(Expense).filter(
            Expense.grant_id == grant_id,
            Expense.is_deleted == False
        ).order_by(Expense.purchased_at.desc()).all()

        return [
            {
                "id": e.id,
                "item": e.item,
                "description": e.description,
                "category": e.category,
                "vendor": e.vendor,
                "purchased_at": e.purchased_at.isoformat() if e.purchased_at else None,
                "quantity": e.quantity,
                "unit_price": float(e.unit_price) if e.unit_price else 0,
                "total_price": float(e.total_price) if e.total_price else 0,
            }
            for e in expenses
        ]

    def _generate_narrative(self, summary: dict) -> str:
        """Generate a brief narrative for funder report."""
        enrollment = summary["enrollment"]
        outcomes = summary["outcomes"]
        financials = summary["financials"]
        progress = summary["progress"]

        parts = []

        # Enrollment narrative
        parts.append(
            f"The program has enrolled {enrollment['total_enrolled']} participants, "
            f"with {enrollment['currently_active']} currently active and "
            f"{enrollment['completed']} having completed the program."
        )

        # Retention
        if enrollment["retention_rate"] >= 80:
            parts.append(
                f"The program maintains a strong retention rate of {enrollment['retention_rate']}%."
            )
        else:
            parts.append(
                f"The current retention rate is {enrollment['retention_rate']}%. "
                "Staff are implementing strategies to improve participant engagement."
            )

        # Outcomes
        if outcomes["placement_count"] > 0:
            parts.append(
                f"{outcomes['placement_count']} participants have achieved positive "
                "employment outcomes (apprenticeship entry or direct employment)."
            )

        # Budget
        parts.append(
            f"Budget utilization is at {financials['utilization_rate']}%, "
            f"with ${financials['remaining']:,.2f} remaining."
        )

        # Progress toward targets
        if progress:
            targets_met = []
            targets_behind = []
            for metric, pct in progress.items():
                if pct >= 100:
                    targets_met.append(metric)
                elif pct < 75:
                    targets_behind.append(metric)

            if targets_met:
                parts.append(
                    f"Targets met or exceeded: {', '.join(targets_met)}."
                )
            if targets_behind:
                parts.append(
                    f"Targets requiring attention: {', '.join(targets_behind)}."
                )

        return " ".join(parts)

    def generate_excel_report(self, grant_id: int) -> Optional[BytesIO]:
        """Generate Excel workbook with grant data."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return None

        summary = self.metrics.get_grant_summary(grant_id)
        if not summary:
            return None

        enrollments = self._get_enrollment_details(grant_id)
        expenses = self._get_expense_details(grant_id)

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Summary data
        ws_summary["A1"] = "Grant Summary Report"
        ws_summary["A1"].font = Font(bold=True, size=14)

        summary_data = [
            ("Grant Name", summary["grant_name"]),
            ("Grant Number", summary.get("grant_number", "N/A")),
            ("Funder", summary.get("funder", "N/A")),
            ("Status", summary["status"].value if summary["status"] else "N/A"),
            ("Start Date", str(summary["period"]["start"])),
            ("End Date", str(summary["period"]["end"]) if summary["period"]["end"] else "N/A"),
            ("", ""),
            ("Enrollment Metrics", ""),
            ("Total Enrolled", summary["enrollment"]["total_enrolled"]),
            ("Currently Active", summary["enrollment"]["currently_active"]),
            ("Completed", summary["enrollment"]["completed"]),
            ("Retention Rate", f"{summary['enrollment']['retention_rate']}%"),
            ("", ""),
            ("Financial Metrics", ""),
            ("Total Budget", f"${summary['financials']['total_budget']:,.2f}"),
            ("Total Spent", f"${summary['financials']['total_spent']:,.2f}"),
            ("Remaining", f"${summary['financials']['remaining']:,.2f}"),
            ("Utilization Rate", f"{summary['financials']['utilization_rate']}%"),
            ("", ""),
            ("Outcomes", ""),
            ("Placements", summary["outcomes"]["placement_count"]),
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=3):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)

        # Enrollments sheet
        ws_enroll = wb.create_sheet("Enrollments")
        enroll_headers = [
            "Student Name", "Student #", "Enrollment Date", "Status",
            "Completion Date", "Outcome", "Placement Employer", "Placement Date"
        ]
        for col_idx, header in enumerate(enroll_headers, start=1):
            cell = ws_enroll.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, e in enumerate(enrollments, start=2):
            ws_enroll.cell(row=row_idx, column=1, value=e["student_name"])
            ws_enroll.cell(row=row_idx, column=2, value=e["student_number"])
            ws_enroll.cell(row=row_idx, column=3, value=e["enrollment_date"])
            ws_enroll.cell(row=row_idx, column=4, value=e["status"])
            ws_enroll.cell(row=row_idx, column=5, value=e["completion_date"])
            ws_enroll.cell(row=row_idx, column=6, value=e["outcome"])
            ws_enroll.cell(row=row_idx, column=7, value=e["placement_employer"])
            ws_enroll.cell(row=row_idx, column=8, value=e["placement_date"])

        # Expenses sheet
        ws_expense = wb.create_sheet("Expenses")
        expense_headers = [
            "Date", "Item", "Category", "Vendor", "Quantity", "Unit Price", "Total"
        ]
        for col_idx, header in enumerate(expense_headers, start=1):
            cell = ws_expense.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, exp in enumerate(expenses, start=2):
            ws_expense.cell(row=row_idx, column=1, value=exp["purchased_at"])
            ws_expense.cell(row=row_idx, column=2, value=exp["item"])
            ws_expense.cell(row=row_idx, column=3, value=exp["category"])
            ws_expense.cell(row=row_idx, column=4, value=exp["vendor"])
            ws_expense.cell(row=row_idx, column=5, value=exp["quantity"])
            ws_expense.cell(row=row_idx, column=6, value=exp["unit_price"])
            ws_expense.cell(row=row_idx, column=7, value=exp["total_price"])

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
