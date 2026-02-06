"""
Referral Report Service for Phase 7 reports.

Created: February 5, 2026 (Week 33A)
Updated: February 5, 2026 (Week 33B - Dispatch reports added)
Phase 7 - Referral & Dispatch System

Generates PDF and Excel reports for out-of-work lists, dispatch logs,
and other referral/dispatch operational reports.
"""

from datetime import datetime, date, timedelta, time
from typing import Optional
from io import BytesIO

from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader

from src.models.book_registration import BookRegistration
from src.models.referral_book import ReferralBook
from src.models.member import Member
from src.models.dispatch import Dispatch
from src.models.labor_request import LaborRequest
from src.models.job_bid import JobBid
from src.db.enums import (
    RegistrationStatus,
    DispatchStatus,
    LaborRequestStatus,
    BidStatus,
    OrganizationType,
    TermReason,
)


# Lazy import for WeasyPrint (requires system libraries)
_weasyprint_available = None


def _get_weasyprint():
    """Lazy load WeasyPrint - only imported when PDF generation is needed."""
    global _weasyprint_available
    if _weasyprint_available is None:
        try:
            from weasyprint import HTML

            _weasyprint_available = HTML
        except (ImportError, OSError):
            _weasyprint_available = False
    return _weasyprint_available


# Morning processing order from Business Rule #2
MORNING_PROCESSING_ORDER = [
    # 8:30 AM - Inside Wireperson
    "WIRE SEATTLE",
    "WIRE BREMERTON",
    "WIRE PT ANGELES",
    # 9:00 AM - Other classifications
    "SOUND & COMM",
    "MARINE",
    "STOCKMAN",
    "LT FXT MAINT",
    "RESIDENTIAL",
    # 9:30 AM - Supplemental
    "TRADESHOW",
    # Unscheduled
    "TECHNICIAN",
    "UTILITY WORKER",
    "TERO APPR WIRE",
]


class ReferralReportService:
    """Generates referral and dispatch reports (PDF + Excel).

    All reports follow these principles:
    - Data assembly happens in service methods (not templates)
    - PDF rendering uses WeasyPrint with Jinja2 HTML templates
    - Excel rendering uses openpyxl with consistent formatting
    - All report generation is audit-logged
    - Union branding (Local 46 header, generation timestamp, user) on all PDFs
    """

    def __init__(self, db: Session):
        self.db = db
        self.template_env = Environment(
            loader=FileSystemLoader("src/templates/reports/referral")
        )

    # --- Report 1: Out-of-Work List (by Book) ---

    def get_out_of_work_by_book(self, book_id: int, tier: Optional[int] = None) -> dict:
        """Assemble data for a single book's out-of-work list.

        Returns: {
            'book': ReferralBook object,
            'registrations': [sorted by APN ascending],
            'total_count': int,
            'tier_counts': {1: n, 2: n, 3: n, 4: n},
            'generated_at': datetime,
        }

        Sort order: APN ascending (DECIMAL sort - this is dispatch priority order).
        If tier is specified, filter to that tier only.
        """
        book = self.db.query(ReferralBook).filter(ReferralBook.id == book_id).first()
        if not book:
            return None

        query = self.db.query(BookRegistration).filter(
            BookRegistration.book_id == book_id,
            BookRegistration.status == RegistrationStatus.REGISTERED,
        )

        # Note: tier filter not implemented - tier is per-book not per-registration
        # The book.book_number represents the tier (1=local, 2=travelers, etc.)

        # Sort by APN (Decimal) - critical for dispatch order
        registrations = query.order_by(BookRegistration.registration_number.asc()).all()

        # Count by tier using the book's tier
        book_tier = book.book_number or 1
        tier_counts = {1: 0, 2: 0, 3: 0, 4: 0}
        tier_counts[book_tier] = len(registrations)

        return {
            "book": book,
            "registrations": registrations,
            "total_count": len(registrations),
            "tier_counts": tier_counts,
            "generated_at": datetime.now(),
        }

    def render_out_of_work_by_book_pdf(
        self, book_id: int, tier: Optional[int] = None
    ) -> BytesIO:
        """Render PDF for single book out-of-work list."""
        data = self.get_out_of_work_by_book(book_id, tier)
        if not data:
            return None

        return self._render_pdf("out_of_work_by_book.html", data)

    def render_out_of_work_by_book_excel(
        self, book_id: int, tier: Optional[int] = None
    ) -> BytesIO:
        """Render Excel for single book out-of-work list."""
        data = self.get_out_of_work_by_book(book_id, tier)
        if not data:
            return None

        headers = [
            "#",
            "APN",
            "Member Name",
            "Card #",
            "Tier",
            "Registered",
            "Last Re-Sign",
            "Check Marks",
            "Status",
        ]

        rows = []
        for i, reg in enumerate(data["registrations"], 1):
            member = reg.member
            rows.append(
                [
                    i,
                    f"{reg.registration_number:.2f}",
                    f"{member.last_name}, {member.first_name}" if member else "Unknown",
                    member.member_number if member else "",
                    data["book"].book_number or 1,
                    reg.registration_date.strftime("%m/%d/%Y")
                    if reg.registration_date
                    else "",
                    reg.last_re_sign_date.strftime("%m/%d/%Y")
                    if reg.last_re_sign_date
                    else "Never",
                    reg.check_marks,
                    reg.status.value if reg.status else "",
                ]
            )

        title = f"Out-of-Work List - {data['book'].name}"
        return self._render_excel(headers, rows, "Out-of-Work", title)

    # --- Report 2: Out-of-Work List (All Books) ---

    def get_out_of_work_all_books(self) -> dict:
        """Assemble data for combined out-of-work list across all books.

        Returns: {
            'books': [{
                'book': ReferralBook,
                'registrations': [sorted by APN],
                'count': int,
            }],
            'total_across_all': int,
            'generated_at': datetime,
        }

        Books sorted by morning processing order (Rule #2).
        """
        books = (
            self.db.query(ReferralBook).filter(ReferralBook.is_active.is_(True)).all()
        )

        # Sort books by morning processing order
        def get_book_order(book):
            name = book.name.upper() if book.name else ""
            try:
                return MORNING_PROCESSING_ORDER.index(name)
            except ValueError:
                return len(MORNING_PROCESSING_ORDER)

        books = sorted(books, key=get_book_order)

        result = []
        total = 0
        for book in books:
            registrations = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book.id,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .order_by(BookRegistration.registration_number.asc())
                .all()
            )
            result.append(
                {
                    "book": book,
                    "registrations": registrations,
                    "count": len(registrations),
                }
            )
            total += len(registrations)

        return {
            "books": result,
            "total_across_all": total,
            "generated_at": datetime.now(),
        }

    def render_out_of_work_all_books_pdf(self) -> BytesIO:
        """Render PDF for all-books out-of-work list."""
        data = self.get_out_of_work_all_books()
        return self._render_pdf("out_of_work_all_books.html", data)

    def render_out_of_work_all_books_excel(self) -> BytesIO:
        """Render Excel for all-books out-of-work list."""
        data = self.get_out_of_work_all_books()

        wb = Workbook()
        wb.remove(wb.active)

        for book_data in data["books"]:
            book = book_data["book"]
            ws = wb.create_sheet(title=book.name[:31] if book.name else "Unknown")

            # Title
            ws.merge_cells("A1:I1")
            ws["A1"] = f"Out-of-Work List - {book.name}"
            ws["A1"].font = Font(bold=True, size=14)

            # Headers
            headers = [
                "#",
                "APN",
                "Name",
                "Card #",
                "Tier",
                "Registered",
                "Re-Sign",
                "Checks",
                "Status",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="003366", end_color="003366", fill_type="solid"
                )
                cell.font = Font(bold=True, color="FFFFFF")

            # Data
            for row_idx, reg in enumerate(book_data["registrations"], 4):
                member = reg.member
                ws.cell(row=row_idx, column=1, value=row_idx - 3)
                ws.cell(row=row_idx, column=2, value=float(reg.registration_number))
                ws.cell(
                    row=row_idx,
                    column=3,
                    value=f"{member.last_name}, {member.first_name}" if member else "",
                )
                ws.cell(
                    row=row_idx, column=4, value=member.member_number if member else ""
                )
                ws.cell(row=row_idx, column=5, value=book.book_number or 1)
                ws.cell(
                    row=row_idx,
                    column=6,
                    value=reg.registration_date.strftime("%m/%d/%Y")
                    if reg.registration_date
                    else "",
                )
                ws.cell(
                    row=row_idx,
                    column=7,
                    value=reg.last_re_sign_date.strftime("%m/%d/%Y")
                    if reg.last_re_sign_date
                    else "",
                )
                ws.cell(row=row_idx, column=8, value=reg.check_marks)
                ws.cell(
                    row=row_idx, column=9, value=reg.status.value if reg.status else ""
                )

            # Auto-fit columns
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 12

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # --- Report 3: Out-of-Work Summary ---

    def get_out_of_work_summary(self) -> dict:
        """Assemble summary counts per book per tier.

        Returns: {
            'summary': [{
                'book_name': str,
                'book_id': int,
                'tier_1': int,
                'tier_2': int,
                'tier_3': int,
                'tier_4': int,
                'total': int,
            }],
            'grand_total': int,
            'generated_at': datetime,
        }
        """
        books = (
            self.db.query(ReferralBook).filter(ReferralBook.is_active.is_(True)).all()
        )

        summary = []
        grand_total = 0

        for book in sorted(books, key=lambda b: b.name or ""):
            registrations = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book.id,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .all()
            )

            # Tier is per-book (book_number), not per-registration
            book_tier = book.book_number or 1
            tier_counts = {1: 0, 2: 0, 3: 0, 4: 0}
            tier_counts[book_tier] = len(registrations)

            total = len(registrations)
            grand_total += total

            summary.append(
                {
                    "book_name": book.name,
                    "book_id": book.id,
                    "tier_1": tier_counts[1],
                    "tier_2": tier_counts[2],
                    "tier_3": tier_counts[3],
                    "tier_4": tier_counts[4],
                    "total": total,
                }
            )

        return {
            "summary": summary,
            "grand_total": grand_total,
            "generated_at": datetime.now(),
        }

    def render_out_of_work_summary_pdf(self) -> BytesIO:
        """Render PDF summary - compact, fits on one page."""
        data = self.get_out_of_work_summary()
        return self._render_pdf("out_of_work_summary.html", data)

    # --- Report 4: Active Registrations by Member ---

    def get_member_registrations(self, member_id: int) -> dict:
        """Assemble all active registrations for a specific member.

        Returns: {
            'member': Member object,
            'registrations': [{
                'book': ReferralBook,
                'tier': int,
                'apn': Decimal,
                'registered_date': date,
                'last_re_sign': date,
                're_sign_due': date,
                'check_marks': int,
                'status': str,
            }],
            'total_books': int,
            'generated_at': datetime,
        }
        """
        member = self.db.query(Member).filter(Member.id == member_id).first()
        if not member:
            return None

        registrations = (
            self.db.query(BookRegistration)
            .filter(
                BookRegistration.member_id == member_id,
                BookRegistration.status == RegistrationStatus.REGISTERED,
            )
            .all()
        )

        result_regs = []
        for reg in registrations:
            # Calculate re-sign due date
            base_date = reg.last_re_sign_date or reg.registration_date
            re_sign_due = None
            if base_date:
                if isinstance(base_date, datetime):
                    base_date = base_date.date()
                re_sign_due = base_date + timedelta(days=30)

            result_regs.append(
                {
                    "book": reg.book,
                    "tier": reg.book.book_number if reg.book else 1,
                    "apn": reg.registration_number,
                    "registered_date": reg.registration_date.date()
                    if isinstance(reg.registration_date, datetime)
                    else reg.registration_date,
                    "last_re_sign": reg.last_re_sign_date.date()
                    if isinstance(reg.last_re_sign_date, datetime)
                    else reg.last_re_sign_date,
                    "re_sign_due": re_sign_due,
                    "check_marks": reg.check_marks,
                    "status": reg.status.value if reg.status else "",
                }
            )

        return {
            "member": member,
            "registrations": result_regs,
            "total_books": len(result_regs),
            "generated_at": datetime.now(),
        }

    def render_member_registrations_pdf(self, member_id: int) -> BytesIO:
        """Render PDF for single member's registration overview."""
        data = self.get_member_registrations(member_id)
        if not data:
            return None
        return self._render_pdf("member_registrations.html", data)

    # --- Shared Infrastructure ---

    def _render_pdf(self, template_name: str, context: dict) -> BytesIO:
        """Render HTML template to PDF via WeasyPrint.

        All PDFs include:
        - IBEW Local 46 header
        - Report title and generation timestamp
        - Page numbers in footer
        """
        HTML = _get_weasyprint()
        if not HTML:
            # Fallback: return empty buffer if WeasyPrint not available
            return BytesIO()

        try:
            template = self.template_env.get_template(template_name)
            html_content = template.render(**context)

            buffer = BytesIO()
            HTML(string=html_content).write_pdf(buffer)
            buffer.seek(0)
            return buffer
        except Exception:
            # Return empty buffer on error
            return BytesIO()

    def _render_excel(
        self, headers: list, rows: list, sheet_name: str, title: str
    ) -> BytesIO:
        """Render data to Excel via openpyxl.

        All Excel files include:
        - Title row (merged, bold)
        - Generation timestamp
        - Auto-column-width
        - Freeze panes (header row frozen)
        - Table formatting with alternating row colors
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Title row
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)

        # Timestamp
        ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")

        # Headers
        header_fill = PatternFill(
            start_color="003366", end_color="003366", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data rows
        alt_fill = PatternFill(
            start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"
        )
        for row_idx, row_data in enumerate(rows, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            max_length = (
                max(
                    len(str(headers[col - 1])),
                    *[len(str(row[col - 1])) for row in rows if col - 1 < len(row)],
                )
                if rows
                else len(str(headers[col - 1]))
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 30)

        # Freeze header row
        ws.freeze_panes = "A5"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # --- Report 5: Daily Dispatch Log ---

    def get_daily_dispatch_log(
        self, start_date: date, end_date: Optional[date] = None
    ) -> dict:
        """Assemble all dispatches within a date range.

        Args:
            start_date: Start of date range (inclusive)
            end_date: End of date range (inclusive). Defaults to start_date (single day).

        Returns: {
            'dispatches': list of dispatch dicts,
            'date_range': {'start': date, 'end': date},
            'total_dispatches': int,
            'by_book_count': {book_name: count},
            'by_employer_count': {employer_name: count},
            'generated_at': datetime,
        }
        """
        if end_date is None:
            end_date = start_date

        # Query dispatches in date range
        query = (
            self.db.query(Dispatch)
            .filter(
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
            )
            .order_by(Dispatch.dispatch_date.desc())
        )

        dispatches = query.all()

        # Build result
        result_dispatches = []
        by_book_count = {}
        by_employer_count = {}

        for d in dispatches:
            member = d.member
            employer = d.employer
            book_name = d.book_code or "Unknown"
            employer_name = employer.name if employer else "Unknown"

            # Count by book
            by_book_count[book_name] = by_book_count.get(book_name, 0) + 1

            # Count by employer
            by_employer_count[employer_name] = (
                by_employer_count.get(employer_name, 0) + 1
            )

            result_dispatches.append(
                {
                    "dispatch_id": d.id,
                    "dispatch_date": d.dispatch_date,
                    "member_name": f"{member.last_name}, {member.first_name}"
                    if member
                    else "Unknown",
                    "member_card": member.member_number if member else "",
                    "book_name": book_name,
                    "employer_name": employer_name,
                    "job_site": d.worksite or "",
                    "labor_request_id": d.labor_request_id,
                    "dispatch_type": d.dispatch_type.value
                    if d.dispatch_type
                    else "normal",
                    "status": d.dispatch_status.value if d.dispatch_status else "",
                    "generates_checkmark": not d.is_short_call,  # Short calls don't generate check marks
                }
            )

        return {
            "dispatches": result_dispatches,
            "date_range": {"start": start_date, "end": end_date},
            "total_dispatches": len(dispatches),
            "by_book_count": by_book_count,
            "by_employer_count": by_employer_count,
            "generated_at": datetime.now(),
        }

    def render_daily_dispatch_log_pdf(
        self, start_date: date, end_date: Optional[date] = None
    ) -> BytesIO:
        """Render PDF for daily dispatch log."""
        data = self.get_daily_dispatch_log(start_date, end_date)
        return self._render_pdf("daily_dispatch_log.html", data)

    def render_daily_dispatch_log_excel(
        self, start_date: date, end_date: Optional[date] = None
    ) -> BytesIO:
        """Render Excel for daily dispatch log."""
        data = self.get_daily_dispatch_log(start_date, end_date)

        headers = [
            "Dispatch #",
            "Date/Time",
            "Member",
            "Card #",
            "Book",
            "Employer",
            "Site",
            "Type",
            "Checkmark?",
            "Status",
        ]

        rows = []
        for d in data["dispatches"]:
            rows.append(
                [
                    d["dispatch_id"],
                    d["dispatch_date"].strftime("%m/%d/%Y %I:%M %p")
                    if d["dispatch_date"]
                    else "",
                    d["member_name"],
                    d["member_card"],
                    d["book_name"],
                    d["employer_name"],
                    d["job_site"],
                    d["dispatch_type"],
                    "Yes" if d["generates_checkmark"] else "No",
                    d["status"],
                ]
            )

        date_range = data["date_range"]
        if date_range["start"] == date_range["end"]:
            title = f"Daily Dispatch Log - {date_range['start'].strftime('%B %d, %Y')}"
        else:
            title = f"Dispatch Log - {date_range['start'].strftime('%m/%d/%Y')} to {date_range['end'].strftime('%m/%d/%Y')}"

        return self._render_excel(headers, rows, "Dispatch Log", title)

    # --- Report 6: Dispatch History by Member ---

    def get_member_dispatch_history(self, member_id: int) -> Optional[dict]:
        """Assemble complete dispatch history for one member.

        Returns: {
            'member': Member object,
            'dispatches': list of dispatch dicts,
            'total_dispatches': int,
            'total_employers': int,
            'average_duration': float (days),
            'check_marks_total': int,
            'generated_at': datetime,
        }
        """
        member = self.db.query(Member).filter(Member.id == member_id).first()
        if not member:
            return None

        dispatches = (
            self.db.query(Dispatch)
            .filter(Dispatch.member_id == member_id)
            .order_by(Dispatch.dispatch_date.desc())
            .all()
        )

        result_dispatches = []
        employers_seen = set()
        total_days = 0
        completed_count = 0
        check_marks = 0

        for d in dispatches:
            employer = d.employer
            employer_name = employer.name if employer else "Unknown"
            employers_seen.add(employer_name)

            # Calculate duration if completed
            duration_days = None
            if d.days_worked:
                duration_days = d.days_worked
                total_days += duration_days
                completed_count += 1
            elif d.term_date and d.start_date:
                duration_days = (d.term_date - d.start_date).days + 1
                total_days += duration_days
                completed_count += 1

            # Determine outcome
            outcome = None
            if d.term_reason:
                outcome = d.term_reason.value
            elif d.dispatch_status:
                outcome = d.dispatch_status.value

            # Count check marks (non-short-call dispatches)
            if not d.is_short_call and d.is_completed:
                check_marks += 1

            result_dispatches.append(
                {
                    "dispatch_date": d.dispatch_date,
                    "book_name": d.book_code or "Unknown",
                    "employer_name": employer_name,
                    "job_site": d.worksite or "",
                    "dispatch_type": d.dispatch_type.value
                    if d.dispatch_type
                    else "normal",
                    "duration_days": duration_days,
                    "outcome": outcome,
                    "was_short_call": d.is_short_call,
                    "generated_checkmark": not d.is_short_call and d.is_completed,
                }
            )

        avg_duration = total_days / completed_count if completed_count > 0 else 0

        return {
            "member": member,
            "dispatches": result_dispatches,
            "total_dispatches": len(dispatches),
            "total_employers": len(employers_seen),
            "average_duration": round(avg_duration, 1),
            "check_marks_total": check_marks,
            "generated_at": datetime.now(),
        }

    def render_member_dispatch_history_pdf(self, member_id: int) -> Optional[BytesIO]:
        """Render PDF for member dispatch history."""
        data = self.get_member_dispatch_history(member_id)
        if not data:
            return None
        return self._render_pdf("member_dispatch_history.html", data)

    # --- Report 7: Labor Request Status ---

    def get_labor_request_status(
        self,
        status_filter: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> dict:
        """Assemble labor request status report.

        Args:
            status_filter: OPEN, FILLED, CANCELLED, EXPIRED, or None (all)
            start_date/end_date: Filter by request creation date range

        Returns: {
            'requests': list of request dicts,
            'status_counts': {'OPEN': n, 'FILLED': n, ...},
            'total_workers_requested': int,
            'total_workers_filled': int,
            'fill_rate': float (percentage),
            'generated_at': datetime,
        }
        """
        query = self.db.query(LaborRequest)

        # Apply status filter
        if status_filter:
            try:
                status_enum = LaborRequestStatus(status_filter.lower())
                query = query.filter(LaborRequest.status == status_enum)
            except ValueError:
                pass

        # Apply date range filter
        if start_date:
            query = query.filter(func.date(LaborRequest.request_date) >= start_date)
        if end_date:
            query = query.filter(func.date(LaborRequest.request_date) <= end_date)

        # Order by date descending
        query = query.order_by(LaborRequest.request_date.desc())

        requests = query.all()

        # Build result
        result_requests = []
        status_counts = {
            "open": 0,
            "filled": 0,
            "cancelled": 0,
            "expired": 0,
            "partially_filled": 0,
        }
        total_workers_requested = 0
        total_workers_filled = 0

        for r in requests:
            status_val = r.status.value if r.status else "open"
            if status_val in status_counts:
                status_counts[status_val] += 1

            total_workers_requested += r.workers_requested
            total_workers_filled += r.workers_dispatched

            # Get filled_at from first dispatch if available
            filled_at = None
            if r.dispatches and len(r.dispatches) > 0:
                first_dispatch = min(r.dispatches, key=lambda x: x.dispatch_date)
                filled_at = first_dispatch.dispatch_date

            result_requests.append(
                {
                    "request_id": r.id,
                    "employer_name": r.employer_name
                    or (r.employer.name if r.employer else "Unknown"),
                    "book_name": r.book.name if r.book else "Unknown",
                    "workers_requested": r.workers_requested,
                    "workers_filled": r.workers_dispatched,
                    "status": status_val,
                    "created_at": r.request_date,
                    "filled_at": filled_at,
                    "agreement_type": r.agreement_type.value
                    if r.agreement_type
                    else "standard",
                    "is_short_call": r.is_short_call,
                    "is_by_name": r.is_foreperson_by_name,
                }
            )

        fill_rate = (
            (total_workers_filled / total_workers_requested * 100)
            if total_workers_requested > 0
            else 0
        )

        return {
            "requests": result_requests,
            "status_counts": status_counts,
            "total_workers_requested": total_workers_requested,
            "total_workers_filled": total_workers_filled,
            "fill_rate": round(fill_rate, 1),
            "generated_at": datetime.now(),
        }

    def render_labor_request_status_pdf(
        self,
        status_filter: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> BytesIO:
        """Render PDF for labor request status."""
        data = self.get_labor_request_status(status_filter, start_date, end_date)
        return self._render_pdf("labor_request_status.html", data)

    def render_labor_request_status_excel(
        self,
        status_filter: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> BytesIO:
        """Render Excel for labor request status."""
        data = self.get_labor_request_status(status_filter, start_date, end_date)

        headers = [
            "Request #",
            "Employer",
            "Book",
            "Workers Req",
            "Workers Filled",
            "Status",
            "Agreement",
            "Type",
            "Created",
            "Filled",
        ]

        rows = []
        for r in data["requests"]:
            req_type = []
            if r["is_short_call"]:
                req_type.append("Short Call")
            if r["is_by_name"]:
                req_type.append("By Name")
            type_str = ", ".join(req_type) if req_type else "Regular"

            rows.append(
                [
                    r["request_id"],
                    r["employer_name"],
                    r["book_name"],
                    r["workers_requested"],
                    r["workers_filled"],
                    r["status"].upper(),
                    r["agreement_type"],
                    type_str,
                    r["created_at"].strftime("%m/%d/%Y") if r["created_at"] else "",
                    r["filled_at"].strftime("%m/%d/%Y") if r["filled_at"] else "",
                ]
            )

        title = "Labor Request Status Report"
        return self._render_excel(headers, rows, "Request Status", title)

    # --- Report 8: Morning Referral Sheet ---

    def _previous_working_day(self, target: date) -> date:
        """Return the previous working day (skip weekends, no holiday awareness yet).

        Monday → Friday, Sunday → Friday, Saturday → Friday
        Tuesday-Friday → previous day

        TODO: Add holiday calendar integration for full Rule #1 compliance.
        """
        delta = 1
        if target.weekday() == 0:  # Monday
            delta = 3
        elif target.weekday() == 6:  # Sunday
            delta = 2
        return target - timedelta(days=delta)

    def get_morning_referral_sheet(self, target_date: Optional[date] = None) -> dict:
        """Assemble the morning referral processing sheet.

        THIS IS THE MOST CRITICAL DAILY REPORT.

        Logic:
        1. Get all OPEN labor requests received before 3:00 PM the previous working day
        2. Group by book, sorted in morning processing order (Rule #2)
        3. For each request, show the next N members on the book (by APN)
           where N = workers_requested
        4. Flag any members with 2 check marks (at-limit warning)
        5. Flag any web bids received during the 5:30 PM - 7:00 AM window

        Args:
            target_date: The morning being processed. Defaults to today.
                         Uses previous working day's 3 PM cutoff for request inclusion.

        Returns: dict with processing_groups, total_requests, etc.
        """
        if target_date is None:
            target_date = date.today()

        # Calculate cutoff: 3:00 PM previous working day
        prev_day = self._previous_working_day(target_date)
        cutoff_datetime = datetime.combine(prev_day, time(15, 0))

        # Get OPEN requests received before cutoff
        open_requests = (
            self.db.query(LaborRequest)
            .filter(
                LaborRequest.status == LaborRequestStatus.OPEN,
                LaborRequest.request_date <= cutoff_datetime,
            )
            .order_by(LaborRequest.request_date)
            .all()
        )

        # Group by book
        requests_by_book = {}
        for req in open_requests:
            book_id = req.book_id
            if book_id not in requests_by_book:
                requests_by_book[book_id] = {
                    "book": req.book,
                    "requests": [],
                }
            requests_by_book[book_id]["requests"].append(req)

        # Build processing groups by morning order
        # 8:30 AM: Wire books, 9:00 AM: Other classifications, 9:30 AM: Supplemental
        processing_groups = [
            {"time_slot": "8:30 AM", "books": []},
            {"time_slot": "9:00 AM", "books": []},
            {"time_slot": "9:30 AM", "books": []},
        ]

        # Determine time slot based on book name
        for book_id, book_data in requests_by_book.items():
            book = book_data["book"]
            book_name = book.name.upper() if book and book.name else ""

            # Determine slot
            if "WIRE" in book_name:
                slot_idx = 0
            elif "TRADESHOW" in book_name or "TERO" in book_name:
                slot_idx = 2
            else:
                slot_idx = 1

            # Get queue for this book
            queue = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book_id,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .order_by(BookRegistration.registration_number)
                .all()
            )

            # Process each request
            processed_requests = []
            for req in book_data["requests"]:
                workers_needed = req.workers_remaining

                # Get web bids for this request
                web_bids = (
                    self.db.query(JobBid)
                    .filter(
                        JobBid.labor_request_id == req.id,
                        JobBid.bid_status == BidStatus.PENDING,
                    )
                    .all()
                )
                bid_member_ids = {b.member_id for b in web_bids}

                # Get next N members in queue
                next_in_queue = []
                for i, reg in enumerate(queue[: workers_needed + 5]):  # Get a few extra
                    member = reg.member
                    next_in_queue.append(
                        {
                            "position": i + 1,
                            "member_name": f"{member.last_name}, {member.first_name}"
                            if member
                            else "Unknown",
                            "apn": reg.registration_number,
                            "tier": book.book_number if book else 1,
                            "check_marks": reg.check_marks,
                            "at_check_limit": reg.check_marks >= 2,
                            "has_web_bid": member.id in bid_member_ids
                            if member
                            else False,
                        }
                    )

                processed_requests.append(
                    {
                        "request_id": req.id,
                        "employer_name": req.employer_name
                        or (req.employer.name if req.employer else "Unknown"),
                        "workers_needed": workers_needed,
                        "agreement_type": req.agreement_type.value
                        if req.agreement_type
                        else "standard",
                        "is_short_call": req.is_short_call,
                        "is_by_name": req.is_foreperson_by_name,
                        "web_bids": [
                            {
                                "member_name": f"{b.member.last_name}, {b.member.first_name}"
                                if b.member
                                else "Unknown",
                                "bid_time": b.bid_submitted_at,
                            }
                            for b in web_bids
                        ],
                        "next_in_queue": next_in_queue[
                            : workers_needed + 3
                        ],  # Show a few extra
                    }
                )

            processing_groups[slot_idx]["books"].append(
                {
                    "book_name": book.name if book else "Unknown",
                    "requests": processed_requests,
                }
            )

        # Sort books within each slot by morning processing order
        for group in processing_groups:
            group["books"].sort(
                key=lambda b: MORNING_PROCESSING_ORDER.index(b["book_name"].upper())
                if b["book_name"].upper() in MORNING_PROCESSING_ORDER
                else len(MORNING_PROCESSING_ORDER)
            )

        total_requests = len(open_requests)
        total_workers_needed = sum(r.workers_remaining for r in open_requests)

        return {
            "target_date": target_date,
            "cutoff_datetime": cutoff_datetime,
            "processing_groups": processing_groups,
            "total_requests": total_requests,
            "total_workers_needed": total_workers_needed,
            "generated_at": datetime.now(),
        }

    def render_morning_referral_sheet_pdf(
        self, target_date: Optional[date] = None
    ) -> BytesIO:
        """Render PDF for morning referral sheet.

        Layout: Landscape orientation for wide tables.
        Grouped by processing time slot, then by book.
        Visual flags: yellow highlight for at-limit check marks,
                      blue highlight for web bids.
        """
        data = self.get_morning_referral_sheet(target_date)
        return self._render_pdf("morning_referral_sheet.html", data)

    # --- Report 9: Active Dispatches ---

    def get_active_dispatches(self) -> dict:
        """Assemble currently active (outstanding) dispatches.

        "Active" = dispatched but not yet completed/terminated.

        Returns: {
            'dispatches': list of dispatch dicts,
            'total_active': int,
            'by_employer': {employer_name: count},
            'by_book': {book_name: count},
            'short_call_count': int,
            'long_call_count': int,
            'generated_at': datetime,
        }
        """
        active_statuses = [
            DispatchStatus.DISPATCHED,
            DispatchStatus.CHECKED_IN,
            DispatchStatus.WORKING,
        ]

        dispatches = (
            self.db.query(Dispatch)
            .filter(Dispatch.dispatch_status.in_(active_statuses))
            .order_by(Dispatch.employer_id, Dispatch.dispatch_date)
            .all()
        )

        result_dispatches = []
        by_employer = {}
        by_book = {}
        short_call_count = 0
        long_call_count = 0
        today = date.today()

        for d in dispatches:
            member = d.member
            employer = d.employer
            employer_name = employer.name if employer else "Unknown"
            book_name = d.book_code or "Unknown"

            # Count by employer and book
            by_employer[employer_name] = by_employer.get(employer_name, 0) + 1
            by_book[book_name] = by_book.get(book_name, 0) + 1

            # Calculate days on job
            dispatch_dt = d.dispatch_date
            if isinstance(dispatch_dt, datetime):
                dispatch_date = dispatch_dt.date()
            else:
                dispatch_date = dispatch_dt
            days_on_job = (today - dispatch_date).days + 1 if dispatch_date else 0

            # Short call tracking
            if d.is_short_call:
                short_call_count += 1
                # Calculate days remaining (10 - days worked)
                short_call_days = (
                    d.short_call_days
                    if hasattr(d, "short_call_days") and d.short_call_days
                    else 10
                )
                days_remaining = max(0, short_call_days - days_on_job)
            else:
                long_call_count += 1
                days_remaining = None

            result_dispatches.append(
                {
                    "member_name": f"{member.last_name}, {member.first_name}"
                    if member
                    else "Unknown",
                    "member_card": member.member_number if member else "",
                    "employer_name": employer_name,
                    "book_name": book_name,
                    "dispatch_date": d.dispatch_date,
                    "days_on_job": days_on_job,
                    "is_short_call": d.is_short_call,
                    "short_call_days_remaining": days_remaining,
                    "agreement_type": d.labor_request.agreement_type.value
                    if d.labor_request and d.labor_request.agreement_type
                    else "standard",
                    "status": d.dispatch_status.value if d.dispatch_status else "",
                }
            )

        return {
            "dispatches": result_dispatches,
            "total_active": len(dispatches),
            "by_employer": by_employer,
            "by_book": by_book,
            "short_call_count": short_call_count,
            "long_call_count": long_call_count,
            "generated_at": datetime.now(),
        }

    def render_active_dispatches_pdf(self) -> BytesIO:
        """Render PDF for active dispatches."""
        data = self.get_active_dispatches()
        return self._render_pdf("active_dispatches.html", data)

    def render_active_dispatches_excel(self) -> BytesIO:
        """Render Excel for active dispatches."""
        data = self.get_active_dispatches()

        headers = [
            "Member",
            "Card #",
            "Employer",
            "Book",
            "Dispatched",
            "Days on Job",
            "Short Call?",
            "Days Left",
            "Agreement",
            "Status",
        ]

        rows = []
        for d in data["dispatches"]:
            rows.append(
                [
                    d["member_name"],
                    d["member_card"],
                    d["employer_name"],
                    d["book_name"],
                    d["dispatch_date"].strftime("%m/%d/%Y")
                    if d["dispatch_date"]
                    else "",
                    d["days_on_job"],
                    "Yes" if d["is_short_call"] else "No",
                    d["short_call_days_remaining"]
                    if d["short_call_days_remaining"] is not None
                    else "",
                    d["agreement_type"],
                    d["status"],
                ]
            )

        title = f"Active Dispatches - {date.today().strftime('%B %d, %Y')}"
        return self._render_excel(headers, rows, "Active Dispatches", title)

    # --- Report 10: Employer Active List (P0) ---

    def get_employer_active_list(self, contract_code: Optional[str] = None) -> dict:
        """Assemble list of employers with active labor requests or dispatched workers.

        Args:
            contract_code: Filter to specific contract (e.g., 'WIREPERSON').
                          None = all contracts.

        Returns: dict with employers, totals, generated_at
        """
        from src.models.organization import Organization
        from src.db.enums import OrganizationType

        # Get employers with open requests OR active dispatches
        # Subquery for open requests count
        open_request_counts = (
            self.db.query(
                LaborRequest.employer_id,
                func.count(LaborRequest.id).label("open_count"),
            )
            .filter(LaborRequest.status == LaborRequestStatus.OPEN)
            .group_by(LaborRequest.employer_id)
            .subquery()
        )

        # Subquery for active dispatches count
        active_dispatch_counts = (
            self.db.query(
                Dispatch.employer_id,
                func.count(Dispatch.id).label("active_count"),
            )
            .filter(
                Dispatch.dispatch_status.in_(
                    [DispatchStatus.DISPATCHED, DispatchStatus.CHECKED_IN, DispatchStatus.WORKING]
                )
            )
            .group_by(Dispatch.employer_id)
            .subquery()
        )

        # Subquery for YTD dispatches
        year_start = date(date.today().year, 1, 1)
        ytd_dispatch_counts = (
            self.db.query(
                Dispatch.employer_id,
                func.count(Dispatch.id).label("ytd_count"),
                func.max(Dispatch.dispatch_date).label("last_dispatch"),
            )
            .filter(func.date(Dispatch.dispatch_date) >= year_start)
            .group_by(Dispatch.employer_id)
            .subquery()
        )

        # Main query - all employers with activity
        query = (
            self.db.query(
                Organization,
                open_request_counts.c.open_count,
                active_dispatch_counts.c.active_count,
                ytd_dispatch_counts.c.ytd_count,
                ytd_dispatch_counts.c.last_dispatch,
            )
            .outerjoin(
                open_request_counts,
                Organization.id == open_request_counts.c.employer_id,
            )
            .outerjoin(
                active_dispatch_counts,
                Organization.id == active_dispatch_counts.c.employer_id,
            )
            .outerjoin(
                ytd_dispatch_counts,
                Organization.id == ytd_dispatch_counts.c.employer_id,
            )
            .filter(Organization.org_type == OrganizationType.EMPLOYER)
            .filter(
                (open_request_counts.c.open_count > 0)
                | (active_dispatch_counts.c.active_count > 0)
            )
            .order_by(Organization.name)
        )

        results = query.all()

        employers = []
        total_open_requests = 0
        total_dispatched = 0

        for org, open_count, active_count, ytd_count, last_dispatch in results:
            open_requests = open_count or 0
            workers_dispatched = active_count or 0
            ytd_total = ytd_count or 0

            total_open_requests += open_requests
            total_dispatched += workers_dispatched

            # Contact info
            contact_parts = []
            if org.phone:
                contact_parts.append(org.phone)
            if org.email:
                contact_parts.append(org.email)
            contact_info = " | ".join(contact_parts) if contact_parts else ""

            employers.append(
                {
                    "employer_id": org.id,
                    "employer_name": org.name,
                    "contract_codes": [],  # Would need employer_contracts join
                    "open_requests": open_requests,
                    "workers_currently_dispatched": workers_dispatched,
                    "total_dispatches_ytd": ytd_total,
                    "last_dispatch_date": last_dispatch,
                    "contact_info": contact_info,
                }
            )

        return {
            "employers": employers,
            "total_employers": len(employers),
            "total_open_requests": total_open_requests,
            "total_dispatched_workers": total_dispatched,
            "contract_code_filter": contract_code or "All",
            "generated_at": datetime.now(),
        }

    def render_employer_active_list_pdf(
        self, contract_code: Optional[str] = None
    ) -> BytesIO:
        """Render PDF for employer active list."""
        data = self.get_employer_active_list(contract_code)
        return self._render_pdf("employer_active_list.html", data)

    def render_employer_active_list_excel(
        self, contract_code: Optional[str] = None
    ) -> BytesIO:
        """Render Excel for employer active list."""
        data = self.get_employer_active_list(contract_code)

        headers = [
            "Employer",
            "Open Requests",
            "Workers Dispatched",
            "YTD Total",
            "Last Dispatch",
            "Contact",
        ]

        rows = []
        for e in data["employers"]:
            rows.append(
                [
                    e["employer_name"],
                    e["open_requests"],
                    e["workers_currently_dispatched"],
                    e["total_dispatches_ytd"],
                    e["last_dispatch_date"].strftime("%m/%d/%Y")
                    if e["last_dispatch_date"]
                    else "",
                    e["contact_info"],
                ]
            )

        title = "Employer Active List"
        return self._render_excel(headers, rows, "Active Employers", title)

    # --- Report 11: Employer Dispatch History (P1) ---

    def get_employer_dispatch_history(
        self,
        employer_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> Optional[dict]:
        """Assemble dispatch history for a specific employer.

        Args:
            employer_id: Target employer
            start_date/end_date: Filter range. Defaults to current year.

        Returns: dict with employer, dispatches, stats, or None if not found
        """
        from src.models.organization import Organization

        employer = self.db.query(Organization).filter(Organization.id == employer_id).first()
        if not employer:
            return None

        # Default to current year
        if not start_date:
            start_date = date(date.today().year, 1, 1)
        if not end_date:
            end_date = date.today()

        dispatches = (
            self.db.query(Dispatch)
            .filter(
                Dispatch.employer_id == employer_id,
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
            )
            .order_by(Dispatch.dispatch_date.desc())
            .all()
        )

        result_dispatches = []
        members_seen = set()
        total_days = 0
        completed_count = 0
        short_call_count = 0

        for d in dispatches:
            member = d.member
            member_name = (
                f"{member.last_name}, {member.first_name}" if member else "Unknown"
            )
            member_card = member.member_number if member else ""
            if member:
                members_seen.add(member.id)

            # Calculate duration
            duration_days = None
            if d.days_worked:
                duration_days = d.days_worked
                total_days += duration_days
                completed_count += 1
            elif d.term_date and d.start_date:
                duration_days = (d.term_date - d.start_date).days + 1
                total_days += duration_days
                completed_count += 1

            # Outcome
            outcome = None
            if d.term_reason:
                outcome = d.term_reason.value
            elif d.dispatch_status:
                outcome = d.dispatch_status.value

            if d.is_short_call:
                short_call_count += 1

            result_dispatches.append(
                {
                    "dispatch_date": d.dispatch_date,
                    "member_name": member_name,
                    "member_card": member_card,
                    "book_name": d.book_code or "Unknown",
                    "dispatch_type": d.dispatch_type.value if d.dispatch_type else "normal",
                    "duration_days": duration_days,
                    "outcome": outcome,
                    "was_short_call": d.is_short_call,
                }
            )

        avg_duration = total_days / completed_count if completed_count > 0 else 0
        short_call_pct = (
            (short_call_count / len(dispatches) * 100) if dispatches else 0
        )

        return {
            "employer": employer,
            "contract_codes": [],
            "dispatches": result_dispatches,
            "total_dispatches": len(dispatches),
            "unique_members": len(members_seen),
            "average_duration": round(avg_duration, 1),
            "short_call_percentage": round(short_call_pct, 1),
            "date_range": {"start": start_date, "end": end_date},
            "generated_at": datetime.now(),
        }

    def render_employer_dispatch_history_pdf(
        self,
        employer_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> Optional[BytesIO]:
        """Render PDF for employer dispatch history."""
        data = self.get_employer_dispatch_history(employer_id, start_date, end_date)
        if not data:
            return None
        return self._render_pdf("employer_dispatch_history.html", data)

    def render_employer_dispatch_history_excel(
        self,
        employer_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> Optional[BytesIO]:
        """Render Excel for employer dispatch history."""
        data = self.get_employer_dispatch_history(employer_id, start_date, end_date)
        if not data:
            return None

        headers = [
            "Date",
            "Member",
            "Card #",
            "Book",
            "Type",
            "Duration (days)",
            "Short Call?",
            "Outcome",
        ]

        rows = []
        for d in data["dispatches"]:
            rows.append(
                [
                    d["dispatch_date"].strftime("%m/%d/%Y") if d["dispatch_date"] else "",
                    d["member_name"],
                    d["member_card"],
                    d["book_name"],
                    d["dispatch_type"],
                    d["duration_days"] if d["duration_days"] else "",
                    "Yes" if d["was_short_call"] else "No",
                    d["outcome"] or "",
                ]
            )

        employer_name = data["employer"].name if data["employer"] else "Unknown"
        title = f"Dispatch History - {employer_name}"
        return self._render_excel(headers, rows, "Dispatch History", title)

    # --- Report 12: Registration History (P1, Excel only) ---

    def get_registration_history(
        self,
        book_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        activity_type: Optional[str] = None,
    ) -> dict:
        """Assemble registration activity history.

        Uses the registration_activities table for activity log.

        Args:
            book_id: Filter to specific book. None = all books.
            start_date/end_date: Filter range. Defaults to last 30 days.
            activity_type: Filter by type (REGISTER, RE_SIGN, etc.). None = all.

        Returns: dict with activities, counts, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        # Default to last 30 days
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        query = self.db.query(RegistrationActivity).filter(
            func.date(RegistrationActivity.activity_date) >= start_date,
            func.date(RegistrationActivity.activity_date) <= end_date,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        if activity_type:
            try:
                action_enum = RegistrationAction(activity_type.lower())
                query = query.filter(RegistrationActivity.action == action_enum)
            except ValueError:
                pass

        activities = query.order_by(RegistrationActivity.activity_date.desc()).all()

        result_activities = []
        activity_counts = {}

        for act in activities:
            member = act.member
            member_name = (
                f"{member.last_name}, {member.first_name}" if member else "Unknown"
            )
            member_card = member.member_number if member else ""
            book = act.book
            book_name = book.name if book else "Unknown"
            tier = book.book_number if book else 1

            action_val = act.action.value
            activity_counts[action_val] = activity_counts.get(action_val, 0) + 1

            # Get processor name
            performed_by = act.processor or (
                f"{act.performed_by.email}" if act.performed_by else "System"
            )

            result_activities.append(
                {
                    "activity_date": act.activity_date,
                    "member_name": member_name,
                    "member_card": member_card,
                    "book_name": book_name,
                    "tier": tier,
                    "activity_type": action_val,
                    "reason": act.reason or "",
                    "previous_apn": act.previous_position,
                    "new_apn": act.new_position,
                    "performed_by": performed_by,
                }
            )

        # Build book filter display
        book_filter = "All Books"
        if book_id:
            book_obj = self.db.query(ReferralBook).filter(ReferralBook.id == book_id).first()
            if book_obj:
                book_filter = book_obj.name

        return {
            "activities": result_activities,
            "activity_counts": activity_counts,
            "total_activities": len(activities),
            "book_filter": book_filter,
            "date_range": {"start": start_date, "end": end_date},
            "generated_at": datetime.now(),
        }

    def render_registration_history_excel(
        self,
        book_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        activity_type: Optional[str] = None,
    ) -> BytesIO:
        """Render Excel for registration history (multi-sheet).

        Sheets:
        - Sheet 1: Activity Log (all activity records)
        - Sheet 2: Summary (counts by type, by book, by week)
        """
        data = self.get_registration_history(book_id, start_date, end_date, activity_type)

        buffer = BytesIO()
        workbook = Workbook()

        # Sheet 1: Activity Log
        sheet1 = workbook.active
        sheet1.title = "Activity Log"

        headers = [
            "Date",
            "Member",
            "Card #",
            "Book",
            "Tier",
            "Activity",
            "Reason",
            "Previous APN",
            "New APN",
            "Performed By",
        ]
        sheet1.append(headers)

        for act in data["activities"]:
            sheet1.append(
                [
                    act["activity_date"].strftime("%m/%d/%Y %H:%M")
                    if act["activity_date"]
                    else "",
                    act["member_name"],
                    act["member_card"],
                    act["book_name"],
                    act["tier"],
                    act["activity_type"],
                    act["reason"],
                    str(act["previous_apn"]) if act["previous_apn"] else "",
                    str(act["new_apn"]) if act["new_apn"] else "",
                    act["performed_by"],
                ]
            )

        # Sheet 2: Summary
        sheet2 = workbook.create_sheet("Summary")
        sheet2.append(["Activity Type Summary"])
        sheet2.append(["Type", "Count"])
        for act_type, count in sorted(data["activity_counts"].items()):
            sheet2.append([act_type.upper(), count])

        sheet2.append([])
        sheet2.append(["Total Activities", data["total_activities"]])
        sheet2.append(["Date Range", f"{data['date_range']['start']} to {data['date_range']['end']}"])
        sheet2.append(["Book Filter", data["book_filter"]])

        workbook.save(buffer)
        buffer.seek(0)
        return buffer

    # --- Report 13: Check Mark Report (P1) ---

    def get_check_mark_report(self, book_id: Optional[int] = None) -> dict:
        """Assemble check mark status across members.

        Shows all members with 1+ active check marks, organized by book.
        Highlights members at the 2-mark limit.

        Args:
            book_id: Filter to specific book. None = all books.

        Returns: dict with books, totals, generated_at
        """
        query = self.db.query(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED,
            BookRegistration.check_marks > 0,
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        registrations = query.all()

        # Group by book
        books_data = {}
        total_with_marks = 0
        total_at_limit = 0

        for reg in registrations:
            book = reg.book
            book_id_key = book.id if book else 0
            book_name = book.name if book else "Unknown"

            if book_id_key not in books_data:
                books_data[book_id_key] = {
                    "book_name": book_name,
                    "book_id": book_id_key,
                    "members_with_marks": [],
                    "total_with_marks": 0,
                    "at_limit_count": 0,
                }

            member = reg.member
            member_name = (
                f"{member.last_name}, {member.first_name}" if member else "Unknown"
            )
            member_card = member.member_number if member else ""
            at_limit = reg.check_marks >= 2

            books_data[book_id_key]["members_with_marks"].append(
                {
                    "member_name": member_name,
                    "member_card": member_card,
                    "check_mark_count": reg.check_marks,
                    "at_limit": at_limit,
                    "marks": [],  # Would need a separate CheckMark model for details
                }
            )
            books_data[book_id_key]["total_with_marks"] += 1
            if at_limit:
                books_data[book_id_key]["at_limit_count"] += 1

            total_with_marks += 1
            if at_limit:
                total_at_limit += 1

        # Sort members within each book by check_mark_count descending
        for book_data in books_data.values():
            book_data["members_with_marks"].sort(
                key=lambda m: m["check_mark_count"], reverse=True
            )

        # Build book filter display
        book_filter = "All Books"
        if book_id:
            book_obj = self.db.query(ReferralBook).filter(ReferralBook.id == book_id).first()
            if book_obj:
                book_filter = book_obj.name

        return {
            "books": list(books_data.values()),
            "total_members_with_marks": total_with_marks,
            "total_at_limit": total_at_limit,
            "book_filter": book_filter,
            "generated_at": datetime.now(),
        }

    def render_check_mark_report_pdf(self, book_id: Optional[int] = None) -> BytesIO:
        """Render PDF for check mark report.

        Visual indicators:
        - 1 mark: yellow row
        - 2 marks: orange/red row with warning icon
        """
        data = self.get_check_mark_report(book_id)
        return self._render_pdf("check_mark_report.html", data)

    # --- Report 14: Re-Sign Due List (P0) ---

    # Configurable re-sign period (Business Rule #7)
    RE_SIGN_DAYS = 30

    def get_re_sign_due_list(
        self, days_ahead: int = 7, include_overdue: bool = True
    ) -> dict:
        """Assemble list of members whose 30-day re-sign is approaching or overdue.

        THIS IS A CRITICAL DAILY REPORT.
        Missing a re-sign = member dropped from ALL books (Rule #7).

        Args:
            days_ahead: Show members due within this many days (default 7)
            include_overdue: Include already-overdue members (default True)

        Returns: dict with overdue, due_soon, counts, generated_at
        """
        today = date.today()

        # Get all registered members
        registrations = (
            self.db.query(BookRegistration)
            .filter(BookRegistration.status == RegistrationStatus.REGISTERED)
            .all()
        )

        overdue = []
        due_soon = []

        for reg in registrations:
            member = reg.member
            book = reg.book
            member_name = (
                f"{member.last_name}, {member.first_name}" if member else "Unknown"
            )
            member_card = member.member_number if member else ""
            book_name = book.name if book else "Unknown"
            tier = book.book_number if book else 1

            # Calculate re-sign due date
            # Use last_re_sign_date if available, otherwise registered_date
            base_date = reg.last_re_sign_date or reg.registered_date
            if not base_date:
                continue  # Skip if no date available

            re_sign_due = base_date + timedelta(days=self.RE_SIGN_DAYS)
            days_diff = (re_sign_due - today).days

            record = {
                "member_name": member_name,
                "member_card": member_card,
                "book_name": book_name,
                "tier": tier,
                "apn": reg.registration_number,
                "last_re_sign": reg.last_re_sign_date or reg.registered_date,
                "re_sign_due": re_sign_due,
            }

            if days_diff < 0:
                # Overdue
                if include_overdue:
                    record["days_overdue"] = abs(days_diff)
                    overdue.append(record)
            elif days_diff <= days_ahead:
                # Due soon
                record["days_until_due"] = days_diff
                due_soon.append(record)

        # Sort overdue by days_overdue descending (most critical first)
        overdue.sort(key=lambda x: x["days_overdue"], reverse=True)

        # Sort due_soon by days_until_due ascending (soonest first)
        due_soon.sort(key=lambda x: x["days_until_due"])

        return {
            "overdue": overdue,
            "due_soon": due_soon,
            "overdue_count": len(overdue),
            "due_soon_count": len(due_soon),
            "days_ahead": days_ahead,
            "generated_at": datetime.now(),
        }

    def render_re_sign_due_list_pdf(
        self, days_ahead: int = 7, include_overdue: bool = True
    ) -> BytesIO:
        """Render PDF for re-sign due list.

        Urgency styling:
        - Overdue section: red header, bold rows
        - Due within 3 days: orange rows
        - Due within 7 days: yellow rows
        """
        data = self.get_re_sign_due_list(days_ahead, include_overdue)
        return self._render_pdf("re_sign_due_list.html", data)

    # =====================================================================
    # WEEK 36 P1 REPORTS: Registration & Book Analytics
    # =====================================================================

    # --- Report P1-1: Registration Activity Summary ---

    def get_registration_activity_summary(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Aggregate registration activity by time period.

        Shows registrations in, re-signs, drops, and dispatched out.

        Args:
            start_date: Start of date range. Defaults to 30 days ago.
            end_date: End of date range. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        query = self.db.query(RegistrationActivity).filter(
            func.date(RegistrationActivity.activity_date) >= start_date,
            func.date(RegistrationActivity.activity_date) <= end_date,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        activities = query.order_by(RegistrationActivity.activity_date).all()

        # Aggregate by week
        weekly_data = {}
        for act in activities:
            week_start = act.activity_date.date() - timedelta(
                days=act.activity_date.weekday()
            )
            week_key = week_start.isoformat()

            if week_key not in weekly_data:
                weekly_data[week_key] = {
                    "period": week_key,
                    "registrations_in": 0,
                    "re_signs": 0,
                    "drops": 0,
                    "dispatched_out": 0,
                    "net_change": 0,
                }

            action = act.action
            if action == RegistrationAction.REGISTER:
                weekly_data[week_key]["registrations_in"] += 1
            elif action == RegistrationAction.RE_SIGN:
                weekly_data[week_key]["re_signs"] += 1
            elif action == RegistrationAction.RE_REGISTER:
                weekly_data[week_key]["registrations_in"] += 1
            elif action in [RegistrationAction.ROLL_OFF, RegistrationAction.RESIGN]:
                weekly_data[week_key]["drops"] += 1
            elif action == RegistrationAction.DISPATCH:
                weekly_data[week_key]["dispatched_out"] += 1

        # Calculate net change
        data = []
        for week_key in sorted(weekly_data.keys()):
            row = weekly_data[week_key]
            row["net_change"] = (
                row["registrations_in"] - row["drops"] - row["dispatched_out"]
            )
            data.append(row)

        # Summary stats
        total_in = sum(r["registrations_in"] for r in data)
        total_drops = sum(r["drops"] for r in data)
        total_out = sum(r["dispatched_out"] for r in data)
        total_re_signs = sum(r["re_signs"] for r in data)
        net_total = total_in - total_drops - total_out
        busiest_period = max(data, key=lambda x: x["registrations_in"])["period"] if data else None

        return {
            "data": data,
            "summary": {
                "total_registrations_in": total_in,
                "total_re_signs": total_re_signs,
                "total_drops": total_drops,
                "total_dispatched_out": total_out,
                "net_change": net_total,
                "busiest_period": busiest_period,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Registration Activity Summary",
        }

    def render_registration_activity_summary_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for registration activity summary."""
        data = self.get_registration_activity_summary(start_date, end_date, book_id)
        return self._render_pdf("registration_activity_summary.html", data)

    def render_registration_activity_summary_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for registration activity summary."""
        data = self.get_registration_activity_summary(start_date, end_date, book_id)

        headers = [
            "Period",
            "Registrations In",
            "Re-Signs",
            "Drops",
            "Dispatched Out",
            "Net Change",
        ]

        rows = [
            [
                r["period"],
                r["registrations_in"],
                r["re_signs"],
                r["drops"],
                r["dispatched_out"],
                r["net_change"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(
            headers, rows, "Activity Summary", data["report_name"]
        )

    # --- Report P1-2: Registration by Classification ---

    def get_registration_by_classification(
        self,
        snapshot_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Breakdown of active registrations by member classification.

        Args:
            snapshot_date: Date for snapshot. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        if not snapshot_date:
            snapshot_date = date.today()

        query = self.db.query(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED,
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        registrations = query.all()

        # Group by classification and book tier
        classification_data = {}
        for reg in registrations:
            member = reg.member
            book = reg.book
            if not member:
                continue

            classification = (
                member.classification.value if member.classification else "Unknown"
            )
            tier = book.book_number if book else 1

            if classification not in classification_data:
                classification_data[classification] = {
                    "classification": classification,
                    "tier_1": 0,
                    "tier_2": 0,
                    "tier_3": 0,
                    "tier_4": 0,
                    "total": 0,
                }

            classification_data[classification][f"tier_{tier}"] += 1
            classification_data[classification]["total"] += 1

        data = sorted(
            classification_data.values(),
            key=lambda x: x["total"],
            reverse=True,
        )

        # Summary
        tier_totals = {1: 0, 2: 0, 3: 0, 4: 0}
        for row in data:
            for t in range(1, 5):
                tier_totals[t] += row[f"tier_{t}"]

        dominant_per_tier = {}
        for t in range(1, 5):
            tier_key = f"tier_{t}"
            max_class = max(data, key=lambda x: x[tier_key]) if data else None
            dominant_per_tier[tier_key] = max_class["classification"] if max_class and max_class[tier_key] > 0 else None

        return {
            "data": data,
            "summary": {
                "tier_totals": tier_totals,
                "dominant_per_tier": dominant_per_tier,
                "total_registrations": sum(r["total"] for r in data),
            },
            "filters": {
                "snapshot_date": snapshot_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Registration by Classification",
        }

    def render_registration_by_classification_pdf(
        self,
        snapshot_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for registration by classification."""
        data = self.get_registration_by_classification(snapshot_date, book_id)
        return self._render_pdf("registration_by_classification.html", data)

    def render_registration_by_classification_excel(
        self,
        snapshot_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for registration by classification."""
        data = self.get_registration_by_classification(snapshot_date, book_id)

        headers = ["Classification", "Tier 1", "Tier 2", "Tier 3", "Tier 4", "Total"]

        rows = [
            [
                r["classification"],
                r["tier_1"],
                r["tier_2"],
                r["tier_3"],
                r["tier_4"],
                r["total"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(
            headers, rows, "By Classification", data["report_name"]
        )

    # --- Report P1-3: Re-Registration Analysis ---

    def get_re_registration_analysis(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        reason: Optional[str] = None,
    ) -> dict:
        """Patterns of re-registration — short calls, 90-day cycles, voluntary.

        Args:
            start_date: Start of date range. Defaults to 90 days ago.
            end_date: End of date range. Defaults to today.
            book_id: Filter to specific book. None = all books.
            reason: Filter by re-registration reason. None = all.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        query = self.db.query(RegistrationActivity).filter(
            RegistrationActivity.action == RegistrationAction.RE_REGISTER,
            func.date(RegistrationActivity.activity_date) >= start_date,
            func.date(RegistrationActivity.activity_date) <= end_date,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        if reason:
            query = query.filter(RegistrationActivity.reason.ilike(f"%{reason}%"))

        activities = query.order_by(RegistrationActivity.activity_date.desc()).all()

        # Build data rows
        data = []
        reason_counts = {}
        gap_days_list = []
        member_re_reg_counts = {}

        for act in activities:
            member = act.member
            book = act.book
            member_name = (
                f"{member.last_name}, {member.first_name}" if member else "Unknown"
            )
            book_name = book.name if book else "Unknown"

            # Determine reason category
            act_reason = act.reason or "Unknown"
            reason_cat = "Other"
            reason_lower = act_reason.lower()
            if "short" in reason_lower:
                reason_cat = "Short Call"
            elif "90" in reason_lower or "cycle" in reason_lower:
                reason_cat = "90-Day Cycle"
            elif "voluntary" in reason_lower or "re-sign" in reason_lower:
                reason_cat = "Voluntary"

            reason_counts[reason_cat] = reason_counts.get(reason_cat, 0) + 1

            # Track repeat re-registrants
            member_id = member.id if member else 0
            member_re_reg_counts[member_id] = member_re_reg_counts.get(member_id, 0) + 1

            # Calculate gap from previous registration (approximate)
            prev_pos = act.previous_position
            gap = 0
            if prev_pos and act.new_position:
                # APN difference as proxy for gap days (very rough)
                gap = abs(float(act.new_position) - float(prev_pos))

            gap_days_list.append(gap)

            data.append({
                "member_name": member_name,
                "member_id": member_id,
                "book_name": book_name,
                "re_reg_date": act.activity_date.date()
                if isinstance(act.activity_date, datetime)
                else act.activity_date,
                "reason": act_reason,
                "reason_category": reason_cat,
                "previous_apn": prev_pos,
                "new_apn": act.new_position,
            })

        # Summary
        avg_gap = sum(gap_days_list) / len(gap_days_list) if gap_days_list else 0
        repeat_re_registrants = sum(1 for c in member_re_reg_counts.values() if c > 1)

        return {
            "data": data,
            "summary": {
                "total_re_registrations": len(data),
                "by_reason": reason_counts,
                "avg_gap_estimate": round(avg_gap, 1),
                "repeat_re_registrants": repeat_re_registrants,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
                "reason": reason,
            },
            "generated_at": datetime.now(),
            "report_name": "Re-Registration Analysis",
        }

    def render_re_registration_analysis_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        reason: Optional[str] = None,
    ) -> BytesIO:
        """Render PDF for re-registration analysis."""
        data = self.get_re_registration_analysis(start_date, end_date, book_id, reason)
        return self._render_pdf("re_registration_analysis.html", data)

    def render_re_registration_analysis_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        reason: Optional[str] = None,
    ) -> BytesIO:
        """Render Excel for re-registration analysis."""
        data = self.get_re_registration_analysis(start_date, end_date, book_id, reason)

        headers = [
            "Member",
            "Book",
            "Re-Reg Date",
            "Reason",
            "Category",
            "Previous APN",
            "New APN",
        ]

        rows = [
            [
                r["member_name"],
                r["book_name"],
                r["re_reg_date"].strftime("%m/%d/%Y") if r["re_reg_date"] else "",
                r["reason"],
                r["reason_category"],
                str(r["previous_apn"]) if r["previous_apn"] else "",
                str(r["new_apn"]) if r["new_apn"] else "",
            ]
            for r in data["data"]
        ]

        return self._render_excel(
            headers, rows, "Re-Registration", data["report_name"]
        )

    # --- Report P1-4: Registration Duration Report ---

    def get_registration_duration(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Average time on book before dispatch or drop.

        Args:
            start_date: Start of date range. Defaults to 90 days ago.
            end_date: End of date range. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction
        from statistics import median

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Get all books
        book_query = self.db.query(ReferralBook).filter(ReferralBook.is_active.is_(True))
        if book_id:
            book_query = book_query.filter(ReferralBook.id == book_id)
        books = book_query.all()

        data = []
        all_durations = []

        for book in books:
            tier = book.book_number or 1

            # Get dispatches/rolloffs for this book in the period
            exit_activities = (
                self.db.query(RegistrationActivity)
                .filter(
                    RegistrationActivity.book_id == book.id,
                    RegistrationActivity.action.in_([
                        RegistrationAction.DISPATCH,
                        RegistrationAction.ROLL_OFF,
                    ]),
                    func.date(RegistrationActivity.activity_date) >= start_date,
                    func.date(RegistrationActivity.activity_date) <= end_date,
                )
                .all()
            )

            # Calculate durations (simplified - using registration to exit time)
            durations = []
            for act in exit_activities:
                # Find corresponding registration activity
                reg_activity = (
                    self.db.query(RegistrationActivity)
                    .filter(
                        RegistrationActivity.member_id == act.member_id,
                        RegistrationActivity.book_id == book.id,
                        RegistrationActivity.action.in_([
                            RegistrationAction.REGISTER,
                            RegistrationAction.RE_REGISTER,
                        ]),
                        RegistrationActivity.activity_date < act.activity_date,
                    )
                    .order_by(RegistrationActivity.activity_date.desc())
                    .first()
                )

                if reg_activity:
                    reg_date = (
                        reg_activity.activity_date.date()
                        if isinstance(reg_activity.activity_date, datetime)
                        else reg_activity.activity_date
                    )
                    exit_date = (
                        act.activity_date.date()
                        if isinstance(act.activity_date, datetime)
                        else act.activity_date
                    )
                    duration = (exit_date - reg_date).days
                    if duration >= 0:
                        durations.append(duration)
                        all_durations.append(duration)

            # Still active count
            still_active = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book.id,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .count()
            )

            if durations:
                avg_days = round(sum(durations) / len(durations), 1)
                med_days = round(median(durations), 1)
                min_days = min(durations)
                max_days = max(durations)
            else:
                avg_days = med_days = min_days = max_days = 0

            data.append({
                "book_name": book.name,
                "book_id": book.id,
                "tier": tier,
                "avg_days": avg_days,
                "median_days": med_days,
                "min_days": min_days,
                "max_days": max_days,
                "still_active_count": still_active,
                "completed_count": len(durations),
            })

        # Sort by avg_days descending
        data.sort(key=lambda x: x["avg_days"], reverse=True)

        # Summary
        overall_avg = (
            round(sum(all_durations) / len(all_durations), 1) if all_durations else 0
        )
        fastest_book = min(data, key=lambda x: x["avg_days"])["book_name"] if data else None
        slowest_book = max(data, key=lambda x: x["avg_days"])["book_name"] if data else None

        return {
            "data": data,
            "summary": {
                "overall_avg_wait": overall_avg,
                "fastest_book": fastest_book,
                "slowest_book": slowest_book,
                "total_completed": len(all_durations),
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Registration Duration Report",
        }

    def render_registration_duration_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for registration duration report."""
        data = self.get_registration_duration(start_date, end_date, book_id)
        return self._render_pdf("registration_duration.html", data)

    def render_registration_duration_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for registration duration report."""
        data = self.get_registration_duration(start_date, end_date, book_id)

        headers = [
            "Book",
            "Tier",
            "Avg Days",
            "Median Days",
            "Min",
            "Max",
            "Still Active",
            "Completed",
        ]

        rows = [
            [
                r["book_name"],
                r["tier"],
                r["avg_days"],
                r["median_days"],
                r["min_days"],
                r["max_days"],
                r["still_active_count"],
                r["completed_count"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Duration", data["report_name"])

    # --- Report P1-5: Book Health Summary ---

    def get_book_health_summary(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Per-book dashboard stats — health indicators.

        Args:
            start_date: Start of period. Defaults to 30 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction
        from statistics import median

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Get all books
        book_query = self.db.query(ReferralBook).filter(ReferralBook.is_active.is_(True))
        if book_id:
            book_query = book_query.filter(ReferralBook.id == book_id)
        books = book_query.all()

        data = []

        for book in books:
            # Current active count
            active_count = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book.id,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .count()
            )

            # Activity counts in period
            activities = (
                self.db.query(RegistrationActivity)
                .filter(
                    RegistrationActivity.book_id == book.id,
                    func.date(RegistrationActivity.activity_date) >= start_date,
                    func.date(RegistrationActivity.activity_date) <= end_date,
                )
                .all()
            )

            dispatched = sum(1 for a in activities if a.action == RegistrationAction.DISPATCH)
            dropped = sum(1 for a in activities if a.action in [RegistrationAction.ROLL_OFF, RegistrationAction.RESIGN])
            re_signed = sum(1 for a in activities if a.action == RegistrationAction.RE_SIGN)

            # Calculate average wait for active registrations
            active_regs = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book.id,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .all()
            )

            wait_days = []
            today = date.today()
            for reg in active_regs:
                reg_date = reg.registration_date
                if reg_date:
                    if isinstance(reg_date, datetime):
                        reg_date = reg_date.date()
                    days = (today - reg_date).days
                    if days >= 0:
                        wait_days.append(days)

            avg_wait = round(sum(wait_days) / len(wait_days), 1) if wait_days else 0

            # Fill rate: dispatched / (dispatched + dropped + still_active)
            total_exits = dispatched + dropped
            fill_rate = round(dispatched / total_exits * 100, 1) if total_exits > 0 else 0

            data.append({
                "book_name": book.name,
                "book_id": book.id,
                "active": active_count,
                "dispatched": dispatched,
                "dropped": dropped,
                "re_signed": re_signed,
                "avg_wait_days": avg_wait,
                "fill_rate": fill_rate,
            })

        # Sort by fill rate descending
        data.sort(key=lambda x: x["fill_rate"], reverse=True)

        # Summary
        healthiest = data[0]["book_name"] if data else None
        concerning = [d["book_name"] for d in data if d["fill_rate"] < 50] if data else []

        return {
            "data": data,
            "summary": {
                "healthiest_book": healthiest,
                "concerning_books": concerning[:3],
                "total_active_across_books": sum(d["active"] for d in data),
                "total_dispatched": sum(d["dispatched"] for d in data),
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Book Health Summary",
        }

    def render_book_health_summary_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for book health summary (landscape)."""
        data = self.get_book_health_summary(start_date, end_date, book_id)
        return self._render_pdf("book_health_summary.html", data)

    # --- Report P1-6: Book Comparison ---

    def get_book_comparison(
        self,
        book_ids: list[int],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> dict:
        """Side-by-side metrics for selected books.

        Args:
            book_ids: List of book IDs to compare (min 2).
            start_date: Start of period. Defaults to 30 days ago.
            end_date: End of period. Defaults to today.

        Returns: dict with data, summary, filters, generated_at
        """
        if len(book_ids) < 2:
            return {
                "data": [],
                "summary": {"error": "At least 2 books required"},
                "filters": {"book_ids": book_ids},
                "generated_at": datetime.now(),
                "report_name": "Book Comparison",
            }

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Get health data for selected books
        health_data = self.get_book_health_summary(start_date, end_date)

        # Filter to selected books
        book_data = {
            d["book_id"]: d for d in health_data["data"] if d["book_id"] in book_ids
        }

        # Build comparison rows (metrics as rows, books as columns)
        metrics = [
            ("Active Registrations", "active"),
            ("Dispatched (Period)", "dispatched"),
            ("Dropped (Period)", "dropped"),
            ("Re-Signed (Period)", "re_signed"),
            ("Avg Wait Days", "avg_wait_days"),
            ("Fill Rate %", "fill_rate"),
        ]

        data = []
        for metric_name, metric_key in metrics:
            row = {"metric": metric_name}
            for bid in book_ids:
                if bid in book_data:
                    row[f"book_{bid}"] = book_data[bid][metric_key]
                else:
                    row[f"book_{bid}"] = "N/A"
            data.append(row)

        # Summary: key differentiators
        differentiators = []
        if book_data:
            fill_rates = [(bid, book_data[bid]["fill_rate"]) for bid in book_ids if bid in book_data]
            if fill_rates:
                max_fr = max(fill_rates, key=lambda x: x[1])
                min_fr = min(fill_rates, key=lambda x: x[1])
                if max_fr[1] - min_fr[1] > 10:
                    max_book = book_data[max_fr[0]]["book_name"]
                    min_book = book_data[min_fr[0]]["book_name"]
                    differentiators.append(
                        f"{max_book} has {max_fr[1] - min_fr[1]:.1f}% higher fill rate than {min_book}"
                    )

        # Get book names for header
        book_names = {bid: book_data[bid]["book_name"] for bid in book_ids if bid in book_data}

        return {
            "data": data,
            "book_ids": book_ids,
            "book_names": book_names,
            "summary": {
                "differentiators": differentiators,
                "recommendations": ["Review low fill rate books for bottlenecks"] if differentiators else [],
            },
            "filters": {
                "book_ids": book_ids,
                "start_date": start_date,
                "end_date": end_date,
            },
            "generated_at": datetime.now(),
            "report_name": "Book Comparison",
        }

    def render_book_comparison_pdf(
        self,
        book_ids: list[int],
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> BytesIO:
        """Render PDF for book comparison."""
        data = self.get_book_comparison(book_ids, start_date, end_date)
        return self._render_pdf("book_comparison.html", data)

    # --- Report P1-7: Book Position Report (Detailed Queue) ---

    def get_book_position_report(
        self,
        book_id: int,
        tier: Optional[int] = None,
        status: Optional[str] = None,
    ) -> Optional[dict]:
        """Full position listing for a specific book.

        CRITICAL: APN must be DECIMAL(10,2) — never truncate.

        Args:
            book_id: The book to report on (required).
            tier: Filter by tier. None = all tiers.
            status: Filter by status. None = REGISTERED only.

        Returns: dict with data, summary, filters, generated_at or None if book not found
        """
        book = self.db.query(ReferralBook).filter(ReferralBook.id == book_id).first()
        if not book:
            return None

        reg_status = RegistrationStatus.REGISTERED
        if status:
            try:
                reg_status = RegistrationStatus(status.lower())
            except ValueError:
                pass

        query = self.db.query(BookRegistration).filter(
            BookRegistration.book_id == book_id,
            BookRegistration.status == reg_status,
        )

        # Sort by APN ascending (dispatch order)
        registrations = query.order_by(BookRegistration.registration_number.asc()).all()

        data = []
        tier_counts = {1: 0, 2: 0, 3: 0, 4: 0}
        upcoming_re_signs = 0
        today = date.today()

        for i, reg in enumerate(registrations, 1):
            member = reg.member
            member_name = (
                f"{member.last_name}, {member.first_name}" if member else "Unknown"
            )

            reg_date = reg.registration_date
            if isinstance(reg_date, datetime):
                reg_date = reg_date.date()

            days_waiting = (today - reg_date).days if reg_date else 0

            # Calculate re-sign due
            base_date = reg.last_re_sign_date or reg_date
            if isinstance(base_date, datetime):
                base_date = base_date.date()
            re_sign_due = base_date + timedelta(days=30) if base_date else None

            days_until_re_sign = (re_sign_due - today).days if re_sign_due else None
            if days_until_re_sign is not None and days_until_re_sign <= 7:
                upcoming_re_signs += 1

            book_tier = book.book_number or 1
            tier_counts[book_tier] = tier_counts.get(book_tier, 0) + 1

            data.append({
                "position": i,
                "apn": f"{reg.registration_number:.2f}",  # CRITICAL: Always 2 decimal places
                "apn_raw": reg.registration_number,
                "member_name": member_name,
                "member_id": member.id if member else None,
                "tier": book_tier,
                "reg_date": reg_date,
                "days_waiting": days_waiting,
                "re_sign_due": re_sign_due,
                "check_marks": reg.check_marks,
            })

        return {
            "data": data,
            "book": {
                "id": book.id,
                "name": book.name,
                "code": book.code,
            },
            "summary": {
                "total_active": len(data),
                "tier_counts": tier_counts,
                "upcoming_re_signs": upcoming_re_signs,
            },
            "filters": {
                "book_id": book_id,
                "tier": tier,
                "status": status or "registered",
            },
            "generated_at": datetime.now(),
            "report_name": f"Book Position Report - {book.name}",
        }

    def render_book_position_report_pdf(
        self,
        book_id: int,
        tier: Optional[int] = None,
        status: Optional[str] = None,
    ) -> Optional[BytesIO]:
        """Render PDF for book position report."""
        data = self.get_book_position_report(book_id, tier, status)
        if not data:
            return None
        return self._render_pdf("book_position_report.html", data)

    def render_book_position_report_excel(
        self,
        book_id: int,
        tier: Optional[int] = None,
        status: Optional[str] = None,
    ) -> Optional[BytesIO]:
        """Render Excel for book position report."""
        data = self.get_book_position_report(book_id, tier, status)
        if not data:
            return None

        headers = [
            "Position",
            "APN",
            "Member Name",
            "Tier",
            "Reg Date",
            "Days Waiting",
            "Re-Sign Due",
            "Check Marks",
        ]

        rows = [
            [
                r["position"],
                r["apn"],  # Pre-formatted with 2 decimal places
                r["member_name"],
                r["tier"],
                r["reg_date"].strftime("%m/%d/%Y") if r["reg_date"] else "",
                r["days_waiting"],
                r["re_sign_due"].strftime("%m/%d/%Y") if r["re_sign_due"] else "",
                r["check_marks"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Queue", data["report_name"])

    # --- Report P1-8: Book Turnover Report ---

    def get_book_turnover(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        granularity: str = "weekly",
    ) -> dict:
        """Registrations in vs out by period — measures churn.

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.
            granularity: "weekly" or "monthly".

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Get all books
        book_query = self.db.query(ReferralBook).filter(ReferralBook.is_active.is_(True))
        if book_id:
            book_query = book_query.filter(ReferralBook.id == book_id)
        books = {b.id: b.name for b in book_query.all()}

        # Get activities
        activities = (
            self.db.query(RegistrationActivity)
            .filter(
                func.date(RegistrationActivity.activity_date) >= start_date,
                func.date(RegistrationActivity.activity_date) <= end_date,
            )
            .all()
        )

        if book_id:
            activities = [a for a in activities if a.book_id == book_id]

        # Aggregate by period and book
        period_data = {}

        for act in activities:
            act_date = (
                act.activity_date.date()
                if isinstance(act.activity_date, datetime)
                else act.activity_date
            )

            if granularity == "monthly":
                period_key = act_date.strftime("%Y-%m")
            else:
                week_start = act_date - timedelta(days=act_date.weekday())
                period_key = week_start.isoformat()

            book_name = books.get(act.book_id, "Unknown")
            full_key = (period_key, book_name)

            if full_key not in period_data:
                period_data[full_key] = {
                    "period": period_key,
                    "book": book_name,
                    "in_new": 0,
                    "in_re_reg": 0,
                    "out_dispatched": 0,
                    "out_dropped": 0,
                    "out_expired": 0,
                }

            action = act.action
            if action == RegistrationAction.REGISTER:
                period_data[full_key]["in_new"] += 1
            elif action == RegistrationAction.RE_REGISTER:
                period_data[full_key]["in_re_reg"] += 1
            elif action == RegistrationAction.DISPATCH:
                period_data[full_key]["out_dispatched"] += 1
            elif action in [RegistrationAction.ROLL_OFF, RegistrationAction.RESIGN]:
                period_data[full_key]["out_dropped"] += 1

        # Build output
        data = []
        for key in sorted(period_data.keys()):
            row = period_data[key]
            total_in = row["in_new"] + row["in_re_reg"]
            total_out = row["out_dispatched"] + row["out_dropped"] + row["out_expired"]
            net = total_in - total_out
            turnover_rate = round((total_in + total_out) / 2 / max(total_in, 1) * 100, 1)

            data.append({
                "period": row["period"],
                "book": row["book"],
                "in_total": total_in,
                "out_total": total_out,
                "net": net,
                "turnover_rate": turnover_rate,
            })

        # Summary
        if data:
            by_book = {}
            for row in data:
                book = row["book"]
                if book not in by_book:
                    by_book[book] = {"turnover_sum": 0, "count": 0}
                by_book[book]["turnover_sum"] += row["turnover_rate"]
                by_book[book]["count"] += 1

            avg_turnovers = {
                b: d["turnover_sum"] / d["count"] for b, d in by_book.items()
            }
            highest_churn = max(avg_turnovers, key=avg_turnovers.get)
            steadiest = min(avg_turnovers, key=avg_turnovers.get)
        else:
            highest_churn = steadiest = None

        return {
            "data": data,
            "summary": {
                "highest_churn_book": highest_churn,
                "steadiest_book": steadiest,
                "total_periods": len(set(d["period"] for d in data)),
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
                "granularity": granularity,
            },
            "generated_at": datetime.now(),
            "report_name": "Book Turnover Report",
        }

    def render_book_turnover_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        granularity: str = "weekly",
    ) -> BytesIO:
        """Render PDF for book turnover report."""
        data = self.get_book_turnover(start_date, end_date, book_id, granularity)
        return self._render_pdf("book_turnover.html", data)

    def render_book_turnover_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        granularity: str = "weekly",
    ) -> BytesIO:
        """Render Excel for book turnover report."""
        data = self.get_book_turnover(start_date, end_date, book_id, granularity)

        headers = ["Period", "Book", "In", "Out", "Net", "Turnover Rate %"]

        rows = [
            [
                r["period"],
                r["book"],
                r["in_total"],
                r["out_total"],
                r["net"],
                r["turnover_rate"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Turnover", data["report_name"])

    # --- Report P1-9: Check Mark Summary ---

    def get_check_mark_summary(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Aggregate check mark statistics by book and period.

        Args:
            start_date: Start of period. Defaults to 30 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=30)

        # Get all books
        book_query = self.db.query(ReferralBook).filter(ReferralBook.is_active.is_(True))
        if book_id:
            book_query = book_query.filter(ReferralBook.id == book_id)
        books = {b.id: b.name for b in book_query.all()}

        # Get check mark activities
        cm_activities = (
            self.db.query(RegistrationActivity)
            .filter(
                RegistrationActivity.action == RegistrationAction.CHECK_MARK,
                func.date(RegistrationActivity.activity_date) >= start_date,
                func.date(RegistrationActivity.activity_date) <= end_date,
            )
            .all()
        )

        if book_id:
            cm_activities = [a for a in cm_activities if a.book_id == book_id]

        # Get roll-offs (3rd check mark)
        rolloffs = (
            self.db.query(RegistrationActivity)
            .filter(
                RegistrationActivity.action == RegistrationAction.ROLL_OFF,
                func.date(RegistrationActivity.activity_date) >= start_date,
                func.date(RegistrationActivity.activity_date) <= end_date,
            )
            .all()
        )

        if book_id:
            rolloffs = [a for a in rolloffs if a.book_id == book_id]

        # Aggregate by book
        book_stats = {}
        for book_id_key, book_name in books.items():
            book_cms = [a for a in cm_activities if a.book_id == book_id_key]
            book_rolloffs = [a for a in rolloffs if a.book_id == book_id_key]

            # Count current check mark levels
            current_regs = (
                self.db.query(BookRegistration)
                .filter(
                    BookRegistration.book_id == book_id_key,
                    BookRegistration.status == RegistrationStatus.REGISTERED,
                )
                .all()
            )

            at_1_cm = sum(1 for r in current_regs if r.check_marks == 1)
            at_2_cm = sum(1 for r in current_regs if r.check_marks >= 2)

            book_stats[book_name] = {
                "book": book_name,
                "check_marks_issued": len(book_cms),
                "members_at_1_cm": at_1_cm,
                "members_at_2_cm": at_2_cm,
                "members_rolled_off": len(book_rolloffs),
            }

        data = list(book_stats.values())
        data.sort(key=lambda x: x["check_marks_issued"], reverse=True)

        # Summary
        total_cms = sum(d["check_marks_issued"] for d in data)
        total_at_limit = sum(d["members_at_2_cm"] for d in data)
        total_rolled = sum(d["members_rolled_off"] for d in data)

        return {
            "data": data,
            "summary": {
                "total_check_marks_issued": total_cms,
                "total_at_limit": total_at_limit,
                "total_rolled_off": total_rolled,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Check Mark Summary",
        }

    def render_check_mark_summary_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for check mark summary."""
        data = self.get_check_mark_summary(start_date, end_date, book_id)
        return self._render_pdf("check_mark_summary.html", data)

    def render_check_mark_summary_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for check mark summary."""
        data = self.get_check_mark_summary(start_date, end_date, book_id)

        headers = [
            "Book",
            "Check Marks Issued",
            "At 1 CM",
            "At 2 CM (Warning)",
            "Rolled Off (3rd CM)",
        ]

        rows = [
            [
                r["book"],
                r["check_marks_issued"],
                r["members_at_1_cm"],
                r["members_at_2_cm"],
                r["members_rolled_off"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "CM Summary", data["report_name"])

    # --- Report P1-10: Check Mark Trend ---

    def get_check_mark_trend(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Check mark issuance over time with trend analysis.

        Args:
            start_date: Start of period. Defaults to 90 days ago (min 3 months for trend).
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Get check mark activities
        query = self.db.query(RegistrationActivity).filter(
            RegistrationActivity.action == RegistrationAction.CHECK_MARK,
            func.date(RegistrationActivity.activity_date) >= start_date,
            func.date(RegistrationActivity.activity_date) <= end_date,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        activities = query.order_by(RegistrationActivity.activity_date).all()

        # Aggregate by week
        weekly_data = {}
        for act in activities:
            act_date = (
                act.activity_date.date()
                if isinstance(act.activity_date, datetime)
                else act.activity_date
            )
            week_start = act_date - timedelta(days=act_date.weekday())
            week_key = week_start.isoformat()

            if week_key not in weekly_data:
                weekly_data[week_key] = {"period": week_key, "cms_issued": 0}
            weekly_data[week_key]["cms_issued"] += 1

        # Build trend data
        data = []
        cumulative = 0
        values = []

        for week_key in sorted(weekly_data.keys()):
            row = weekly_data[week_key]
            cumulative += row["cms_issued"]
            values.append(row["cms_issued"])

            # Rolling average (last 4 weeks)
            recent = values[-4:] if len(values) >= 4 else values
            rolling_avg = round(sum(recent) / len(recent), 1)

            # Count members at risk (2 CMs)
            # This is a point-in-time query — approximate for the period
            if book_id:
                at_risk = (
                    self.db.query(BookRegistration)
                    .filter(
                        BookRegistration.book_id == book_id,
                        BookRegistration.status == RegistrationStatus.REGISTERED,
                        BookRegistration.check_marks >= 2,
                    )
                    .count()
                )
            else:
                at_risk = (
                    self.db.query(BookRegistration)
                    .filter(
                        BookRegistration.status == RegistrationStatus.REGISTERED,
                        BookRegistration.check_marks >= 2,
                    )
                    .count()
                )

            data.append({
                "period": row["period"],
                "cms_issued": row["cms_issued"],
                "cumulative": cumulative,
                "members_at_risk": at_risk,
                "rolling_avg": rolling_avg,
            })

        # Trend analysis
        if len(values) >= 2:
            first_half = values[: len(values) // 2]
            second_half = values[len(values) // 2 :]
            first_avg = sum(first_half) / len(first_half)
            second_avg = sum(second_half) / len(second_half)

            if second_avg > first_avg * 1.1:
                trend = "increasing"
            elif second_avg < first_avg * 0.9:
                trend = "decreasing"
            else:
                trend = "stable"
        else:
            trend = "insufficient data"

        # Projection
        projected_at_risk = data[-1]["members_at_risk"] if data else 0
        if trend == "increasing" and data:
            projected_at_risk = int(projected_at_risk * 1.1)

        return {
            "data": data,
            "summary": {
                "trend_direction": trend,
                "total_cms_period": cumulative,
                "current_at_risk": data[-1]["members_at_risk"] if data else 0,
                "projected_at_risk_next_period": projected_at_risk,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Check Mark Trend",
        }

    def render_check_mark_trend_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for check mark trend."""
        data = self.get_check_mark_trend(start_date, end_date, book_id)
        return self._render_pdf("check_mark_trend.html", data)

    # =====================================================================
    # WEEK 37 P1 REPORTS: Dispatch Operations & Employer Analytics
    # =====================================================================

    # --- Report W37-1: Weekly Dispatch Summary ---

    def get_weekly_dispatch_summary(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
    ) -> dict:
        """Weekly dispatch summary with breakdown by book and employer.

        Args:
            start_date: Start of period. Defaults to 4 weeks ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.
            employer_id: Filter to specific employer. None = all employers.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=28)

        # Query dispatches with joins
        query = (
            self.db.query(Dispatch)
            .join(BookRegistration, Dispatch.registration_id == BookRegistration.id)
            .join(ReferralBook, BookRegistration.book_id == ReferralBook.id)
            .filter(
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
            )
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)
        if employer_id:
            query = query.filter(Dispatch.employer_id == employer_id)

        dispatches = query.all()

        # Get book and employer names
        books = {b.id: b.name for b in self.db.query(ReferralBook).all()}
        employers = {
            o.id: o.name
            for o in self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER).all()
        }

        # Aggregate by week
        weekly_data = {}
        employer_counts = {}

        for d in dispatches:
            dispatch_date = (
                d.dispatch_date.date()
                if isinstance(d.dispatch_date, datetime)
                else d.dispatch_date
            )
            week_start = dispatch_date - timedelta(days=dispatch_date.weekday())
            week_key = week_start.isoformat()

            if week_key not in weekly_data:
                weekly_data[week_key] = {
                    "week": week_key,
                    "dispatches": 0,
                    "short_calls": 0,
                    "long_calls": 0,
                    "total_days": 0,
                    "books": {},
                    "employers": {},
                }

            weekly_data[week_key]["dispatches"] += 1

            # Count duration type
            if d.end_date:
                end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
                duration = (end - dispatch_date).days
                weekly_data[week_key]["total_days"] += duration
                if duration <= 10:
                    weekly_data[week_key]["short_calls"] += 1
                else:
                    weekly_data[week_key]["long_calls"] += 1

            # Track by book
            reg = self.db.query(BookRegistration).filter(
                BookRegistration.id == d.registration_id
            ).first()
            if reg:
                book_name = books.get(reg.book_id, f"Book {reg.book_id}")
                weekly_data[week_key]["books"][book_name] = (
                    weekly_data[week_key]["books"].get(book_name, 0) + 1
                )

            # Track by employer
            emp_name = employers.get(d.employer_id, f"Employer {d.employer_id}")
            weekly_data[week_key]["employers"][emp_name] = (
                weekly_data[week_key]["employers"].get(emp_name, 0) + 1
            )
            employer_counts[emp_name] = employer_counts.get(emp_name, 0) + 1

        # Build data array
        data = []
        for week_key in sorted(weekly_data.keys()):
            row = weekly_data[week_key]
            avg_duration = (
                round(row["total_days"] / row["dispatches"], 1)
                if row["dispatches"] > 0
                else 0
            )
            data.append({
                "week": row["week"],
                "dispatches": row["dispatches"],
                "short_calls": row["short_calls"],
                "long_calls": row["long_calls"],
                "avg_duration_days": avg_duration,
                "top_book": max(row["books"], key=row["books"].get) if row["books"] else "N/A",
                "top_employer": max(row["employers"], key=row["employers"].get) if row["employers"] else "N/A",
            })

        # Summary
        total_dispatches = sum(d["dispatches"] for d in data)
        busiest_week = max(data, key=lambda x: x["dispatches"])["week"] if data else "N/A"
        top_employer = max(employer_counts, key=employer_counts.get) if employer_counts else "N/A"

        return {
            "data": data,
            "summary": {
                "total_dispatches": total_dispatches,
                "busiest_week": busiest_week,
                "top_employer": top_employer,
                "weeks_covered": len(data),
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
                "employer_id": employer_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Weekly Dispatch Summary",
        }

    def render_weekly_dispatch_summary_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for weekly dispatch summary (landscape)."""
        data = self.get_weekly_dispatch_summary(start_date, end_date, book_id, employer_id)
        return self._render_pdf("weekly_dispatch_summary.html", data)

    def render_weekly_dispatch_summary_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for weekly dispatch summary."""
        data = self.get_weekly_dispatch_summary(start_date, end_date, book_id, employer_id)

        headers = [
            "Week",
            "Dispatches",
            "Short Calls",
            "Long Calls",
            "Avg Duration (Days)",
            "Top Book",
            "Top Employer",
        ]

        rows = [
            [
                r["week"],
                r["dispatches"],
                r["short_calls"],
                r["long_calls"],
                r["avg_duration_days"],
                r["top_book"],
                r["top_employer"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Weekly Summary", data["report_name"])

    # --- Report W37-2: Monthly Dispatch Summary ---

    def get_monthly_dispatch_summary(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Monthly dispatch summary with trend indicators.

        Args:
            start_date: Start of period. Defaults to 12 months ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = date(end_date.year - 1, end_date.month, 1)

        # Query dispatches
        query = (
            self.db.query(Dispatch)
            .join(BookRegistration, Dispatch.registration_id == BookRegistration.id)
            .filter(
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
            )
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        dispatches = query.all()

        # Aggregate by month
        monthly_data = {}
        for d in dispatches:
            dispatch_date = (
                d.dispatch_date.date()
                if isinstance(d.dispatch_date, datetime)
                else d.dispatch_date
            )
            month_key = dispatch_date.strftime("%Y-%m")

            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    "month": month_key,
                    "total": 0,
                    "short_calls": 0,
                    "by_agreement": {"STANDARD": 0, "PLA": 0, "CWA": 0, "TERO": 0},
                }

            monthly_data[month_key]["total"] += 1

            # Count short calls
            if d.end_date:
                end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
                duration = (end - dispatch_date).days
                if duration <= 10:
                    monthly_data[month_key]["short_calls"] += 1

            # Count by agreement type
            agreement = getattr(d, "agreement_type", "STANDARD") or "STANDARD"
            if agreement in monthly_data[month_key]["by_agreement"]:
                monthly_data[month_key]["by_agreement"][agreement] += 1
            else:
                monthly_data[month_key]["by_agreement"]["STANDARD"] += 1

        # Build data with trend indicators
        data = []
        sorted_months = sorted(monthly_data.keys())
        prev_total = None

        for month_key in sorted_months:
            row = monthly_data[month_key]
            short_pct = (
                round(100 * row["short_calls"] / row["total"], 1)
                if row["total"] > 0
                else 0
            )

            # Calculate trend
            if prev_total is not None:
                change_pct = (
                    (row["total"] - prev_total) / prev_total * 100
                    if prev_total > 0
                    else 0
                )
                if change_pct > 5:
                    trend = "↑"
                elif change_pct < -5:
                    trend = "↓"
                else:
                    trend = "→"
            else:
                trend = "→"

            data.append({
                "month": row["month"],
                "total_dispatches": row["total"],
                "short_calls_pct": short_pct,
                "standard": row["by_agreement"]["STANDARD"],
                "pla": row["by_agreement"]["PLA"],
                "cwa": row["by_agreement"]["CWA"],
                "tero": row["by_agreement"]["TERO"],
                "trend": trend,
            })
            prev_total = row["total"]

        # Summary
        if data:
            totals = [d["total_dispatches"] for d in data]
            peak_month = max(data, key=lambda x: x["total_dispatches"])["month"]
            trough_month = min(data, key=lambda x: x["total_dispatches"])["month"]
            total_all = sum(totals)
            total_pla = sum(d["pla"] for d in data)
            total_cwa = sum(d["cwa"] for d in data)
        else:
            peak_month = "N/A"
            trough_month = "N/A"
            total_all = 0
            total_pla = 0
            total_cwa = 0

        return {
            "data": data,
            "summary": {
                "total_dispatches": total_all,
                "peak_month": peak_month,
                "trough_month": trough_month,
                "pla_cwa_share_pct": round(100 * (total_pla + total_cwa) / total_all, 1) if total_all > 0 else 0,
                "months_covered": len(data),
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Monthly Dispatch Summary",
        }

    def render_monthly_dispatch_summary_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for monthly dispatch summary."""
        data = self.get_monthly_dispatch_summary(start_date, end_date, book_id)
        return self._render_pdf("monthly_dispatch_summary.html", data)

    def render_monthly_dispatch_summary_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for monthly dispatch summary."""
        data = self.get_monthly_dispatch_summary(start_date, end_date, book_id)

        headers = [
            "Month",
            "Total Dispatches",
            "Short Calls %",
            "Standard",
            "PLA",
            "CWA",
            "TERO",
            "Trend",
        ]

        rows = [
            [
                r["month"],
                r["total_dispatches"],
                r["short_calls_pct"],
                r["standard"],
                r["pla"],
                r["cwa"],
                r["tero"],
                r["trend"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Monthly Summary", data["report_name"])

    # --- Report W37-3: Dispatch by Agreement Type ---

    def get_dispatch_by_agreement_type(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Breakdown of dispatches by agreement type (PLA, CWA, TERO, Standard).

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Query dispatches
        query = (
            self.db.query(Dispatch)
            .join(BookRegistration, Dispatch.registration_id == BookRegistration.id)
            .filter(
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
            )
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        dispatches = query.all()

        # Aggregate by agreement type
        agreement_stats = {}
        members_by_agreement = {}
        employers_by_agreement = {}

        for d in dispatches:
            agreement = getattr(d, "agreement_type", "STANDARD") or "STANDARD"

            if agreement not in agreement_stats:
                agreement_stats[agreement] = {
                    "type": agreement,
                    "count": 0,
                    "total_days": 0,
                    "completed": 0,
                }
                members_by_agreement[agreement] = set()
                employers_by_agreement[agreement] = set()

            agreement_stats[agreement]["count"] += 1
            members_by_agreement[agreement].add(d.member_id)
            employers_by_agreement[agreement].add(d.employer_id)

            # Calculate duration
            if d.end_date:
                dispatch_date = (
                    d.dispatch_date.date()
                    if isinstance(d.dispatch_date, datetime)
                    else d.dispatch_date
                )
                end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
                agreement_stats[agreement]["total_days"] += (end - dispatch_date).days
                agreement_stats[agreement]["completed"] += 1

        # Build data
        data = []
        total_dispatches = sum(s["count"] for s in agreement_stats.values())

        for agreement, stats in sorted(agreement_stats.items()):
            avg_duration = (
                round(stats["total_days"] / stats["completed"], 1)
                if stats["completed"] > 0
                else 0
            )
            fill_rate = (
                round(100 * stats["completed"] / stats["count"], 1)
                if stats["count"] > 0
                else 0
            )

            data.append({
                "agreement_type": agreement,
                "dispatch_count": stats["count"],
                "avg_duration_days": avg_duration,
                "fill_rate_pct": fill_rate,
                "unique_employers": len(employers_by_agreement.get(agreement, set())),
                "unique_members": len(members_by_agreement.get(agreement, set())),
            })

        # Sort by count descending
        data.sort(key=lambda x: x["dispatch_count"], reverse=True)

        # Summary
        dominant = data[0]["agreement_type"] if data else "N/A"
        pla_cwa_count = sum(d["dispatch_count"] for d in data if d["agreement_type"] in ["PLA", "CWA"])

        return {
            "data": data,
            "summary": {
                "total_dispatches": total_dispatches,
                "dominant_type": dominant,
                "pla_cwa_share_pct": round(100 * pla_cwa_count / total_dispatches, 1) if total_dispatches > 0 else 0,
                "types_count": len(data),
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Dispatch by Agreement Type",
        }

    def render_dispatch_by_agreement_type_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for dispatch by agreement type."""
        data = self.get_dispatch_by_agreement_type(start_date, end_date, book_id)
        return self._render_pdf("dispatch_by_agreement_type.html", data)

    def render_dispatch_by_agreement_type_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for dispatch by agreement type."""
        data = self.get_dispatch_by_agreement_type(start_date, end_date, book_id)

        headers = [
            "Agreement Type",
            "Dispatch Count",
            "Avg Duration (Days)",
            "Fill Rate %",
            "Unique Employers",
            "Unique Members",
        ]

        rows = [
            [
                r["agreement_type"],
                r["dispatch_count"],
                r["avg_duration_days"],
                r["fill_rate_pct"],
                r["unique_employers"],
                r["unique_members"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "By Agreement", data["report_name"])

    # --- Report W37-4: Dispatch Duration Analysis ---

    def get_dispatch_duration_analysis(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
        group_by: str = "book",
    ) -> dict:
        """Dispatch duration analysis by book, employer, or classification.

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.
            employer_id: Filter to specific employer. None = all employers.
            group_by: Grouping (book, employer, or classification). Default: book.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization
        import statistics

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Query completed dispatches (with end_date)
        query = (
            self.db.query(Dispatch)
            .join(BookRegistration, Dispatch.registration_id == BookRegistration.id)
            .join(ReferralBook, BookRegistration.book_id == ReferralBook.id)
            .filter(
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
                Dispatch.end_date.isnot(None),
            )
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)
        if employer_id:
            query = query.filter(Dispatch.employer_id == employer_id)

        dispatches = query.all()

        # Get lookup maps
        books = {b.id: b.name for b in self.db.query(ReferralBook).all()}
        employers = {
            o.id: o.name
            for o in self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER).all()
        }
        members = {
            m.id: m.classification.value if m.classification else "Unknown"
            for m in self.db.query(Member).all()
        }

        # Aggregate by group
        duration_data = {}

        for d in dispatches:
            dispatch_date = (
                d.dispatch_date.date()
                if isinstance(d.dispatch_date, datetime)
                else d.dispatch_date
            )
            end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
            duration = (end - dispatch_date).days

            # Determine grouping key
            reg = self.db.query(BookRegistration).filter(
                BookRegistration.id == d.registration_id
            ).first()

            if group_by == "book":
                key = books.get(reg.book_id if reg else None, "Unknown")
            elif group_by == "employer":
                key = employers.get(d.employer_id, f"Employer {d.employer_id}")
            else:  # classification
                key = members.get(d.member_id, "Unknown")

            if key not in duration_data:
                duration_data[key] = []
            duration_data[key].append(duration)

        # Build data
        data = []
        all_durations = []

        for key, durations in sorted(duration_data.items()):
            all_durations.extend(durations)
            avg_days = round(statistics.mean(durations), 1) if durations else 0
            median_days = round(statistics.median(durations), 1) if durations else 0
            std_dev = round(statistics.stdev(durations), 1) if len(durations) > 1 else 0
            short_calls = sum(1 for d in durations if d <= 10)

            data.append({
                "group": key,
                "avg_days": avg_days,
                "median_days": median_days,
                "min_days": min(durations) if durations else 0,
                "max_days": max(durations) if durations else 0,
                "std_dev": std_dev,
                "count": len(durations),
                "short_calls": short_calls,
            })

        # Sort by avg duration descending
        data.sort(key=lambda x: x["avg_days"], reverse=True)

        # Summary
        overall_median = round(statistics.median(all_durations), 1) if all_durations else 0
        longest_avg = data[0]["group"] if data else "N/A"
        shortest_avg = data[-1]["group"] if data else "N/A"

        return {
            "data": data,
            "summary": {
                "overall_median_days": overall_median,
                "longest_avg_group": longest_avg,
                "shortest_avg_group": shortest_avg,
                "total_dispatches": len(all_durations),
                "group_by": group_by,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
                "employer_id": employer_id,
                "group_by": group_by,
            },
            "generated_at": datetime.now(),
            "report_name": "Dispatch Duration Analysis",
        }

    def render_dispatch_duration_analysis_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
        group_by: str = "book",
    ) -> BytesIO:
        """Render PDF for dispatch duration analysis."""
        data = self.get_dispatch_duration_analysis(
            start_date, end_date, book_id, employer_id, group_by
        )
        return self._render_pdf("dispatch_duration_analysis.html", data)

    def render_dispatch_duration_analysis_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
        group_by: str = "book",
    ) -> BytesIO:
        """Render Excel for dispatch duration analysis."""
        data = self.get_dispatch_duration_analysis(
            start_date, end_date, book_id, employer_id, group_by
        )

        headers = [
            "Group",
            "Avg Days",
            "Median Days",
            "Min",
            "Max",
            "Std Dev",
            "Count",
            "Short Calls (≤10d)",
        ]

        rows = [
            [
                r["group"],
                r["avg_days"],
                r["median_days"],
                r["min_days"],
                r["max_days"],
                r["std_dev"],
                r["count"],
                r["short_calls"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Duration Analysis", data["report_name"])

    # --- Report W37-5: Short Call Analysis ---

    def get_short_call_analysis(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
    ) -> dict:
        """Short call frequency, average duration, and re-registration patterns.

        Business rule: ≤3 days = treated as Long Call (Rule 9).
        Max 2 short call dispatches per registration cycle.

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.
            employer_id: Filter to specific employer. None = all employers.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization
        from src.models.registration_activity import RegistrationActivity

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Query completed dispatches
        query = (
            self.db.query(Dispatch)
            .join(BookRegistration, Dispatch.registration_id == BookRegistration.id)
            .filter(
                func.date(Dispatch.dispatch_date) >= start_date,
                func.date(Dispatch.dispatch_date) <= end_date,
                Dispatch.end_date.isnot(None),
            )
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)
        if employer_id:
            query = query.filter(Dispatch.employer_id == employer_id)

        dispatches = query.all()

        # Get employer lookup
        employers = {
            o.id: o.name
            for o in self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER).all()
        }

        # Analyze short calls
        period_data = {}
        short_call_members = {}
        employer_short_calls = {}
        total_dispatches = len(dispatches)
        short_calls_count = 0
        long_call_rule_count = 0
        total_short_call_days = 0

        for d in dispatches:
            dispatch_date = (
                d.dispatch_date.date()
                if isinstance(d.dispatch_date, datetime)
                else d.dispatch_date
            )
            end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
            duration = (end - dispatch_date).days

            # Weekly period
            week_start = dispatch_date - timedelta(days=dispatch_date.weekday())
            period_key = week_start.isoformat()

            if period_key not in period_data:
                period_data[period_key] = {
                    "period": period_key,
                    "short_calls": 0,
                    "long_call_rule": 0,
                    "total_days": 0,
                    "re_registrations": 0,
                }

            if duration <= 10:
                short_calls_count += 1
                period_data[period_key]["short_calls"] += 1
                period_data[period_key]["total_days"] += duration
                total_short_call_days += duration

                # Track by member
                short_call_members[d.member_id] = short_call_members.get(d.member_id, 0) + 1

                # Track by employer
                emp_name = employers.get(d.employer_id, f"Employer {d.employer_id}")
                employer_short_calls[emp_name] = employer_short_calls.get(emp_name, 0) + 1

                if duration <= 3:
                    long_call_rule_count += 1
                    period_data[period_key]["long_call_rule"] += 1

        # Count re-registrations after short calls (simplified)
        # Look for REGISTER activities in the period
        re_regs = (
            self.db.query(RegistrationActivity)
            .filter(
                func.date(RegistrationActivity.activity_date) >= start_date,
                func.date(RegistrationActivity.activity_date) <= end_date,
            )
            .count()
        )

        # Build data
        data = []
        for period_key in sorted(period_data.keys()):
            row = period_data[period_key]
            avg_duration = (
                round(row["total_days"] / row["short_calls"], 1)
                if row["short_calls"] > 0
                else 0
            )
            data.append({
                "period": row["period"],
                "short_calls": row["short_calls"],
                "avg_duration_days": avg_duration,
                "long_call_rule": row["long_call_rule"],
            })

        # Count max 2 per cycle violations (members with >2 short calls)
        violations = sum(1 for count in short_call_members.values() if count > 2)

        # Top employers using short calls
        top_employers = sorted(
            employer_short_calls.items(), key=lambda x: x[1], reverse=True
        )[:5]

        # Members with multiple short calls
        members_multiple = sum(1 for count in short_call_members.values() if count > 1)

        return {
            "data": data,
            "summary": {
                "total_short_calls": short_calls_count,
                "short_call_rate_pct": round(100 * short_calls_count / total_dispatches, 1) if total_dispatches > 0 else 0,
                "avg_short_call_duration": round(total_short_call_days / short_calls_count, 1) if short_calls_count > 0 else 0,
                "long_call_rule_count": long_call_rule_count,
                "max_2_violations": violations,
                "top_short_call_employers": [{"employer": e[0], "count": e[1]} for e in top_employers],
                "members_with_multiple": members_multiple,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
                "employer_id": employer_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Short Call Analysis",
        }

    def render_short_call_analysis_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for short call analysis."""
        data = self.get_short_call_analysis(start_date, end_date, book_id, employer_id)
        return self._render_pdf("short_call_analysis.html", data)

    def render_short_call_analysis_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        employer_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for short call analysis."""
        data = self.get_short_call_analysis(start_date, end_date, book_id, employer_id)

        headers = [
            "Period",
            "Short Calls",
            "Avg Duration (Days)",
            "≤3 Days (Long Call Rule)",
        ]

        rows = [
            [
                r["period"],
                r["short_calls"],
                r["avg_duration_days"],
                r["long_call_rule"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Short Calls", data["report_name"])

    # --- Report W37-6: Employer Utilization Report ---

    def get_employer_utilization(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
        contract_code: Optional[str] = None,
    ) -> dict:
        """Employer utilization — requests vs dispatches, fill rates.

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            employer_id: Filter to specific employer. None = all employers.
            contract_code: Filter by contract code. None = all.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Get employers
        employers_query = self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER)
        if employer_id:
            employers_query = employers_query.filter(Organization.id == employer_id)
        employers = {e.id: e for e in employers_query.all()}

        # Query labor requests
        request_query = self.db.query(LaborRequest).filter(
            func.date(LaborRequest.request_date) >= start_date,
            func.date(LaborRequest.request_date) <= end_date,
        )
        if employer_id:
            request_query = request_query.filter(LaborRequest.employer_id == employer_id)
        requests = request_query.all()

        # Query dispatches
        dispatch_query = self.db.query(Dispatch).filter(
            func.date(Dispatch.dispatch_date) >= start_date,
            func.date(Dispatch.dispatch_date) <= end_date,
        )
        if employer_id:
            dispatch_query = dispatch_query.filter(Dispatch.employer_id == employer_id)
        dispatches = dispatch_query.all()

        # Aggregate by employer
        employer_data = {}

        for r in requests:
            emp_id = r.employer_id
            if emp_id not in employer_data:
                employer_data[emp_id] = {
                    "employer_id": emp_id,
                    "requests_made": 0,
                    "workers_requested": 0,
                    "workers_dispatched": 0,
                    "cancellations": 0,
                    "fill_days": [],
                }
            employer_data[emp_id]["requests_made"] += 1
            employer_data[emp_id]["workers_requested"] += getattr(r, "workers_needed", 1) or 1
            if r.status == LaborRequestStatus.CANCELLED:
                employer_data[emp_id]["cancellations"] += 1

        for d in dispatches:
            emp_id = d.employer_id
            if emp_id not in employer_data:
                employer_data[emp_id] = {
                    "employer_id": emp_id,
                    "requests_made": 0,
                    "workers_requested": 0,
                    "workers_dispatched": 0,
                    "cancellations": 0,
                    "fill_days": [],
                }
            employer_data[emp_id]["workers_dispatched"] += 1

        # Build data
        data = []
        for emp_id, stats in employer_data.items():
            emp = employers.get(emp_id)
            emp_name = emp.name if emp else f"Employer {emp_id}"

            fill_rate = (
                round(100 * stats["workers_dispatched"] / stats["workers_requested"], 1)
                if stats["workers_requested"] > 0
                else 0
            )
            avg_fill_time = 0  # Simplified - would need request-to-dispatch matching

            data.append({
                "employer": emp_name,
                "requests_made": stats["requests_made"],
                "workers_requested": stats["workers_requested"],
                "workers_dispatched": stats["workers_dispatched"],
                "fill_rate_pct": fill_rate,
                "avg_fill_time_days": avg_fill_time,
                "cancellations": stats["cancellations"],
            })

        # Sort by dispatched volume descending
        data.sort(key=lambda x: x["workers_dispatched"], reverse=True)

        # Summary
        total_unfilled = sum(
            d["workers_requested"] - d["workers_dispatched"]
            for d in data
            if d["workers_requested"] > d["workers_dispatched"]
        )
        low_fill = [d for d in data if d["fill_rate_pct"] < 70]

        return {
            "data": data,
            "summary": {
                "total_employers": len(data),
                "total_requests": sum(d["requests_made"] for d in data),
                "total_dispatched": sum(d["workers_dispatched"] for d in data),
                "total_unfilled": total_unfilled,
                "low_fill_rate_count": len(low_fill),
                "top_10_volume": data[:10] if len(data) >= 10 else data,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "employer_id": employer_id,
                "contract_code": contract_code,
            },
            "generated_at": datetime.now(),
            "report_name": "Employer Utilization Report",
        }

    def render_employer_utilization_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
        contract_code: Optional[str] = None,
    ) -> BytesIO:
        """Render PDF for employer utilization."""
        data = self.get_employer_utilization(start_date, end_date, employer_id, contract_code)
        return self._render_pdf("employer_utilization.html", data)

    def render_employer_utilization_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
        contract_code: Optional[str] = None,
    ) -> BytesIO:
        """Render Excel for employer utilization."""
        data = self.get_employer_utilization(start_date, end_date, employer_id, contract_code)

        headers = [
            "Employer",
            "Requests Made",
            "Workers Requested",
            "Workers Dispatched",
            "Fill Rate %",
            "Avg Fill Time (Days)",
            "Cancellations",
        ]

        rows = [
            [
                r["employer"],
                r["requests_made"],
                r["workers_requested"],
                r["workers_dispatched"],
                r["fill_rate_pct"],
                r["avg_fill_time_days"],
                r["cancellations"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Utilization", data["report_name"])

    # --- Report W37-7: Employer Request Patterns ---

    def get_employer_request_patterns(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
        contract_code: Optional[str] = None,
    ) -> dict:
        """Employer request patterns — frequency, size, seasonal trends.

        Args:
            start_date: Start of period. Defaults to 6 months ago.
            end_date: End of period. Defaults to today.
            employer_id: Filter to specific employer. None = all employers.
            contract_code: Filter by contract code. None = all.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization
        from collections import Counter

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=180)

        # Get employers
        employers = {
            e.id: e.name
            for e in self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER).all()
        }

        # Query requests
        query = self.db.query(LaborRequest).filter(
            func.date(LaborRequest.request_date) >= start_date,
            func.date(LaborRequest.request_date) <= end_date,
        )
        if employer_id:
            query = query.filter(LaborRequest.employer_id == employer_id)
        requests = query.all()

        # Calculate months in range
        months_in_range = max(1, (end_date - start_date).days // 30)

        # Aggregate by employer
        employer_patterns = {}

        for r in requests:
            emp_id = r.employer_id
            if emp_id not in employer_patterns:
                employer_patterns[emp_id] = {
                    "requests": [],
                    "workers": [],
                    "months": Counter(),
                    "days": Counter(),
                }

            request_date = (
                r.request_date.date()
                if isinstance(r.request_date, datetime)
                else r.request_date
            )

            employer_patterns[emp_id]["requests"].append(r)
            employer_patterns[emp_id]["workers"].append(getattr(r, "workers_needed", 1) or 1)
            employer_patterns[emp_id]["months"][request_date.strftime("%B")] += 1
            employer_patterns[emp_id]["days"][request_date.strftime("%A")] += 1

        # Build data
        data = []
        for emp_id, patterns in employer_patterns.items():
            emp_name = employers.get(emp_id, f"Employer {emp_id}")
            requests_per_month = round(len(patterns["requests"]) / months_in_range, 1)
            avg_workers = round(sum(patterns["workers"]) / len(patterns["workers"]), 1) if patterns["workers"] else 0

            peak_month = patterns["months"].most_common(1)[0][0] if patterns["months"] else "N/A"
            trough_month = patterns["months"].most_common()[-1][0] if patterns["months"] else "N/A"
            peak_day = patterns["days"].most_common(1)[0][0] if patterns["days"] else "N/A"

            data.append({
                "employer": emp_name,
                "requests_per_month": requests_per_month,
                "avg_workers_per_request": avg_workers,
                "peak_month": peak_month,
                "trough_month": trough_month,
                "peak_day_of_week": peak_day,
                "total_requests": len(patterns["requests"]),
            })

        # Sort by requests per month descending
        data.sort(key=lambda x: x["requests_per_month"], reverse=True)

        # Summary
        most_active = data[:5] if len(data) >= 5 else data
        overall_peak = Counter()
        for emp_id, patterns in employer_patterns.items():
            overall_peak.update(patterns["months"])

        return {
            "data": data,
            "summary": {
                "total_employers": len(data),
                "total_requests": sum(d["total_requests"] for d in data),
                "most_active_employers": [d["employer"] for d in most_active],
                "overall_peak_month": overall_peak.most_common(1)[0][0] if overall_peak else "N/A",
                "months_analyzed": months_in_range,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "employer_id": employer_id,
                "contract_code": contract_code,
            },
            "generated_at": datetime.now(),
            "report_name": "Employer Request Patterns",
        }

    def render_employer_request_patterns_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
        contract_code: Optional[str] = None,
    ) -> BytesIO:
        """Render PDF for employer request patterns."""
        data = self.get_employer_request_patterns(start_date, end_date, employer_id, contract_code)
        return self._render_pdf("employer_request_patterns.html", data)

    def render_employer_request_patterns_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
        contract_code: Optional[str] = None,
    ) -> BytesIO:
        """Render Excel for employer request patterns."""
        data = self.get_employer_request_patterns(start_date, end_date, employer_id, contract_code)

        headers = [
            "Employer",
            "Requests/Month",
            "Avg Workers/Request",
            "Peak Month",
            "Trough Month",
            "Peak Day",
            "Total Requests",
        ]

        rows = [
            [
                r["employer"],
                r["requests_per_month"],
                r["avg_workers_per_request"],
                r["peak_month"],
                r["trough_month"],
                r["peak_day_of_week"],
                r["total_requests"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Patterns", data["report_name"])

    # --- Report W37-8: Top Employers Report ---

    def get_top_employers(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: int = 20,
        sort_by: str = "dispatches",
    ) -> dict:
        """Top employers by dispatch volume, request frequency, or fill rate.

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            limit: Number of employers to return. Default 20.
            sort_by: Sort metric (dispatches, requests, fill_rate). Default: dispatches.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Get employers
        employers = {
            e.id: e for e in self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER).all()
        }

        # Query data
        requests = self.db.query(LaborRequest).filter(
            func.date(LaborRequest.request_date) >= start_date,
            func.date(LaborRequest.request_date) <= end_date,
        ).all()

        dispatches = self.db.query(Dispatch).filter(
            func.date(Dispatch.dispatch_date) >= start_date,
            func.date(Dispatch.dispatch_date) <= end_date,
        ).all()

        # Aggregate
        employer_stats = {}

        for r in requests:
            emp_id = r.employer_id
            if emp_id not in employer_stats:
                employer_stats[emp_id] = {
                    "requests": 0,
                    "dispatches": 0,
                    "workers_requested": 0,
                    "durations": [],
                }
            employer_stats[emp_id]["requests"] += 1
            employer_stats[emp_id]["workers_requested"] += getattr(r, "workers_needed", 1) or 1

        for d in dispatches:
            emp_id = d.employer_id
            if emp_id not in employer_stats:
                employer_stats[emp_id] = {
                    "requests": 0,
                    "dispatches": 0,
                    "workers_requested": 0,
                    "durations": [],
                }
            employer_stats[emp_id]["dispatches"] += 1

            if d.end_date:
                dispatch_date = (
                    d.dispatch_date.date()
                    if isinstance(d.dispatch_date, datetime)
                    else d.dispatch_date
                )
                end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
                employer_stats[emp_id]["durations"].append((end - dispatch_date).days)

        # Build data
        data = []
        total_dispatches = sum(s["dispatches"] for s in employer_stats.values())

        for emp_id, stats in employer_stats.items():
            emp = employers.get(emp_id)
            emp_name = emp.name if emp else f"Employer {emp_id}"

            fill_rate = (
                round(100 * stats["dispatches"] / stats["workers_requested"], 1)
                if stats["workers_requested"] > 0
                else 0
            )
            avg_duration = (
                round(sum(stats["durations"]) / len(stats["durations"]), 1)
                if stats["durations"]
                else 0
            )

            data.append({
                "employer": emp_name,
                "total_dispatches": stats["dispatches"],
                "total_requests": stats["requests"],
                "fill_rate_pct": fill_rate,
                "avg_dispatch_duration": avg_duration,
            })

        # Sort
        if sort_by == "requests":
            data.sort(key=lambda x: x["total_requests"], reverse=True)
        elif sort_by == "fill_rate":
            data.sort(key=lambda x: x["fill_rate_pct"], reverse=True)
        else:
            data.sort(key=lambda x: x["total_dispatches"], reverse=True)

        # Apply limit and add rank
        data = data[:limit]
        for i, d in enumerate(data, 1):
            d["rank"] = i

        # Summary
        top_5_dispatches = sum(d["total_dispatches"] for d in data[:5])
        top_5_pct = round(100 * top_5_dispatches / total_dispatches, 1) if total_dispatches > 0 else 0

        return {
            "data": data,
            "summary": {
                "total_employers_ranked": len(data),
                "total_dispatches_all": total_dispatches,
                "top_5_share_pct": top_5_pct,
                "sort_by": sort_by,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "limit": limit,
                "sort_by": sort_by,
            },
            "generated_at": datetime.now(),
            "report_name": "Top Employers Report",
        }

    def render_top_employers_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: int = 20,
        sort_by: str = "dispatches",
    ) -> BytesIO:
        """Render PDF for top employers."""
        data = self.get_top_employers(start_date, end_date, limit, sort_by)
        return self._render_pdf("top_employers.html", data)

    def render_top_employers_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        limit: int = 20,
        sort_by: str = "dispatches",
    ) -> BytesIO:
        """Render Excel for top employers."""
        data = self.get_top_employers(start_date, end_date, limit, sort_by)

        headers = [
            "Rank",
            "Employer",
            "Total Dispatches",
            "Total Requests",
            "Fill Rate %",
            "Avg Dispatch Duration",
        ]

        rows = [
            [
                r["rank"],
                r["employer"],
                r["total_dispatches"],
                r["total_requests"],
                r["fill_rate_pct"],
                r["avg_dispatch_duration"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Top Employers", data["report_name"])

    # --- Report W37-9: Employer Compliance Report ---

    def get_employer_compliance(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
    ) -> dict:
        """Employer compliance — by-name requests, agreement adherence, blackout violations.

        Business rules: Rule 13 (Foreperson By Name anti-collusion).
                       Rule 12 (2-week blackout after quit/discharge).

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            employer_id: Filter to specific employer. None = all employers.

        Returns: dict with data, summary, filters, generated_at
        """
        from src.models.organization import Organization

        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Get employers
        employers = {
            e.id: e.name
            for e in self.db.query(Organization).filter(Organization.org_type == OrganizationType.EMPLOYER).all()
        }

        # Query labor requests
        query = self.db.query(LaborRequest).filter(
            func.date(LaborRequest.request_date) >= start_date,
            func.date(LaborRequest.request_date) <= end_date,
        )
        if employer_id:
            query = query.filter(LaborRequest.employer_id == employer_id)
        requests = query.all()

        # Aggregate by employer
        employer_compliance = {}

        for r in requests:
            emp_id = r.employer_id
            if emp_id not in employer_compliance:
                employer_compliance[emp_id] = {
                    "total_requests": 0,
                    "by_name_requests": 0,
                    "agreement_violations": 0,
                    "blackout_violations": 0,
                    "notes": [],
                }

            employer_compliance[emp_id]["total_requests"] += 1

            # Check for by-name (foreperson request)
            if getattr(r, "is_by_name", False) or getattr(r, "foreperson_requested", False):
                employer_compliance[emp_id]["by_name_requests"] += 1

        # Build data
        data = []
        flagged_employers = []

        for emp_id, stats in employer_compliance.items():
            emp_name = employers.get(emp_id, f"Employer {emp_id}")
            by_name_pct = (
                round(100 * stats["by_name_requests"] / stats["total_requests"], 1)
                if stats["total_requests"] > 0
                else 0
            )

            notes = []
            if by_name_pct > 50:
                notes.append("High by-name rate (>50%)")
                flagged_employers.append(emp_name)
            if stats["blackout_violations"] > 0:
                notes.append(f"{stats['blackout_violations']} blackout violation(s)")

            data.append({
                "employer": emp_name,
                "by_name_requests": stats["by_name_requests"],
                "by_name_pct": by_name_pct,
                "blackout_violations": stats["blackout_violations"],
                "agreement_violations": stats["agreement_violations"],
                "notes": "; ".join(notes) if notes else "None",
            })

        # Sort by by-name percentage descending
        data.sort(key=lambda x: x["by_name_pct"], reverse=True)

        # Summary
        total_by_name = sum(d["by_name_requests"] for d in data)
        total_requests = sum(
            employer_compliance[emp_id]["total_requests"]
            for emp_id in employer_compliance
        )
        total_blackout = sum(d["blackout_violations"] for d in data)

        return {
            "data": data,
            "summary": {
                "total_employers": len(data),
                "total_by_name_requests": total_by_name,
                "overall_by_name_pct": round(100 * total_by_name / total_requests, 1) if total_requests > 0 else 0,
                "employers_over_50_pct": len(flagged_employers),
                "total_blackout_violations": total_blackout,
                "flagged_for_review": flagged_employers[:10],
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "employer_id": employer_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Employer Compliance Report",
        }

    def render_employer_compliance_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for employer compliance (sensitive - Officer+ only)."""
        data = self.get_employer_compliance(start_date, end_date, employer_id)
        return self._render_pdf("employer_compliance.html", data)

    # --- Report W37-10: Member Dispatch Frequency ---

    def get_member_dispatch_frequency(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        min_dispatches: int = 0,
    ) -> dict:
        """Member dispatch frequency — dispatch count per member.

        Args:
            start_date: Start of period. Defaults to 90 days ago.
            end_date: End of period. Defaults to today.
            book_id: Filter to specific book. None = all books.
            min_dispatches: Minimum dispatch count threshold. Default 0.

        Returns: dict with data, summary, filters, generated_at
        """
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        # Query dispatches
        query = self.db.query(Dispatch).filter(
            func.date(Dispatch.dispatch_date) >= start_date,
            func.date(Dispatch.dispatch_date) <= end_date,
        )

        if book_id:
            query = (
                query.join(BookRegistration, Dispatch.registration_id == BookRegistration.id)
                .filter(BookRegistration.book_id == book_id)
            )

        dispatches = query.all()

        # Get member lookup
        members = {
            m.id: m for m in self.db.query(Member).all()
        }

        # Aggregate by member
        member_stats = {}

        for d in dispatches:
            member_id = d.member_id
            if member_id not in member_stats:
                member_stats[member_id] = {
                    "dispatches": 0,
                    "durations": [],
                    "employers": set(),
                    "short_calls": 0,
                    "last_dispatch": None,
                }

            member_stats[member_id]["dispatches"] += 1
            member_stats[member_id]["employers"].add(d.employer_id)

            dispatch_date = (
                d.dispatch_date.date()
                if isinstance(d.dispatch_date, datetime)
                else d.dispatch_date
            )

            if member_stats[member_id]["last_dispatch"] is None or dispatch_date > member_stats[member_id]["last_dispatch"]:
                member_stats[member_id]["last_dispatch"] = dispatch_date

            if d.end_date:
                end = d.end_date.date() if isinstance(d.end_date, datetime) else d.end_date
                duration = (end - dispatch_date).days
                member_stats[member_id]["durations"].append(duration)
                if duration <= 10:
                    member_stats[member_id]["short_calls"] += 1

        # Build data
        data = []

        for member_id, stats in member_stats.items():
            if stats["dispatches"] < min_dispatches:
                continue

            member = members.get(member_id)
            member_name = f"{member.first_name} {member.last_name}" if member else f"Member {member_id}"

            avg_days = (
                round(sum(stats["durations"]) / len(stats["durations"]), 1)
                if stats["durations"]
                else 0
            )
            short_call_pct = (
                round(100 * stats["short_calls"] / stats["dispatches"], 1)
                if stats["dispatches"] > 0
                else 0
            )

            data.append({
                "member": member_name,
                "member_id": member_id,
                "total_dispatches": stats["dispatches"],
                "avg_days_per_dispatch": avg_days,
                "unique_employers": len(stats["employers"]),
                "short_call_pct": short_call_pct,
                "last_dispatch_date": stats["last_dispatch"].isoformat() if stats["last_dispatch"] else "N/A",
            })

        # Sort by total dispatches descending
        data.sort(key=lambda x: x["total_dispatches"], reverse=True)

        # Summary
        dispatch_counts = [d["total_dispatches"] for d in data]
        avg_dispatches = round(sum(dispatch_counts) / len(dispatch_counts), 1) if dispatch_counts else 0

        # Distribution histogram data
        hist_bins = {
            "0": sum(1 for c in dispatch_counts if c == 0),
            "1-2": sum(1 for c in dispatch_counts if 1 <= c <= 2),
            "3-5": sum(1 for c in dispatch_counts if 3 <= c <= 5),
            "6-10": sum(1 for c in dispatch_counts if 6 <= c <= 10),
            "10+": sum(1 for c in dispatch_counts if c > 10),
        }

        # Get members with 0 dispatches but still registered
        zero_dispatch_members = (
            self.db.query(BookRegistration)
            .filter(
                BookRegistration.status == RegistrationStatus.REGISTERED,
                ~BookRegistration.member_id.in_(member_stats.keys()) if member_stats else True,
            )
            .count()
        )

        return {
            "data": data[:100],  # Limit for PDF
            "summary": {
                "total_members_dispatched": len(data),
                "most_dispatched": data[0]["member"] if data else "N/A",
                "avg_dispatches_per_member": avg_dispatches,
                "members_with_zero": zero_dispatch_members,
                "distribution": hist_bins,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
                "min_dispatches": min_dispatches,
            },
            "generated_at": datetime.now(),
            "report_name": "Member Dispatch Frequency",
        }

    def render_member_dispatch_frequency_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        min_dispatches: int = 0,
    ) -> BytesIO:
        """Render PDF for member dispatch frequency (Officer+ only)."""
        data = self.get_member_dispatch_frequency(start_date, end_date, book_id, min_dispatches)
        return self._render_pdf("member_dispatch_frequency.html", data)

    def render_member_dispatch_frequency_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
        min_dispatches: int = 0,
    ) -> BytesIO:
        """Render Excel for member dispatch frequency."""
        data = self.get_member_dispatch_frequency(start_date, end_date, book_id, min_dispatches)

        headers = [
            "Member",
            "Total Dispatches",
            "Avg Days/Dispatch",
            "Unique Employers",
            "Short Call %",
            "Last Dispatch",
        ]

        rows = [
            [
                r["member"],
                r["total_dispatches"],
                r["avg_days_per_dispatch"],
                r["unique_employers"],
                r["short_call_pct"],
                r["last_dispatch_date"],
            ]
            for r in data["data"]
        ]

        return self._render_excel(headers, rows, "Member Frequency", data["report_name"])

    # =========================================================================
    # WEEK 38 P1 REPORTS: Compliance, Operational & Cross-Book Analytics
    # =========================================================================

    # --- Report: Internet Bidding Activity ---

    def get_internet_bidding_activity(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        member_id: Optional[int] = None,
        status: Optional[str] = None,
    ) -> dict:
        """Assemble data for internet bidding activity report.
        
        Tracks web bid submissions, acceptances, rejections, and ban tracking.
        Business rule: Rule 8 — 5:30 PM to 7:00 AM window. 2nd rejection in 12 months = lose privileges 1 year.
        """
        if not start_date:
            start_date = date.today() - timedelta(days=90)
        if not end_date:
            end_date = date.today()

        query = (
            self.db.query(JobBid)
            .filter(
                JobBid.bid_submitted_at >= datetime.combine(start_date, time.min),
                JobBid.bid_submitted_at <= datetime.combine(end_date, time.max),
            )
        )

        if member_id:
            query = query.filter(JobBid.member_id == member_id)

        if status:
            status_map = {
                "accepted": BidStatus.ACCEPTED,
                "rejected": BidStatus.REJECTED,
                "pending": BidStatus.PENDING,
            }
            if status in status_map:
                query = query.filter(JobBid.bid_status == status_map[status])

        bids = query.order_by(JobBid.bid_submitted_at.desc()).all()

        # Build data with bidding window check
        data = []
        rejections_by_member = {}
        for bid in bids:
            bid_time = bid.bid_submitted_at
            # Bidding window: 5:30 PM to 7:00 AM (17:30 to 07:00)
            hour = bid_time.hour if bid_time else 12
            minute = bid_time.minute if bid_time else 0
            within_window = (hour >= 17 and minute >= 30) or hour >= 18 or hour < 7

            member = bid.member
            member_name = f"{member.last_name}, {member.first_name}" if member else "Unknown"
            member_id_val = bid.member_id

            # Track rejections for rolling 12-month count
            if bid.bid_status == BidStatus.REJECTED:
                if member_id_val not in rejections_by_member:
                    rejections_by_member[member_id_val] = []
                rejections_by_member[member_id_val].append(bid.bid_submitted_at)

            # Get rolling 12-month rejection count for this member
            cutoff = datetime.now() - timedelta(days=365)
            rejection_count = (
                self.db.query(func.count(JobBid.id))
                .filter(
                    JobBid.member_id == member_id_val,
                    JobBid.bid_status == BidStatus.REJECTED,
                    JobBid.bid_submitted_at >= cutoff,
                )
                .scalar() or 0
            )

            data.append({
                "date": bid_time.strftime("%m/%d/%Y %H:%M") if bid_time else "N/A",
                "member": member_name,
                "labor_request_id": bid.labor_request_id,
                "bid_time": bid_time.strftime("%H:%M") if bid_time else "N/A",
                "status": bid.bid_status.value if bid.bid_status else "unknown",
                "within_window": "Y" if within_window else "N",
                "rejection_count_12mo": rejection_count,
                "at_ban_threshold": rejection_count >= 2,
            })

        # Summary
        total_bids = len(data)
        accepted = sum(1 for d in data if d["status"] == "accepted")
        rejected = sum(1 for d in data if d["status"] == "rejected")
        pending = sum(1 for d in data if d["status"] == "pending")
        members_at_threshold = sum(1 for d in data if d["at_ban_threshold"])

        return {
            "data": data,
            "summary": {
                "total_bids": total_bids,
                "accepted": accepted,
                "rejected": rejected,
                "pending": pending,
                "acceptance_rate": round(accepted / total_bids * 100, 1) if total_bids else 0,
                "members_at_ban_threshold": members_at_threshold,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "member_id": member_id,
                "status": status,
            },
            "generated_at": datetime.now(),
            "report_name": "Internet Bidding Activity",
        }

    def render_internet_bidding_activity_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        member_id: Optional[int] = None,
        status: Optional[str] = None,
    ) -> BytesIO:
        """Render PDF for internet bidding activity."""
        data = self.get_internet_bidding_activity(start_date, end_date, member_id, status)
        return self._render_pdf("internet_bidding_activity.html", data)

    def render_internet_bidding_activity_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        member_id: Optional[int] = None,
        status: Optional[str] = None,
    ) -> BytesIO:
        """Render Excel for internet bidding activity."""
        data = self.get_internet_bidding_activity(start_date, end_date, member_id, status)

        headers = ["Date", "Member", "Labor Request ID", "Bid Time", "Status", "Within Window", "12mo Rejections"]
        rows = [
            [d["date"], d["member"], d["labor_request_id"], d["bid_time"], 
             d["status"], d["within_window"], d["rejection_count_12mo"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Bidding Activity", data["report_name"])

    # --- Report: Exempt Status Report ---

    def get_exempt_status_report(
        self,
        snapshot_date: Optional[date] = None,
        exempt_type: Optional[str] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for exempt status report.
        
        Members currently on exempt status by type.
        Business rule: Rule 14 — exempt members retain position but are not eligible for dispatch.
        """
        if not snapshot_date:
            snapshot_date = date.today()

        query = (
            self.db.query(BookRegistration)
            .filter(
                BookRegistration.is_exempt == True,
                BookRegistration.status == RegistrationStatus.REGISTERED,
            )
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        registrations = query.all()

        data = []
        exempt_type_counts = {}
        for reg in registrations:
            member = reg.member
            book = reg.book
            exempt_reason = reg.exempt_reason.value if reg.exempt_reason else "unknown"

            if exempt_type and exempt_reason != exempt_type:
                continue

            # Calculate days exempt
            days_exempt = 0
            if reg.exempt_start_date:
                days_exempt = (snapshot_date - reg.exempt_start_date).days

            # Count by type
            exempt_type_counts[exempt_reason] = exempt_type_counts.get(exempt_reason, 0) + 1

            data.append({
                "member": f"{member.last_name}, {member.first_name}" if member else "Unknown",
                "member_id": reg.member_id,
                "exempt_type": exempt_reason,
                "start_date": reg.exempt_start_date.strftime("%m/%d/%Y") if reg.exempt_start_date else "N/A",
                "expected_end_date": reg.exempt_end_date.strftime("%m/%d/%Y") if reg.exempt_end_date else "Indefinite",
                "days_exempt": days_exempt,
                "book": book.name if book else "Unknown",
            })

        # Find expirations in next 30 days
        upcoming_expirations = [
            d for d in data 
            if d["expected_end_date"] != "Indefinite"
            and datetime.strptime(d["expected_end_date"], "%m/%d/%Y").date() <= snapshot_date + timedelta(days=30)
        ]

        # Find longest current exemptions
        data.sort(key=lambda x: x["days_exempt"], reverse=True)

        return {
            "data": data,
            "summary": {
                "total_exempt": len(data),
                "by_type": exempt_type_counts,
                "longest_exemption_days": data[0]["days_exempt"] if data else 0,
                "upcoming_expirations_30d": len(upcoming_expirations),
            },
            "filters": {
                "snapshot_date": snapshot_date,
                "exempt_type": exempt_type,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Exempt Status Report",
        }

    def render_exempt_status_report_pdf(
        self,
        snapshot_date: Optional[date] = None,
        exempt_type: Optional[str] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for exempt status report."""
        data = self.get_exempt_status_report(snapshot_date, exempt_type, book_id)
        return self._render_pdf("exempt_status.html", data)

    def render_exempt_status_report_excel(
        self,
        snapshot_date: Optional[date] = None,
        exempt_type: Optional[str] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for exempt status report."""
        data = self.get_exempt_status_report(snapshot_date, exempt_type, book_id)

        headers = ["Member", "Exempt Type", "Start Date", "Expected End", "Days Exempt", "Book"]
        rows = [
            [d["member"], d["exempt_type"], d["start_date"], d["expected_end_date"], d["days_exempt"], d["book"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Exempt Status", data["report_name"])

    # --- Report: Penalty Report ---

    def get_penalty_report(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        penalty_type: Optional[str] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for penalty report.
        
        All penalty actions over period — check marks, bid rejections, quit/discharge roll-offs.
        """
        if not start_date:
            start_date = date.today() - timedelta(days=90)
        if not end_date:
            end_date = date.today()

        penalties = []

        # Check marks from registrations with check_marks > 0
        check_mark_query = (
            self.db.query(BookRegistration)
            .filter(
                BookRegistration.check_marks > 0,
                BookRegistration.registration_date >= start_date,
                BookRegistration.registration_date <= end_date,
            )
        )
        if book_id:
            check_mark_query = check_mark_query.filter(BookRegistration.book_id == book_id)

        for reg in check_mark_query.all():
            member = reg.member
            book = reg.book
            for i in range(reg.check_marks):
                penalties.append({
                    "date": reg.registration_date.strftime("%m/%d/%Y") if reg.registration_date else "N/A",
                    "member": f"{member.last_name}, {member.first_name}" if member else "Unknown",
                    "penalty_type": "check_mark",
                    "book": book.name if book else "Unknown",
                    "details": f"Check mark #{i+1}",
                    "resulting_action": "Roll-off" if i+1 >= 3 else "Warning",
                    "severity": 2 if i+1 >= 2 else 1,
                })

        # Bid rejections
        rejection_query = (
            self.db.query(JobBid)
            .filter(
                JobBid.bid_status == BidStatus.REJECTED,
                JobBid.bid_submitted_at >= datetime.combine(start_date, time.min),
                JobBid.bid_submitted_at <= datetime.combine(end_date, time.max),
            )
        )
        for bid in rejection_query.all():
            member = bid.member
            penalties.append({
                "date": bid.bid_submitted_at.strftime("%m/%d/%Y") if bid.bid_submitted_at else "N/A",
                "member": f"{member.last_name}, {member.first_name}" if member else "Unknown",
                "penalty_type": "bid_rejection",
                "book": "All Books",
                "details": f"Rejected bid for request #{bid.labor_request_id}",
                "resulting_action": "Warning (1-year ban if 2nd in 12mo)",
                "severity": 1,
            })

        # Quit/Discharge roll-offs from Dispatch terminations
        rolloff_query = (
            self.db.query(Dispatch)
            .filter(
                Dispatch.dispatch_status.in_([DispatchStatus.TERMINATED]),
                Dispatch.end_date >= start_date,
                Dispatch.end_date <= end_date,
            )
        )
        for dispatch in rolloff_query.all():
            member = dispatch.member
            book = dispatch.labor_request.book if dispatch.labor_request else None
            penalties.append({
                "date": dispatch.end_date.strftime("%m/%d/%Y") if dispatch.end_date else "N/A",
                "member": f"{member.last_name}, {member.first_name}" if member else "Unknown",
                "penalty_type": "quit_discharge" if dispatch.term_reason == TermReason.QUIT else "termination",
                "book": book.name if book else "All Books",
                "details": f"Dispatch ended: {dispatch.dispatch_status.value}",
                "resulting_action": "All-books rolloff + 2-week blackout",
                "severity": 3,
            })

        # Filter by penalty type if specified
        if penalty_type:
            penalties = [p for p in penalties if p["penalty_type"] == penalty_type]

        # Sort by date descending
        penalties.sort(key=lambda x: x["date"], reverse=True)

        # Summary
        type_counts = {}
        for p in penalties:
            type_counts[p["penalty_type"]] = type_counts.get(p["penalty_type"], 0) + 1

        book_counts = {}
        for p in penalties:
            book_counts[p["book"]] = book_counts.get(p["book"], 0) + 1

        most_penalized_book = max(book_counts.items(), key=lambda x: x[1])[0] if book_counts else "N/A"

        return {
            "data": penalties,
            "summary": {
                "total_penalties": len(penalties),
                "by_type": type_counts,
                "most_penalized_book": most_penalized_book,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "penalty_type": penalty_type,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Penalty Report",
        }

    def render_penalty_report_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        penalty_type: Optional[str] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for penalty report (Officer+ only)."""
        data = self.get_penalty_report(start_date, end_date, penalty_type, book_id)
        return self._render_pdf("penalty_report.html", data)

    def render_penalty_report_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        penalty_type: Optional[str] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for penalty report."""
        data = self.get_penalty_report(start_date, end_date, penalty_type, book_id)

        headers = ["Date", "Member", "Penalty Type", "Book", "Details", "Resulting Action"]
        rows = [
            [d["date"], d["member"], d["penalty_type"], d["book"], d["details"], d["resulting_action"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Penalties", data["report_name"])

    # --- Report: Foreperson By Name Audit ---

    def get_foreperson_by_name_audit(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for foreperson by-name audit report.
        
        By-name request tracking for anti-collusion compliance review.
        Business rules: Rule 13 — anti-collusion. Rule 12 — 2-week blackout after quit/discharge.
        SENSITIVE: Officer+ only.
        """
        if not start_date:
            start_date = date.today() - timedelta(days=90)
        if not end_date:
            end_date = date.today()

        # Query labor requests that were by-name
        query = (
            self.db.query(LaborRequest)
            .filter(
                LaborRequest.is_foreperson_by_name == True,
                LaborRequest.created_at >= datetime.combine(start_date, time.min),
                LaborRequest.created_at <= datetime.combine(end_date, time.max),
            )
        )

        if employer_id:
            query = query.filter(LaborRequest.employer_id == employer_id)

        requests = query.order_by(LaborRequest.created_at.desc()).all()

        # Get total dispatches for percentage calculation
        total_dispatches = (
            self.db.query(func.count(Dispatch.id))
            .filter(
                Dispatch.start_date >= start_date,
                Dispatch.start_date <= end_date,
            )
            .scalar() or 0
        )

        data = []
        employer_byname_counts = {}
        for req in requests:
            employer = req.employer
            employer_name = employer.name if employer else "Unknown"
            employer_byname_counts[employer_name] = employer_byname_counts.get(employer_name, 0) + 1

            # Check for any dispatches related to this request
            dispatches = (
                self.db.query(Dispatch)
                .filter(Dispatch.labor_request_id == req.id)
                .all()
            )

            for dispatch in dispatches:
                member = dispatch.member
                data.append({
                    "date": req.created_at.strftime("%m/%d/%Y") if req.created_at else "N/A",
                    "employer": employer_name,
                    "foreperson": f"{req.foreperson.last_name}, {req.foreperson.first_name}" if req.foreperson else "Not specified",
                    "requested_member": f"{member.last_name}, {member.first_name}" if member else "Unknown",
                    "approved": "Y",  # If dispatch exists, it was approved
                    "blackout_active": "N",  # Would need to check blackout periods
                    "notes": req.comments or "",
                })

        by_name_count = len(requests)
        by_name_pct = round(by_name_count / total_dispatches * 100, 1) if total_dispatches else 0

        # Top employers by by-name rate
        top_employers = sorted(employer_byname_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        return {
            "data": data,
            "summary": {
                "total_by_name_requests": by_name_count,
                "by_name_percentage": by_name_pct,
                "total_dispatches": total_dispatches,
                "top_employers_by_name": top_employers,
                "blackout_violations": 0,  # Would need separate blackout tracking
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "employer_id": employer_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Foreperson By Name Audit",
        }

    def render_foreperson_by_name_audit_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        employer_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for foreperson by-name audit (Officer+ only, PDF only)."""
        data = self.get_foreperson_by_name_audit(start_date, end_date, employer_id)
        return self._render_pdf("foreperson_by_name_audit.html", data)

    # --- Report: Queue Wait Time Report ---

    def get_queue_wait_time_report(
        self,
        snapshot_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for queue wait time report.
        
        Average wait time by book — identifies bottlenecks and longest-waiting members.
        """
        if not snapshot_date:
            snapshot_date = date.today()

        query = (
            self.db.query(BookRegistration)
            .filter(BookRegistration.status == RegistrationStatus.REGISTERED)
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        registrations = query.all()

        # Group by book
        by_book = {}
        for reg in registrations:
            book = reg.book
            book_name = book.name if book else "Unknown"
            if book_name not in by_book:
                by_book[book_name] = {
                    "registrations": [],
                    "book_tier": book.book_number if book else 1,
                }
            by_book[book_name]["registrations"].append(reg)

        data = []
        for book_name, info in by_book.items():
            regs = info["registrations"]
            wait_days = []
            resign_due_count = 0
            
            for reg in regs:
                if reg.registration_date:
                    days = (snapshot_date - reg.registration_date).days
                    wait_days.append(days)
                
                # Check re-sign due
                if reg.last_re_sign_date:
                    days_since_resign = (snapshot_date - reg.last_re_sign_date).days
                    if days_since_resign >= 23:  # Within 7 days of 30-day deadline
                        resign_due_count += 1

            avg_wait = round(sum(wait_days) / len(wait_days), 1) if wait_days else 0
            median_wait = sorted(wait_days)[len(wait_days) // 2] if wait_days else 0
            max_wait = max(wait_days) if wait_days else 0

            # Find longest waiting member (anonymized for staff)
            longest_member = "N/A"
            for reg in regs:
                if reg.registration_date and (snapshot_date - reg.registration_date).days == max_wait:
                    longest_member = f"Member #{reg.member_id}"
                    break

            data.append({
                "book": book_name,
                "tier": info["book_tier"],
                "active_count": len(regs),
                "avg_wait_days": avg_wait,
                "median_wait": median_wait,
                "longest_wait_days": max_wait,
                "longest_wait_member": longest_member,
                "resign_due_count": resign_due_count,
            })

        # Sort by avg wait descending
        data.sort(key=lambda x: x["avg_wait_days"], reverse=True)

        # Summary
        all_waits = []
        for book_data in by_book.values():
            for reg in book_data["registrations"]:
                if reg.registration_date:
                    all_waits.append((snapshot_date - reg.registration_date).days)

        members_over_90 = sum(1 for w in all_waits if w > 90)
        resign_due_this_week = sum(d["resign_due_count"] for d in data)

        return {
            "data": data,
            "summary": {
                "longest_avg_wait_book": data[0]["book"] if data else "N/A",
                "members_waiting_over_90": members_over_90,
                "resign_due_this_week": resign_due_this_week,
            },
            "filters": {
                "snapshot_date": snapshot_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Queue Wait Time Report",
        }

    def render_queue_wait_time_report_pdf(
        self,
        snapshot_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for queue wait time report."""
        data = self.get_queue_wait_time_report(snapshot_date, book_id)
        return self._render_pdf("queue_wait_time.html", data)

    def render_queue_wait_time_report_excel(
        self,
        snapshot_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for queue wait time report."""
        data = self.get_queue_wait_time_report(snapshot_date, book_id)

        headers = ["Book", "Tier", "Active Count", "Avg Wait (Days)", "Median Wait", "Longest Wait", "Re-Sign Due"]
        rows = [
            [d["book"], d["tier"], d["active_count"], d["avg_wait_days"], 
             d["median_wait"], d["longest_wait_days"], d["resign_due_count"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Queue Wait Times", data["report_name"])

    # --- Report: Morning Referral History ---

    def get_morning_referral_history(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for morning referral history report.
        
        Historical log of morning referral processing — what was dispatched each morning.
        Business rule: Rule 2 — processing order: Wire 8:30 AM → S&C/Marine/Stock/LFM/Residential 9:00 AM → Tradeshow 9:30 AM.
        """
        if not start_date:
            start_date = date.today() - timedelta(days=14)
        if not end_date:
            end_date = date.today()

        # Query dispatches grouped by date
        query = (
            self.db.query(Dispatch)
            .filter(
                Dispatch.start_date >= start_date,
                Dispatch.start_date <= end_date,
            )
        )

        if book_id:
            query = query.filter(Dispatch.book_id == book_id)

        dispatches = query.all()

        # Group by date and book
        by_date = {}
        for dispatch in dispatches:
            dispatch_date = dispatch.start_date.strftime("%m/%d/%Y") if dispatch.start_date else "Unknown"
            if dispatch_date not in by_date:
                by_date[dispatch_date] = {"total": 0, "by_book": {}, "unfilled": 0}
            
            by_date[dispatch_date]["total"] += 1
            book = dispatch.labor_request.book if dispatch.labor_request else None
            book_name = book.name if book else "Unknown"
            by_date[dispatch_date]["by_book"][book_name] = by_date[dispatch_date]["by_book"].get(book_name, 0) + 1

        # Get unfilled requests per date
        for date_str in by_date.keys():
            try:
                d = datetime.strptime(date_str, "%m/%d/%Y").date()
                unfilled = (
                    self.db.query(func.count(LaborRequest.id))
                    .filter(
                        LaborRequest.created_at >= datetime.combine(d, time.min),
                        LaborRequest.created_at <= datetime.combine(d, time.max),
                        LaborRequest.status == LaborRequestStatus.OPEN,
                    )
                    .scalar() or 0
                )
                by_date[date_str]["unfilled"] = unfilled
            except ValueError:
                pass

        # Build data rows
        data = []
        for date_str, info in sorted(by_date.items(), key=lambda x: x[0], reverse=True):
            for book_name, count in info["by_book"].items():
                # Determine processing order slot
                slot = "Other"
                for i, name in enumerate(MORNING_PROCESSING_ORDER):
                    if name in book_name.upper():
                        if i < 3:
                            slot = "8:30 AM"
                        elif i < 8:
                            slot = "9:00 AM"
                        else:
                            slot = "9:30 AM"
                        break

                data.append({
                    "date": date_str,
                    "processing_time": slot,
                    "book": book_name,
                    "dispatched": count,
                    "unfilled": info["unfilled"],
                    "processing_order_slot": slot,
                })

        # Summary
        total_dispatches = sum(d["dispatched"] for d in data)
        days_count = len(by_date)
        avg_per_morning = round(total_dispatches / days_count, 1) if days_count else 0
        total_unfilled = sum(info["unfilled"] for info in by_date.values())

        # Day of week analysis
        dow_counts = {}
        for date_str in by_date.keys():
            try:
                d = datetime.strptime(date_str, "%m/%d/%Y").date()
                dow = d.strftime("%A")
                dow_counts[dow] = dow_counts.get(dow, 0) + by_date[date_str]["total"]
            except ValueError:
                pass

        busiest_dow = max(dow_counts.items(), key=lambda x: x[1])[0] if dow_counts else "N/A"

        return {
            "data": data,
            "summary": {
                "avg_dispatches_per_morning": avg_per_morning,
                "busiest_day_of_week": busiest_dow,
                "total_unfilled": total_unfilled,
                "unfilled_rate": round(total_unfilled / (total_dispatches + total_unfilled) * 100, 1) if total_dispatches else 0,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Morning Referral History",
        }

    def render_morning_referral_history_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for morning referral history."""
        data = self.get_morning_referral_history(start_date, end_date, book_id)
        return self._render_pdf("morning_referral_history.html", data)

    def render_morning_referral_history_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for morning referral history."""
        data = self.get_morning_referral_history(start_date, end_date, book_id)

        headers = ["Date", "Processing Time", "Book", "Dispatched", "Unfilled", "Time Slot"]
        rows = [
            [d["date"], d["processing_time"], d["book"], d["dispatched"], d["unfilled"], d["processing_order_slot"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Morning Referral History", data["report_name"])

    # --- Report: Unfilled Request Report ---

    def get_unfilled_request_report(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
        employer_id: Optional[int] = None,
        book_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for unfilled request report.
        
        Labor requests not fully filled — aging analysis and root cause.
        """
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        if not end_date:
            end_date = date.today()

        query = (
            self.db.query(LaborRequest)
            .filter(
                LaborRequest.created_at >= datetime.combine(start_date, time.min),
                LaborRequest.created_at <= datetime.combine(end_date, time.max),
            )
        )

        # Filter by fill status
        if status == "unfilled":
            query = query.filter(LaborRequest.workers_dispatched == 0)
        elif status == "partially_filled":
            query = query.filter(
                LaborRequest.workers_dispatched > 0,
                LaborRequest.workers_dispatched < LaborRequest.workers_requested,
            )
        else:
            # Include both unfilled and partially filled
            query = query.filter(LaborRequest.workers_dispatched < LaborRequest.workers_requested)

        if employer_id:
            query = query.filter(LaborRequest.employer_id == employer_id)

        requests = query.order_by(LaborRequest.created_at.desc()).all()

        data = []
        employer_shortfall = {}
        classification_shortfall = {}
        for req in requests:
            employer = req.employer
            employer_name = employer.name if employer else "Unknown"
            book = req.book
            book_name = book.name if book else "Unknown"

            shortfall = req.workers_requested - (req.workers_dispatched or 0)
            days_open = (date.today() - req.created_at.date()).days if req.created_at else 0

            employer_shortfall[employer_name] = employer_shortfall.get(employer_name, 0) + shortfall
            classification_shortfall[book_name] = classification_shortfall.get(book_name, 0) + shortfall

            data.append({
                "request_date": req.created_at.strftime("%m/%d/%Y") if req.created_at else "N/A",
                "employer": employer_name,
                "workers_requested": req.workers_requested,
                "workers_filled": req.workers_dispatched or 0,
                "shortfall": shortfall,
                "days_open": days_open,
                "book": book_name,
                "reason": req.comments or "Not documented",
            })

        # Summary
        total_shortfall = sum(d["shortfall"] for d in data)
        top_unfilled_employers = sorted(employer_shortfall.items(), key=lambda x: x[1], reverse=True)[:5]
        avg_age = round(sum(d["days_open"] for d in data) / len(data), 1) if data else 0

        return {
            "data": data,
            "summary": {
                "total_unfilled_positions": total_shortfall,
                "top_unfilled_employers": top_unfilled_employers,
                "avg_days_open": avg_age,
                "shortfall_by_classification": classification_shortfall,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "status": status,
                "employer_id": employer_id,
                "book_id": book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Unfilled Request Report",
        }

    def render_unfilled_request_report_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
        employer_id: Optional[int] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for unfilled request report."""
        data = self.get_unfilled_request_report(start_date, end_date, status, employer_id, book_id)
        return self._render_pdf("unfilled_requests.html", data)

    def render_unfilled_request_report_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        status: Optional[str] = None,
        employer_id: Optional[int] = None,
        book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for unfilled request report."""
        data = self.get_unfilled_request_report(start_date, end_date, status, employer_id, book_id)

        headers = ["Request Date", "Employer", "Requested", "Filled", "Shortfall", "Days Open", "Book", "Reason"]
        rows = [
            [d["request_date"], d["employer"], d["workers_requested"], d["workers_filled"],
             d["shortfall"], d["days_open"], d["book"], d["reason"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Unfilled Requests", data["report_name"])

    # --- Report: Referral Agent Activity ---

    def get_referral_agent_activity(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        staff_member_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for referral agent activity report.
        
        Dispatches processed per staff member — workload distribution.
        Access: Officer+ (performance data).
        """
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        if not end_date:
            end_date = date.today()

        # Query dispatches grouped by dispatched_by_id
        from src.models.user import User
        
        query = (
            self.db.query(Dispatch)
            .filter(
                Dispatch.start_date >= start_date,
                Dispatch.start_date <= end_date,
            )
        )

        if staff_member_id:
            query = query.filter(Dispatch.dispatched_by_id == staff_member_id)

        dispatches = query.all()

        # Group by staff member
        by_staff = {}
        for dispatch in dispatches:
            staff_id = dispatch.dispatched_by_id
            if staff_id not in by_staff:
                by_staff[staff_id] = {
                    "dispatches": 0,
                    "check_marks": 0,
                    "by_name_approvals": 0,
                }
            by_staff[staff_id]["dispatches"] += 1
            # Count check marks issued - would need to track check_mark issuer
            # Count by-name approvals
            if dispatch.labor_request and dispatch.labor_request.is_foreperson_by_name:
                by_staff[staff_id]["by_name_approvals"] += 1

        # Build data with staff names
        data = []
        for staff_id, stats in by_staff.items():
            user = self.db.query(User).filter(User.id == staff_id).first()
            staff_name = f"{user.first_name} {user.last_name}" if user else f"Staff #{staff_id}"

            data.append({
                "staff_member": staff_name,
                "dispatches_processed": stats["dispatches"],
                "registrations_processed": 0,  # Would need registration processing tracking
                "check_marks_issued": stats["check_marks"],
                "by_name_approvals": stats["by_name_approvals"],
                "hours_active": "N/A",  # Would need session tracking
            })

        # Sort by dispatches descending
        data.sort(key=lambda x: x["dispatches_processed"], reverse=True)

        # Summary
        total_processed = sum(d["dispatches_processed"] for d in data)
        agent_count = len(data)
        avg_per_agent = round(total_processed / agent_count, 1) if agent_count else 0
        busiest = data[0]["staff_member"] if data else "N/A"
        quietest = data[-1]["staff_member"] if data else "N/A"

        return {
            "data": data,
            "summary": {
                "total_processed": total_processed,
                "agent_count": agent_count,
                "avg_per_agent": avg_per_agent,
                "busiest_agent": busiest,
                "quietest_agent": quietest,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "staff_member_id": staff_member_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Referral Agent Activity",
        }

    def render_referral_agent_activity_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        staff_member_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for referral agent activity (Officer+ only)."""
        data = self.get_referral_agent_activity(start_date, end_date, staff_member_id)
        return self._render_pdf("referral_agent_activity.html", data)

    # --- Report: Multi-Book Members ---

    def get_multi_book_members(
        self,
        snapshot_date: Optional[date] = None,
        min_books: int = 2,
    ) -> dict:
        """Assemble data for multi-book members report.
        
        Members registered on multiple books simultaneously — validates cross-classification rules.
        Business rule: Rule 5 — one registration per classification. Multiple classifications allowed.
        CRITICAL: Display APN as DECIMAL(10,2) — never truncate.
        """
        if not snapshot_date:
            snapshot_date = date.today()

        # Query members with multiple active registrations
        from sqlalchemy import and_

        subquery = (
            self.db.query(
                BookRegistration.member_id,
                func.count(BookRegistration.id).label("book_count")
            )
            .filter(BookRegistration.status == RegistrationStatus.REGISTERED)
            .group_by(BookRegistration.member_id)
            .having(func.count(BookRegistration.id) >= min_books)
            .subquery()
        )

        members_with_multi = (
            self.db.query(Member)
            .join(subquery, Member.id == subquery.c.member_id)
            .all()
        )

        data = []
        book_combo_counts = {}
        for member in members_with_multi:
            regs = [
                r for r in member.book_registrations
                if r.status == RegistrationStatus.REGISTERED
            ]

            book_info = []
            book_names = []
            for reg in sorted(regs, key=lambda r: r.registration_number or 0):
                book = reg.book
                book_names.append(book.name if book else "Unknown")
                book_info.append({
                    "book": book.name if book else "Unknown",
                    "apn": f"{reg.registration_number:.2f}" if reg.registration_number else "N/A",
                    "tier": book.book_number if book else 1,
                    "reg_date": reg.registration_date.strftime("%m/%d/%Y") if reg.registration_date else "N/A",
                })

            # Count book combinations
            combo = " + ".join(sorted(book_names))
            book_combo_counts[combo] = book_combo_counts.get(combo, 0) + 1

            data.append({
                "member": f"{member.last_name}, {member.first_name}",
                "member_id": member.id,
                "total_books": len(regs),
                "books": book_info,
            })

        # Sort by book count descending
        data.sort(key=lambda x: x["total_books"], reverse=True)

        # Summary
        two_books = sum(1 for d in data if d["total_books"] == 2)
        three_books = sum(1 for d in data if d["total_books"] == 3)
        four_plus = sum(1 for d in data if d["total_books"] >= 4)
        most_common_combos = sorted(book_combo_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        return {
            "data": data,
            "summary": {
                "members_on_2_books": two_books,
                "members_on_3_books": three_books,
                "members_on_4_plus_books": four_plus,
                "most_common_combinations": most_common_combos,
            },
            "filters": {
                "snapshot_date": snapshot_date,
                "min_books": min_books,
            },
            "generated_at": datetime.now(),
            "report_name": "Multi-Book Members",
        }

    def render_multi_book_members_pdf(
        self,
        snapshot_date: Optional[date] = None,
        min_books: int = 2,
    ) -> BytesIO:
        """Render PDF for multi-book members."""
        data = self.get_multi_book_members(snapshot_date, min_books)
        return self._render_pdf("multi_book_members.html", data)

    def render_multi_book_members_excel(
        self,
        snapshot_date: Optional[date] = None,
        min_books: int = 2,
    ) -> BytesIO:
        """Render Excel for multi-book members."""
        data = self.get_multi_book_members(snapshot_date, min_books)

        headers = ["Member", "Total Books", "Book 1", "APN 1", "Book 2", "APN 2", "Book 3", "APN 3"]
        rows = []
        for d in data["data"]:
            row = [d["member"], d["total_books"]]
            for i in range(3):
                if i < len(d["books"]):
                    row.extend([d["books"][i]["book"], d["books"][i]["apn"]])
                else:
                    row.extend(["", ""])
            rows.append(row)

        return self._render_excel(headers, rows, "Multi-Book Members", data["report_name"])

    # --- Report: Book Transfer Report ---

    def get_book_transfer_report(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        source_book_id: Optional[int] = None,
        destination_book_id: Optional[int] = None,
    ) -> dict:
        """Assemble data for book transfer report.
        
        Members who moved between books — re-registration patterns after dispatch/drop.
        """
        if not start_date:
            start_date = date.today() - timedelta(days=90)
        if not end_date:
            end_date = date.today()

        # This would need RegistrationActivity to track book changes
        # For now, we can look at members with rolled-off status on one book
        # and active on another

        rolled_off = (
            self.db.query(BookRegistration)
            .filter(
                BookRegistration.status.in_([
                    RegistrationStatus.ROLLED_OFF,
                    RegistrationStatus.DISPATCHED,
                ]),
                BookRegistration.registration_date >= start_date,
                BookRegistration.registration_date <= end_date,
            )
        )

        if source_book_id:
            rolled_off = rolled_off.filter(BookRegistration.book_id == source_book_id)

        rolled_off_regs = rolled_off.all()

        data = []
        transfer_counts = {}
        for old_reg in rolled_off_regs:
            member = old_reg.member
            if not member:
                continue

            # Find any new registrations for this member
            new_regs = [
                r for r in member.book_registrations
                if r.status == RegistrationStatus.REGISTERED
                and r.book_id != old_reg.book_id
                and r.registration_date and old_reg.registration_date
                and r.registration_date > old_reg.registration_date
            ]

            if destination_book_id:
                new_regs = [r for r in new_regs if r.book_id == destination_book_id]

            for new_reg in new_regs:
                source_book = old_reg.book
                dest_book = new_reg.book
                
                gap_days = (new_reg.registration_date - old_reg.registration_date).days if new_reg.registration_date and old_reg.registration_date else 0

                transfer_key = f"{source_book.name if source_book else 'Unknown'} → {dest_book.name if dest_book else 'Unknown'}"
                transfer_counts[transfer_key] = transfer_counts.get(transfer_key, 0) + 1

                data.append({
                    "member": f"{member.last_name}, {member.first_name}",
                    "source_book": source_book.name if source_book else "Unknown",
                    "source_tier": source_book.book_number if source_book else 1,
                    "drop_date": old_reg.registration_date.strftime("%m/%d/%Y") if old_reg.registration_date else "N/A",
                    "drop_reason": old_reg.status.value if old_reg.status else "unknown",
                    "new_book": dest_book.name if dest_book else "Unknown",
                    "new_tier": dest_book.book_number if dest_book else 1,
                    "re_reg_date": new_reg.registration_date.strftime("%m/%d/%Y") if new_reg.registration_date else "N/A",
                    "gap_days": gap_days,
                })

        # Sort by gap days
        data.sort(key=lambda x: x["gap_days"])

        # Summary
        most_common_transfers = sorted(transfer_counts.items(), key=lambda x: x[1], reverse=True)[:5]
        avg_gap = round(sum(d["gap_days"] for d in data) / len(data), 1) if data else 0

        return {
            "data": data,
            "summary": {
                "total_transfers": len(data),
                "most_common_transfers": most_common_transfers,
                "avg_gap_days": avg_gap,
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "source_book_id": source_book_id,
                "destination_book_id": destination_book_id,
            },
            "generated_at": datetime.now(),
            "report_name": "Book Transfer Report",
        }

    def render_book_transfer_report_pdf(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        source_book_id: Optional[int] = None,
        destination_book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render PDF for book transfer report (Officer+ only)."""
        data = self.get_book_transfer_report(start_date, end_date, source_book_id, destination_book_id)
        return self._render_pdf("book_transfer.html", data)

    def render_book_transfer_report_excel(
        self,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        source_book_id: Optional[int] = None,
        destination_book_id: Optional[int] = None,
    ) -> BytesIO:
        """Render Excel for book transfer report."""
        data = self.get_book_transfer_report(start_date, end_date, source_book_id, destination_book_id)

        headers = ["Member", "Source Book", "Source Tier", "Drop Date", "Drop Reason", "New Book", "New Tier", "Re-Reg Date", "Gap Days"]
        rows = [
            [d["member"], d["source_book"], d["source_tier"], d["drop_date"], d["drop_reason"],
             d["new_book"], d["new_tier"], d["re_reg_date"], d["gap_days"]]
            for d in data["data"]
        ]
        return self._render_excel(headers, rows, "Book Transfers", data["report_name"])

    # ========================================
    # WEEK 40: P2 BATCH 1 - REGISTRATION & BOOK ANALYTICS (12 reports)
    # ========================================

    # --- Report P2-1: Registration Aging Report ---

    def generate_registration_aging_report(
        self,
        format: str = "pdf",
        book_id: Optional[int] = None,
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """Registration aging analysis with duration buckets (0-30, 31-90, 91-180, 180+ days)."""
        as_of = as_of_date or date.today()

        query = self.db.query(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        registrations = query.all()

        # Bucket registrations by age
        buckets = {
            "0-30": [], "31-90": [], "91-180": [], "180+": []
        }

        for reg in registrations:
            # Calculate days since registration (from APN integer part = Excel serial date)
            if reg.registration_number:
                # APN integer part is Excel serial date
                reg_date_serial = int(reg.registration_number)
                # Excel epoch: 1900-01-01 is day 1 (with leap year bug, day 60 = 1900-02-29 doesn't exist)
                # Convert to actual date
                from datetime import timedelta
                excel_epoch = date(1899, 12, 30)  # Adjusted for Excel's quirks
                reg_date = excel_epoch + timedelta(days=reg_date_serial)
                days_on_book = (as_of - reg_date).days

                if days_on_book <= 30:
                    buckets["0-30"].append(reg)
                elif days_on_book <= 90:
                    buckets["31-90"].append(reg)
                elif days_on_book <= 180:
                    buckets["91-180"].append(reg)
                else:
                    buckets["180+"].append(reg)

        data = {
            "data": [
                {"bucket": "0-30 days", "count": len(buckets["0-30"]), "registrations": buckets["0-30"]},
                {"bucket": "31-90 days", "count": len(buckets["31-90"]), "registrations": buckets["31-90"]},
                {"bucket": "91-180 days", "count": len(buckets["91-180"]), "registrations": buckets["91-180"]},
                {"bucket": "180+ days", "count": len(buckets["180+"]), "registrations": buckets["180+"]},
            ],
            "summary": {
                "total_registrations": len(registrations),
                "avg_days_on_book": sum((as_of - (date(1899, 12, 30) + timedelta(days=int(r.registration_number)))).days for r in registrations if r.registration_number) / len(registrations) if registrations else 0,
            },
            "filters": {"book_id": book_id, "as_of_date": as_of},
            "generated_at": datetime.now(),
            "report_name": "Registration Aging Report",
        }

        if format == "pdf":
            return self._render_pdf("registration_aging.html", data).getvalue()
        else:
            headers = ["Bucket", "Count", "Percentage"]
            total = len(registrations)
            rows = [[d["bucket"], d["count"], f"{d['count']/total*100:.1f}%" if total else "0%"] for d in data["data"]]
            return self._render_excel(headers, rows, "Registration Aging", data["report_name"]).getvalue()

    # --- Report P2-2: Registration Turnover Report ---

    def generate_registration_turnover_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> bytes:
        """New registrations vs departures (dispatched, rolled-off, resigned) by period."""
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        query = self.db.query(RegistrationActivity).filter(
            RegistrationActivity.activity_date >= start,
            RegistrationActivity.activity_date <= end,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        activities = query.all()

        # Categorize activities
        new_registrations = [a for a in activities if a.activity_type == RegistrationAction.REGISTRATION]
        departures = [a for a in activities if a.activity_type in [
            RegistrationAction.DISPATCH, RegistrationAction.RESIGNATION
        ]]
        rolled_off = [a for a in activities if hasattr(RegistrationAction, 'ROLL_OFF') and a.activity_type == RegistrationAction.ROLL_OFF]

        data = {
            "data": [
                {"category": "New Registrations", "count": len(new_registrations)},
                {"category": "Dispatched", "count": len([a for a in activities if a.activity_type == RegistrationAction.DISPATCH])},
                {"category": "Resigned", "count": len([a for a in activities if a.activity_type == RegistrationAction.RESIGNATION])},
                {"category": "Rolled Off", "count": len(rolled_off)},
            ],
            "summary": {
                "net_change": len(new_registrations) - len(departures) - len(rolled_off),
                "period_days": (end - start).days,
            },
            "filters": {"start_date": start, "end_date": end, "book_id": book_id},
            "generated_at": datetime.now(),
            "report_name": "Registration Turnover Report",
        }

        if format == "pdf":
            return self._render_pdf("registration_turnover.html", data).getvalue()
        else:
            headers = ["Category", "Count"]
            rows = [[d["category"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Registration Turnover", data["report_name"]).getvalue()

    # --- Report P2-3: Re-Sign Compliance Report (Rule 7) ---

    def generate_re_sign_compliance_report(
        self,
        format: str = "pdf",
        book_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Re-sign on-time vs late vs missed by book (Rule 7 enforcement analytics)."""
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        start = start_date or (date.today() - timedelta(days=90))
        end = end_date or date.today()

        query = self.db.query(RegistrationActivity).filter(
            RegistrationActivity.activity_type == RegistrationAction.RE_SIGN,
            RegistrationActivity.activity_date >= start,
            RegistrationActivity.activity_date <= end,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        re_signs = query.all()

        # Analyze re-sign compliance (30-day cycle)
        on_time = []
        late = []
        missed = []

        for re_sign in re_signs:
            # Get previous re-sign for this registration
            prev_re_sign = self.db.query(RegistrationActivity).filter(
                RegistrationActivity.registration_id == re_sign.registration_id,
                RegistrationActivity.activity_type == RegistrationAction.RE_SIGN,
                RegistrationActivity.activity_date < re_sign.activity_date,
            ).order_by(RegistrationActivity.activity_date.desc()).first()

            if prev_re_sign:
                days_gap = (re_sign.activity_date - prev_re_sign.activity_date).days
                if days_gap <= 30:
                    on_time.append(re_sign)
                elif days_gap <= 35:  # Late within 5 days
                    late.append(re_sign)
                else:
                    missed.append(re_sign)
            else:
                # First re-sign, consider on-time
                on_time.append(re_sign)

        data = {
            "data": [
                {"status": "On Time", "count": len(on_time), "color": "green"},
                {"status": "Late (1-5 days)", "count": len(late), "color": "yellow"},
                {"status": "Missed (5+ days)", "count": len(missed), "color": "red"},
            ],
            "summary": {
                "total_re_signs": len(re_signs),
                "compliance_rate": len(on_time) / len(re_signs) * 100 if re_signs else 0,
            },
            "filters": {"start_date": start, "end_date": end, "book_id": book_id},
            "generated_at": datetime.now(),
            "report_name": "Re-Sign Compliance Report (Rule 7)",
        }

        if format == "pdf":
            return self._render_pdf("re_sign_compliance.html", data).getvalue()
        else:
            headers = ["Status", "Count", "Percentage"]
            total = len(re_signs)
            rows = [[d["status"], d["count"], f"{d['count']/total*100:.1f}%" if total else "0%"] for d in data["data"]]
            return self._render_excel(headers, rows, "Re-Sign Compliance", data["report_name"]).getvalue()

    # --- Report P2-4: Re-Registration Pattern Analysis (Rule 6) ---

    def generate_re_registration_patterns_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        book_id: Optional[int] = None,
    ) -> bytes:
        """Re-registration triggers (short call, under scale, 90-day, turnaround) by frequency."""
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        start = start_date or (date.today() - timedelta(days=90))
        end = end_date or date.today()

        query = self.db.query(RegistrationActivity).filter(
            RegistrationActivity.activity_type == RegistrationAction.RE_REGISTRATION,
            RegistrationActivity.activity_date >= start,
            RegistrationActivity.activity_date <= end,
        )

        if book_id:
            query = query.filter(RegistrationActivity.book_id == book_id)

        re_regs = query.all()

        # Group by reason (if available in activity notes/details)
        reason_counts = {}
        for re_reg in re_regs:
            reason = re_reg.notes or "Unknown"
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

        data = {
            "data": [{"reason": k, "count": v, "percentage": v / len(re_regs) * 100 if re_regs else 0}
                     for k, v in sorted(reason_counts.items(), key=lambda x: x[1], reverse=True)],
            "summary": {
                "total_re_registrations": len(re_regs),
                "unique_reasons": len(reason_counts),
            },
            "filters": {"start_date": start, "end_date": end, "book_id": book_id},
            "generated_at": datetime.now(),
            "report_name": "Re-Registration Pattern Analysis (Rule 6)",
        }

        if format == "pdf":
            return self._render_pdf("re_registration_patterns.html", data).getvalue()
        else:
            headers = ["Reason", "Count", "Percentage"]
            rows = [[d["reason"], d["count"], f"{d['percentage']:.1f}%"] for d in data["data"]]
            return self._render_excel(headers, rows, "Re-Registration Patterns", data["report_name"]).getvalue()

    # --- Report P2-5: Inactive Registration Report ---

    def generate_inactive_registrations_report(
        self,
        format: str = "pdf",
        book_id: Optional[int] = None,
        inactive_days: int = 60,
    ) -> bytes:
        """Registrations with no activity (no re-sign, no dispatch, no bid) for 60+ days."""
        cutoff_date = date.today() - timedelta(days=inactive_days)

        query = self.db.query(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED,
        )

        if book_id:
            query = query.filter(BookRegistration.book_id == book_id)

        registrations = query.all()

        inactive_regs = []
        for reg in registrations:
            # Check last activity date (simplified - would need to check dispatch, bid, re-sign tables)
            # For now, use registration number date as proxy
            if reg.registration_number:
                reg_date_serial = int(reg.registration_number)
                excel_epoch = date(1899, 12, 30)
                reg_date = excel_epoch + timedelta(days=reg_date_serial)
                days_inactive = (date.today() - reg_date).days

                if days_inactive >= inactive_days:
                    inactive_regs.append({
                        "member": f"{reg.member.first_name} {reg.member.last_name}" if reg.member else "Unknown",
                        "book": reg.book.book_name if reg.book else "Unknown",
                        "days_inactive": days_inactive,
                        "apn": reg.registration_number,
                    })

        data = {
            "data": inactive_regs,
            "summary": {
                "total_inactive": len(inactive_regs),
                "inactive_threshold_days": inactive_days,
            },
            "filters": {"book_id": book_id, "inactive_days": inactive_days},
            "generated_at": datetime.now(),
            "report_name": "Inactive Registration Report",
        }

        if format == "pdf":
            return self._render_pdf("inactive_registrations.html", data).getvalue()
        else:
            headers = ["Member", "Book", "Days Inactive", "APN"]
            rows = [[d["member"], d["book"], d["days_inactive"], d["apn"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Inactive Registrations", data["report_name"]).getvalue()

    # --- Report P2-6: Cross-Book Registration Analysis (Rule 5) ---

    def generate_cross_book_registration_report(
        self,
        format: str = "pdf",
    ) -> bytes:
        """Members registered on multiple books simultaneously (Rule 5 validation)."""
        # Query members with multiple active registrations
        from sqlalchemy import func as sql_func

        multi_book_members = self.db.query(
            BookRegistration.member_id,
            sql_func.count(sql_func.distinct(BookRegistration.book_id)).label("book_count")
        ).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED
        ).group_by(
            BookRegistration.member_id
        ).having(
            sql_func.count(sql_func.distinct(BookRegistration.book_id)) > 1
        ).all()

        data_rows = []
        for member_id, book_count in multi_book_members:
            member = self.db.query(Member).filter(Member.id == member_id).first()
            regs = self.db.query(BookRegistration).filter(
                BookRegistration.member_id == member_id,
                BookRegistration.status == RegistrationStatus.REGISTERED
            ).all()

            books = [reg.book.book_name for reg in regs if reg.book]
            data_rows.append({
                "member": f"{member.first_name} {member.last_name}" if member else "Unknown",
                "book_count": book_count,
                "books": ", ".join(books),
            })

        data = {
            "data": data_rows,
            "summary": {
                "total_multi_book_members": len(data_rows),
                "max_books_per_member": max([r["book_count"] for r in data_rows]) if data_rows else 0,
            },
            "filters": {},
            "generated_at": datetime.now(),
            "report_name": "Cross-Book Registration Analysis (Rule 5)",
        }

        if format == "pdf":
            return self._render_pdf("cross_book_registration.html", data).getvalue()
        else:
            headers = ["Member", "Book Count", "Books"]
            rows = [[d["member"], d["book_count"], d["books"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Cross-Book Registration", data["report_name"]).getvalue()

    # --- Report P2-7: Classification Demand Gap ---

    def generate_classification_demand_gap_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Registrations available vs labor requests received per classification."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        # Count registrations by classification (via book)
        from sqlalchemy import func as sql_func

        reg_counts = {}
        books = self.db.query(ReferralBook).all()
        for book in books:
            classification = book.classification or "Unknown"
            count = self.db.query(BookRegistration).filter(
                BookRegistration.book_id == book.id,
                BookRegistration.status == RegistrationStatus.REGISTERED
            ).count()
            reg_counts[classification] = reg_counts.get(classification, 0) + count

        # Count labor requests by classification
        request_counts = {}
        requests = self.db.query(LaborRequest).filter(
            LaborRequest.created_at >= start,
            LaborRequest.created_at <= end,
        ).all()

        for req in requests:
            # Get classification from book
            classification = req.book.classification if req.book else "Unknown"
            request_counts[classification] = request_counts.get(classification, 0) + 1

        # Calculate gap
        all_classifications = set(list(reg_counts.keys()) + list(request_counts.keys()))
        data_rows = []
        for classification in sorted(all_classifications):
            available = reg_counts.get(classification, 0)
            demand = request_counts.get(classification, 0)
            gap = available - demand
            data_rows.append({
                "classification": classification,
                "available": available,
                "demand": demand,
                "gap": gap,
                "gap_status": "surplus" if gap > 0 else "shortage" if gap < 0 else "balanced",
            })

        data = {
            "data": data_rows,
            "summary": {
                "total_available": sum(reg_counts.values()),
                "total_demand": sum(request_counts.values()),
                "overall_gap": sum(reg_counts.values()) - sum(request_counts.values()),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Classification Demand Gap",
        }

        if format == "pdf":
            return self._render_pdf("classification_demand_gap.html", data).getvalue()
        else:
            headers = ["Classification", "Available", "Demand", "Gap", "Status"]
            rows = [[d["classification"], d["available"], d["demand"], d["gap"], d["gap_status"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Classification Demand Gap", data["report_name"]).getvalue()

    # --- Report P2-8: Book Comparison Dashboard ---

    def generate_book_comparison_report(
        self,
        format: str = "pdf",
    ) -> bytes:
        """Side-by-side metrics per book: avg wait time, dispatch rate, turnover, fill rate."""
        books = self.db.query(ReferralBook).all()

        data_rows = []
        for book in books:
            # Count registrations
            reg_count = self.db.query(BookRegistration).filter(
                BookRegistration.book_id == book.id,
                BookRegistration.status == RegistrationStatus.REGISTERED
            ).count()

            # Count dispatches (last 30 days)
            from datetime import datetime as dt
            thirty_days_ago = dt.now() - timedelta(days=30)
            dispatch_count = self.db.query(Dispatch).filter(
                Dispatch.book_id == book.id,
                Dispatch.dispatch_date >= thirty_days_ago.date()
            ).count()

            # Calculate metrics
            dispatch_rate = dispatch_count / reg_count if reg_count else 0

            data_rows.append({
                "book": book.book_name,
                "registrations": reg_count,
                "dispatches_30d": dispatch_count,
                "dispatch_rate": f"{dispatch_rate:.2%}",
                "avg_wait_days": "N/A",  # Would need more complex calculation
                "fill_rate": "N/A",  # Would need labor request data
            })

        data = {
            "data": data_rows,
            "summary": {
                "total_books": len(books),
                "total_registrations": sum(r["registrations"] for r in data_rows),
                "total_dispatches_30d": sum(r["dispatches_30d"] for r in data_rows),
            },
            "filters": {},
            "generated_at": datetime.now(),
            "report_name": "Book Comparison Dashboard",
        }

        if format == "pdf":
            return self._render_pdf("book_comparison.html", data).getvalue()
        else:
            headers = ["Book", "Registrations", "Dispatches (30d)", "Dispatch Rate", "Avg Wait Days", "Fill Rate"]
            rows = [[d["book"], d["registrations"], d["dispatches_30d"], d["dispatch_rate"], d["avg_wait_days"], d["fill_rate"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Book Comparison", data["report_name"]).getvalue()

    # --- Report P2-9: Tier Distribution Report ---

    def generate_tier_distribution_report(
        self,
        format: str = "pdf",
    ) -> bytes:
        """Book 1/2/3/4 registration counts and percentages per classification."""
        from sqlalchemy import func as sql_func

        # Group by book tier (book_priority_number)
        tier_data = self.db.query(
            ReferralBook.classification,
            ReferralBook.book_priority_number,
            sql_func.count(BookRegistration.id).label("count")
        ).join(
            BookRegistration, BookRegistration.book_id == ReferralBook.id
        ).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED
        ).group_by(
            ReferralBook.classification,
            ReferralBook.book_priority_number
        ).all()

        # Organize by classification
        by_classification = {}
        for classification, tier, count in tier_data:
            if classification not in by_classification:
                by_classification[classification] = {1: 0, 2: 0, 3: 0, 4: 0}
            by_classification[classification][tier or 1] = count

        data_rows = []
        for classification, tiers in sorted(by_classification.items()):
            total = sum(tiers.values())
            data_rows.append({
                "classification": classification,
                "tier1": tiers[1],
                "tier1_pct": f"{tiers[1]/total*100:.1f}%" if total else "0%",
                "tier2": tiers[2],
                "tier2_pct": f"{tiers[2]/total*100:.1f}%" if total else "0%",
                "tier3": tiers[3],
                "tier3_pct": f"{tiers[3]/total*100:.1f}%" if total else "0%",
                "tier4": tiers[4],
                "tier4_pct": f"{tiers[4]/total*100:.1f}%" if total else "0%",
                "total": total,
                "inverted": tiers[3] > tiers[1],  # Flag inverted distributions
            })

        data = {
            "data": data_rows,
            "summary": {
                "total_classifications": len(by_classification),
                "inverted_count": sum(1 for r in data_rows if r["inverted"]),
            },
            "filters": {},
            "generated_at": datetime.now(),
            "report_name": "Tier Distribution Report",
        }

        if format == "pdf":
            return self._render_pdf("tier_distribution.html", data).getvalue()
        else:
            headers = ["Classification", "Tier 1", "Tier 1 %", "Tier 2", "Tier 2 %", "Tier 3", "Tier 3 %", "Tier 4", "Tier 4 %", "Total", "Inverted?"]
            rows = [[d["classification"], d["tier1"], d["tier1_pct"], d["tier2"], d["tier2_pct"],
                     d["tier3"], d["tier3_pct"], d["tier4"], d["tier4_pct"], d["total"],
                     "Yes" if d["inverted"] else "No"] for d in data["data"]]
            return self._render_excel(headers, rows, "Tier Distribution", data["report_name"]).getvalue()

    # --- Report P2-10: Book Capacity Trends ---

    def generate_book_capacity_trends_report(
        self,
        format: str = "pdf",
        book_id: Optional[int] = None,
        period: str = "weekly",  # weekly or monthly
    ) -> bytes:
        """Registration counts over time (weekly/monthly) per book with period-over-period change."""
        # This requires historical data tracking - simplified implementation
        from datetime import datetime as dt

        # Get registration counts for last 12 weeks/months
        periods = 12
        period_length = 7 if period == "weekly" else 30

        data_rows = []
        for i in range(periods):
            period_end = date.today() - timedelta(days=i * period_length)
            period_start = period_end - timedelta(days=period_length)

            query = self.db.query(BookRegistration).filter(
                BookRegistration.status == RegistrationStatus.REGISTERED
            )

            if book_id:
                query = query.filter(BookRegistration.book_id == book_id)

            count = query.count()  # Simplified - should check registration date

            data_rows.insert(0, {
                "period": f"{period_start} to {period_end}",
                "count": count,
                "change": 0 if i == periods - 1 else count - data_rows[0]["count"] if data_rows else 0,
            })

        data = {
            "data": data_rows,
            "summary": {
                "periods": periods,
                "period_type": period,
                "avg_capacity": sum(r["count"] for r in data_rows) / len(data_rows) if data_rows else 0,
            },
            "filters": {"book_id": book_id, "period": period},
            "generated_at": datetime.now(),
            "report_name": "Book Capacity Trends",
        }

        if format == "pdf":
            return self._render_pdf("book_capacity_trends.html", data).getvalue()
        else:
            headers = ["Period", "Count", "Change"]
            rows = [[d["period"], d["count"], d["change"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Book Capacity Trends", data["report_name"]).getvalue()

    # --- Report P2-11: APN Wait Time Distribution ---

    def generate_apn_wait_time_report(
        self,
        format: str = "pdf",
        book_id: Optional[int] = None,
    ) -> bytes:
        """Histogram of wait times from registration to first dispatch, bucketed by book."""
        # Query dispatches with their registration data
        dispatches = self.db.query(Dispatch).all()

        wait_time_buckets = {
            "<7 days": 0, "7-30": 0, "31-90": 0, "91-180": 0, "180+": 0, "Still Waiting": 0
        }

        for dispatch in dispatches:
            if dispatch.registration and dispatch.registration.registration_number:
                # Calculate wait time from APN date to dispatch date
                reg_date_serial = int(dispatch.registration.registration_number)
                excel_epoch = date(1899, 12, 30)
                reg_date = excel_epoch + timedelta(days=reg_date_serial)

                if dispatch.dispatch_date:
                    wait_days = (dispatch.dispatch_date - reg_date).days

                    if wait_days < 7:
                        wait_time_buckets["<7 days"] += 1
                    elif wait_days <= 30:
                        wait_time_buckets["7-30"] += 1
                    elif wait_days <= 90:
                        wait_time_buckets["31-90"] += 1
                    elif wait_days <= 180:
                        wait_time_buckets["91-180"] += 1
                    else:
                        wait_time_buckets["180+"] += 1

        # Count members still waiting (active registrations with no dispatch)
        active_regs = self.db.query(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.REGISTERED
        )
        if book_id:
            active_regs = active_regs.filter(BookRegistration.book_id == book_id)

        wait_time_buckets["Still Waiting"] = active_regs.count()

        data = {
            "data": [{"bucket": k, "count": v} for k, v in wait_time_buckets.items()],
            "summary": {
                "total_dispatches": sum(v for k, v in wait_time_buckets.items() if k != "Still Waiting"),
                "still_waiting": wait_time_buckets["Still Waiting"],
            },
            "filters": {"book_id": book_id},
            "generated_at": datetime.now(),
            "report_name": "APN Wait Time Distribution",
        }

        if format == "pdf":
            return self._render_pdf("apn_wait_time.html", data).getvalue()
        else:
            headers = ["Wait Time Bucket", "Count"]
            rows = [[d["bucket"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "APN Wait Time", data["report_name"]).getvalue()

    # --- Report P2-12: Seasonal Registration Patterns ---

    def generate_seasonal_registration_report(
        self,
        format: str = "pdf",
        year: Optional[int] = None,
    ) -> bytes:
        """Registration volume by month/quarter overlaid with dispatch volume for correlation."""
        from src.models.registration_activity import RegistrationActivity
        from src.db.enums import RegistrationAction

        target_year = year or date.today().year

        # Count registrations by month
        monthly_data = []
        for month in range(1, 13):
            month_start = date(target_year, month, 1)
            if month == 12:
                month_end = date(target_year + 1, 1, 1)
            else:
                month_end = date(target_year, month + 1, 1)

            reg_count = self.db.query(RegistrationActivity).filter(
                RegistrationActivity.activity_type == RegistrationAction.REGISTRATION,
                RegistrationActivity.activity_date >= month_start,
                RegistrationActivity.activity_date < month_end,
            ).count()

            dispatch_count = self.db.query(Dispatch).filter(
                Dispatch.dispatch_date >= month_start,
                Dispatch.dispatch_date < month_end,
            ).count()

            monthly_data.append({
                "month": month_start.strftime("%B"),
                "registrations": reg_count,
                "dispatches": dispatch_count,
                "ratio": dispatch_count / reg_count if reg_count else 0,
            })

        data = {
            "data": monthly_data,
            "summary": {
                "year": target_year,
                "total_registrations": sum(d["registrations"] for d in monthly_data),
                "total_dispatches": sum(d["dispatches"] for d in monthly_data),
                "avg_monthly_registrations": sum(d["registrations"] for d in monthly_data) / 12,
                "avg_monthly_dispatches": sum(d["dispatches"] for d in monthly_data) / 12,
            },
            "filters": {"year": target_year},
            "generated_at": datetime.now(),
            "report_name": "Seasonal Registration Patterns",
        }

        if format == "pdf":
            return self._render_pdf("seasonal_registration.html", data).getvalue()
        else:
            headers = ["Month", "Registrations", "Dispatches", "Dispatch:Reg Ratio"]
            rows = [[d["month"], d["registrations"], d["dispatches"], f"{d['ratio']:.2f}"] for d in data["data"]]
            return self._render_excel(headers, rows, "Seasonal Registration", data["report_name"]).getvalue()

    # --- WEEK 41 REPORTS (19 new methods) ---
    # Theme A: Dispatch Operations (6 reports)

    def generate_dispatch_success_rate_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Dispatch success rate: offers accepted vs declined by period."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        # Count dispatches by status
        dispatched = self.db.query(func.count(Dispatch.id)).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end,
            Dispatch.status == DispatchStatus.COMPLETED
        ).scalar() or 0

        cancelled = self.db.query(func.count(Dispatch.id)).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end,
            Dispatch.status == DispatchStatus.CANCELLED
        ).scalar() or 0

        total_offers = dispatched + cancelled
        success_rate = (dispatched / total_offers * 100) if total_offers > 0 else 0

        data = {
            "data": [
                {"status": "Accepted (Completed)", "count": dispatched},
                {"status": "Declined (Cancelled)", "count": cancelled},
            ],
            "summary": {
                "total_offers": total_offers,
                "success_rate": success_rate,
                "acceptance_count": dispatched,
                "decline_count": cancelled,
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Dispatch Success Rate",
        }

        if format == "pdf":
            return self._render_pdf("dispatch_success_rate.html", data).getvalue()
        else:
            headers = ["Status", "Count", "Percentage"]
            rows = [
                [d["status"], d["count"], f"{d['count']/total_offers*100:.1f}%" if total_offers else "0%"]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Success Rate", data["report_name"]).getvalue()

    def generate_time_to_fill_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Time to fill: days from request to dispatch completion."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end,
            Dispatch.status == DispatchStatus.COMPLETED
        ).all()

        fill_times = []
        for dispatch in dispatches:
            if dispatch.labor_request and dispatch.dispatch_date:
                days_to_fill = (dispatch.dispatch_date - dispatch.labor_request.request_date).days
                fill_times.append(days_to_fill)

        buckets = {"Same Day": 0, "1-2 Days": 0, "3-7 Days": 0, "8+ Days": 0}
        for days in fill_times:
            if days == 0:
                buckets["Same Day"] += 1
            elif days <= 2:
                buckets["1-2 Days"] += 1
            elif days <= 7:
                buckets["3-7 Days"] += 1
            else:
                buckets["8+ Days"] += 1

        avg_time = sum(fill_times) / len(fill_times) if fill_times else 0

        data = {
            "data": [{"bucket": k, "count": v} for k, v in buckets.items()],
            "summary": {
                "total_completed": len(fill_times),
                "avg_days_to_fill": avg_time,
                "median_days_to_fill": sorted(fill_times)[len(fill_times)//2] if fill_times else 0,
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Time to Fill Report",
        }

        if format == "pdf":
            return self._render_pdf("time_to_fill.html", data).getvalue()
        else:
            headers = ["Bucket", "Count"]
            rows = [[d["bucket"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Time to Fill", data["report_name"]).getvalue()

    def generate_dispatch_method_comparison_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Dispatch method comparison: queue vs by-name vs short call."""
        from src.db.enums import DispatchMethod

        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        methods = {}
        for method in [DispatchMethod.QUEUE, DispatchMethod.BY_NAME, DispatchMethod.SHORT_CALL]:
            count = self.db.query(func.count(Dispatch.id)).filter(
                Dispatch.dispatch_date >= start,
                Dispatch.dispatch_date <= end,
                Dispatch.dispatch_method == method
            ).scalar() or 0
            methods[method.value] = count

        total = sum(methods.values())

        data = {
            "data": [{"method": k, "count": v} for k, v in methods.items()],
            "summary": {
                "total_dispatches": total,
                "queue_dispatches": methods.get("QUEUE", 0),
                "by_name_dispatches": methods.get("BY_NAME", 0),
                "short_call_dispatches": methods.get("SHORT_CALL", 0),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Dispatch Method Comparison",
        }

        if format == "pdf":
            return self._render_pdf("dispatch_method_comparison.html", data).getvalue()
        else:
            headers = ["Method", "Count", "Percentage"]
            rows = [
                [d["method"], d["count"], f"{d['count']/total*100:.1f}%" if total else "0%"]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Dispatch Methods", data["report_name"]).getvalue()

    def generate_dispatch_geographic_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Geographic dispatch analysis by region."""
        from src.db.enums import BookRegion

        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        regions = {}
        for region in [BookRegion.SEATTLE, BookRegion.BREMERTON, BookRegion.PT_ANGELES]:
            count = self.db.query(func.count(Dispatch.id)).join(
                ReferralBook, Dispatch.book_id == ReferralBook.id
            ).filter(
                Dispatch.dispatch_date >= start,
                Dispatch.dispatch_date <= end,
                ReferralBook.region == region
            ).scalar() or 0
            regions[region.value] = count

        total = sum(regions.values())

        data = {
            "data": [{"region": k, "count": v} for k, v in regions.items()],
            "summary": {
                "total_dispatches": total,
                "regions_active": sum(1 for v in regions.values() if v > 0),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Geographic Dispatch Report",
        }

        if format == "pdf":
            return self._render_pdf("dispatch_geographic.html", data).getvalue()
        else:
            headers = ["Region", "Count", "Percentage"]
            rows = [
                [d["region"], d["count"], f"{d['count']/total*100:.1f}%" if total else "0%"]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Geographic Dispatch", data["report_name"]).getvalue()

    def generate_termination_reason_analysis_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Dispatch termination reason analysis (completion, cancellation, etc.)."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        # Group by status (serve as termination reason proxy)
        termination_reasons = {}
        for status in [DispatchStatus.COMPLETED, DispatchStatus.CANCELLED, DispatchStatus.SHORT_CALL]:
            count = self.db.query(func.count(Dispatch.id)).filter(
                Dispatch.dispatch_date >= start,
                Dispatch.dispatch_date <= end,
                Dispatch.status == status
            ).scalar() or 0
            termination_reasons[status.value] = count

        total = sum(termination_reasons.values())

        data = {
            "data": [{"reason": k, "count": v} for k, v in termination_reasons.items()],
            "summary": {
                "total_dispatches": total,
                "completed_count": termination_reasons.get("COMPLETED", 0),
                "cancelled_count": termination_reasons.get("CANCELLED", 0),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Termination Reason Analysis",
        }

        if format == "pdf":
            return self._render_pdf("termination_reason_analysis.html", data).getvalue()
        else:
            headers = ["Reason", "Count", "Percentage"]
            rows = [
                [d["reason"], d["count"], f"{d['count']/total*100:.1f}%" if total else "0%"]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Termination Reasons", data["report_name"]).getvalue()

    def generate_return_dispatch_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Return dispatch: repeat member-employer pairs within period."""
        start = start_date or (date.today() - timedelta(days=90))
        end = end_date or date.today()

        dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end
        ).all()

        pairs = {}
        for dispatch in dispatches:
            if dispatch.member_id and dispatch.labor_request and dispatch.labor_request.employer_id:
                pair_key = (dispatch.member_id, dispatch.labor_request.employer_id)
                pairs[pair_key] = pairs.get(pair_key, 0) + 1

        repeat_pairs = {k: v for k, v in pairs.items() if v > 1}
        repeat_count = len(repeat_pairs)
        total_pairs = len(pairs)
        repeat_percentage = (repeat_count / total_pairs * 100) if total_pairs > 0 else 0

        data = {
            "data": [{"pair_count": k, "frequency": v} for k, v in sorted(
                {v: pairs.values().count(v) for v in set(pairs.values())}.items()
            )],
            "summary": {
                "total_pairs": total_pairs,
                "repeat_pairs": repeat_count,
                "repeat_percentage": repeat_percentage,
                "unique_members": len(set(k[0] for k in pairs.keys())),
                "unique_employers": len(set(k[1] for k in pairs.keys())),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Return Dispatch Report",
        }

        if format == "pdf":
            return self._render_pdf("return_dispatch.html", data).getvalue()
        else:
            headers = ["Dispatches per Pair", "Count"]
            rows = [[d["pair_count"], d["frequency"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Return Dispatch", data["report_name"]).getvalue()

    # Theme B: Employer Intelligence (6 reports)

    def generate_employer_growth_trends_report(
        self,
        format: str = "pdf",
        current_year: Optional[int] = None,
        require_officer_access: bool = True,
    ) -> bytes:
        """Employer growth trends: YoY volume comparison (Officer+ only)."""
        target_year = current_year or date.today().year
        prev_year = target_year - 1

        # Current year dispatches
        current_year_count = self.db.query(func.count(Dispatch.id)).filter(
            func.extract("year", Dispatch.dispatch_date) == target_year
        ).scalar() or 0

        # Previous year dispatches
        prev_year_count = self.db.query(func.count(Dispatch.id)).filter(
            func.extract("year", Dispatch.dispatch_date) == prev_year
        ).scalar() or 0

        growth = ((current_year_count - prev_year_count) / prev_year_count * 100) if prev_year_count > 0 else 0

        data = {
            "data": [
                {"year": prev_year, "dispatches": prev_year_count},
                {"year": target_year, "dispatches": current_year_count},
            ],
            "summary": {
                "prev_year_total": prev_year_count,
                "current_year_total": current_year_count,
                "yoy_growth_percent": growth,
                "growth_direction": "↑" if growth > 0 else "↓" if growth < 0 else "→",
            },
            "filters": {"current_year": target_year, "officer_restricted": require_officer_access},
            "generated_at": datetime.now(),
            "report_name": "Employer Growth Trends (Officer Only)",
        }

        if format == "pdf":
            return self._render_pdf("employer_growth_trends.html", data).getvalue()
        else:
            headers = ["Year", "Dispatches", "Change"]
            rows = [
                [prev_year, prev_year_count, "—"],
                [target_year, current_year_count, f"{growth:+.1f}%"],
            ]
            return self._render_excel(headers, rows, "Growth Trends", data["report_name"]).getvalue()

    def generate_employer_workforce_size_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Active dispatches per employer: workforce size snapshot."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end,
            Dispatch.status == DispatchStatus.COMPLETED
        ).all()

        employer_counts = {}
        for dispatch in dispatches:
            if dispatch.labor_request and dispatch.labor_request.employer:
                emp_name = dispatch.labor_request.employer.name
                employer_counts[emp_name] = employer_counts.get(emp_name, 0) + 1

        sorted_employers = sorted(employer_counts.items(), key=lambda x: x[1], reverse=True)[:20]

        data = {
            "data": [{"employer": emp, "active_dispatches": count} for emp, count in sorted_employers],
            "summary": {
                "total_employers": len(employer_counts),
                "avg_dispatches_per_employer": sum(employer_counts.values()) / len(employer_counts) if employer_counts else 0,
                "top_employer": sorted_employers[0][0] if sorted_employers else None,
                "top_employer_count": sorted_employers[0][1] if sorted_employers else 0,
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Employer Workforce Size",
        }

        if format == "pdf":
            return self._render_pdf("employer_workforce_size.html", data).getvalue()
        else:
            headers = ["Employer", "Active Dispatches"]
            rows = [[d["employer"], d["active_dispatches"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Workforce Size", data["report_name"]).getvalue()

    def generate_new_employer_activity_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """New employer activity: first-time employers in period."""
        start = start_date or (date.today() - timedelta(days=90))
        end = end_date or date.today()

        # Find all dispatches in period
        dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end
        ).all()

        # Find all employers before this period
        old_dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date < start
        ).all()
        old_employer_ids = set(d.labor_request.employer_id for d in old_dispatches if d.labor_request)

        # New employers = in this period but not before
        new_employers = {}
        for dispatch in dispatches:
            if dispatch.labor_request and dispatch.labor_request.employer:
                emp_id = dispatch.labor_request.employer_id
                if emp_id not in old_employer_ids:
                    emp_name = dispatch.labor_request.employer.name
                    new_employers[emp_name] = new_employers.get(emp_name, 0) + 1

        new_count = len(new_employers)
        total_dispatches = len(dispatches)

        data = {
            "data": [{"employer": emp, "dispatch_count": count} for emp, count in sorted(
                new_employers.items(), key=lambda x: x[1], reverse=True
            )],
            "summary": {
                "new_employers_count": new_count,
                "dispatches_from_new": sum(new_employers.values()),
                "total_period_dispatches": total_dispatches,
                "new_employer_percentage": (sum(new_employers.values()) / total_dispatches * 100) if total_dispatches > 0 else 0,
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "New Employer Activity",
        }

        if format == "pdf":
            return self._render_pdf("new_employer_activity.html", data).getvalue()
        else:
            headers = ["Employer", "Dispatch Count"]
            rows = [[d["employer"], d["dispatch_count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "New Employers", data["report_name"]).getvalue()

    def generate_contract_code_utilization_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Contract code utilization: dispatch volume by CBA code (8 codes)."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        # Fetch dispatches grouped by book contract code
        books = self.db.query(ReferralBook).all()
        contract_codes = {}

        for book in books:
            dispatch_count = self.db.query(func.count(Dispatch.id)).filter(
                Dispatch.book_id == book.id,
                Dispatch.dispatch_date >= start,
                Dispatch.dispatch_date <= end
            ).scalar() or 0

            code = book.contract_code or "UNCLASSIFIED"
            contract_codes[code] = contract_codes.get(code, 0) + dispatch_count

        total = sum(contract_codes.values())

        data = {
            "data": [{"contract_code": k, "dispatches": v} for k, v in sorted(
                contract_codes.items(), key=lambda x: x[1], reverse=True
            )],
            "summary": {
                "total_dispatches": total,
                "contract_codes_active": len([v for v in contract_codes.values() if v > 0]),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Contract Code Utilization",
        }

        if format == "pdf":
            return self._render_pdf("contract_code_utilization.html", data).getvalue()
        else:
            headers = ["Contract Code", "Dispatches", "Percentage"]
            rows = [
                [d["contract_code"], d["dispatches"], f"{d['dispatches']/total*100:.1f}%" if total else "0%"]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Contract Codes", data["report_name"]).getvalue()

    def generate_queue_velocity_report(
        self,
        format: str = "pdf",
        book_id: Optional[int] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Queue velocity: speed of dispatch from referral queue."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        query = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end,
            Dispatch.dispatch_method == "QUEUE"
        )

        if book_id:
            query = query.filter(Dispatch.book_id == book_id)

        dispatches = query.all()

        # Calculate position-to-dispatch velocity
        velocity_data = {}
        for dispatch in dispatches:
            if dispatch.book_registration:
                position = dispatch.book_registration.book_priority_number or 0
                velocity_data[position] = velocity_data.get(position, 0) + 1

        avg_position_dispatched = sum(velocity_data.keys()) / len(velocity_data) if velocity_data else 0

        data = {
            "data": [{"position": k, "count": v} for k, v in sorted(velocity_data.items())[:20]],
            "summary": {
                "total_queue_dispatches": len(dispatches),
                "avg_position_dispatched": avg_position_dispatched,
                "high_position_dispatches": len([p for p in velocity_data.keys() if p > 50]),
            },
            "filters": {"book_id": book_id, "start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Queue Velocity Report",
        }

        if format == "pdf":
            return self._render_pdf("queue_velocity.html", data).getvalue()
        else:
            headers = ["Position", "Dispatches"]
            rows = [[d["position"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Queue Velocity", data["report_name"]).getvalue()

    def generate_peak_demand_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Peak demand: labor request patterns by day of week and hour."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        requests = self.db.query(LaborRequest).filter(
            LaborRequest.request_date >= start,
            LaborRequest.request_date <= end
        ).all()

        # Group by day of week
        dow_counts = {i: 0 for i in range(7)}
        for req in requests:
            dow = req.request_date.weekday()
            dow_counts[dow] += 1

        dow_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

        data = {
            "data": [
                {"day": dow_names[i], "request_count": dow_counts[i]}
                for i in range(7)
            ],
            "summary": {
                "total_requests": len(requests),
                "peak_day": dow_names[max(dow_counts, key=dow_counts.get)],
                "peak_day_count": max(dow_counts.values()),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Peak Demand Report",
        }

        if format == "pdf":
            return self._render_pdf("peak_demand.html", data).getvalue()
        else:
            headers = ["Day of Week", "Request Count"]
            rows = [[d["day"], d["request_count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Peak Demand", data["report_name"]).getvalue()

    # Theme C: Enforcement (7 reports)

    def generate_check_mark_patterns_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Check mark patterns: Rule 10 penalty analysis."""
        from src.models.check_mark import CheckMark

        start = start_date or (date.today() - timedelta(days=90))
        end = end_date or date.today()

        check_marks = self.db.query(CheckMark).filter(
            CheckMark.created_at >= start,
            CheckMark.created_at <= end
        ).all()

        # Bucket by member count
        member_counts = {}
        for cm in check_marks:
            member_id = cm.member_id
            member_counts[member_id] = member_counts.get(member_id, 0) + 1

        buckets = {"1 Mark": 0, "2 Marks": 0, "3+ Marks (Rolled Off)": 0}
        for count in member_counts.values():
            if count == 1:
                buckets["1 Mark"] += 1
            elif count == 2:
                buckets["2 Marks"] += 1
            else:
                buckets["3+ Marks (Rolled Off)"] += 1

        rolled_off = buckets["3+ Marks (Rolled Off)"]

        data = {
            "data": [{"bucket": k, "count": v} for k, v in buckets.items()],
            "summary": {
                "total_members_with_marks": len(member_counts),
                "total_check_marks": len(check_marks),
                "rolled_off_count": rolled_off,
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Check Mark Patterns",
        }

        if format == "pdf":
            return self._render_pdf("check_mark_patterns.html", data).getvalue()
        else:
            headers = ["Bucket", "Count"]
            rows = [[d["bucket"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Check Mark Patterns", data["report_name"]).getvalue()

    def generate_check_mark_exceptions_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Check mark exceptions: Rule 11 exception audit (no-check-mark reasons)."""
        from src.models.check_mark import CheckMark
        from src.db.enums import NoCheckMarkReason

        start = start_date or (date.today() - timedelta(days=90))
        end = end_date or date.today()

        exceptions = self.db.query(CheckMark).filter(
            CheckMark.created_at >= start,
            CheckMark.created_at <= end,
            CheckMark.exception_reason.isnot(None)
        ).all()

        reason_counts = {}
        for exc in exceptions:
            reason = exc.exception_reason or "Unknown"
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

        total_possible = self.db.query(func.count(Dispatch.id)).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end
        ).scalar() or 1

        data = {
            "data": [{"reason": k, "count": v} for k, v in sorted(
                reason_counts.items(), key=lambda x: x[1], reverse=True
            )],
            "summary": {
                "total_exceptions": len(exceptions),
                "exception_percentage": (len(exceptions) / total_possible * 100),
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Check Mark Exceptions",
        }

        if format == "pdf":
            return self._render_pdf("check_mark_exceptions.html", data).getvalue()
        else:
            headers = ["Exception Reason", "Count"]
            rows = [[d["reason"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Check Mark Exceptions", data["report_name"]).getvalue()

    def generate_internet_bidding_analytics_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Internet bidding analytics: Rule 8 bidding participation and infractions."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        bids = self.db.query(JobBid).filter(
            JobBid.created_at >= start,
            JobBid.created_at <= end
        ).all()

        bid_statuses = {}
        for bid in bids:
            status = bid.status.value if bid.status else "UNKNOWN"
            bid_statuses[status] = bid_statuses.get(status, 0) + 1

        total_bids = len(bids)
        unique_members = len(set(b.member_id for b in bids))

        data = {
            "data": [{"status": k, "count": v} for k, v in bid_statuses.items()],
            "summary": {
                "total_bids": total_bids,
                "unique_members": unique_members,
                "avg_bids_per_member": total_bids / unique_members if unique_members > 0 else 0,
            },
            "filters": {"start_date": start, "end_date": end},
            "generated_at": datetime.now(),
            "report_name": "Internet Bidding Analytics",
        }

        if format == "pdf":
            return self._render_pdf("internet_bidding_analytics.html", data).getvalue()
        else:
            headers = ["Status", "Count"]
            rows = [[d["status"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Bidding Analytics", data["report_name"]).getvalue()

    def generate_exemption_status_report(
        self,
        format: str = "pdf",
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """Exemption status: Rule 14 active exemptions by type."""
        from src.models.member_exemption import MemberExemption
        from src.db.enums import ExemptReason

        as_of = as_of_date or date.today()

        # Active exemptions at this date
        exemptions = self.db.query(MemberExemption).filter(
            MemberExemption.start_date <= as_of,
            MemberExemption.end_date >= as_of
        ).all()

        reason_counts = {}
        for exempt in exemptions:
            reason = exempt.reason.value if exempt.reason else "Unknown"
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

        total = len(exemptions)

        data = {
            "data": [{"reason": k, "count": v} for k, v in sorted(
                reason_counts.items(), key=lambda x: x[1], reverse=True
            )],
            "summary": {
                "total_active_exemptions": total,
                "reason_types": len(reason_counts),
            },
            "filters": {"as_of_date": as_of},
            "generated_at": datetime.now(),
            "report_name": "Exemption Status Report",
        }

        if format == "pdf":
            return self._render_pdf("exemption_status.html", data).getvalue()
        else:
            headers = ["Exemption Reason", "Count"]
            rows = [[d["reason"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Exemptions", data["report_name"]).getvalue()

    def generate_agreement_type_performance_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        require_officer_access: bool = True,
    ) -> bytes:
        """Agreement type performance: Rule 4 CBA analytics (Officer+ only)."""
        from src.db.enums import AgreementType

        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        # Group dispatches by agreement type
        agreement_types = {}
        for atype in [AgreementType.STANDARD, AgreementType.PLA, AgreementType.CWA, AgreementType.TERO]:
            count = self.db.query(func.count(Dispatch.id)).join(
                ReferralBook, Dispatch.book_id == ReferralBook.id
            ).filter(
                Dispatch.dispatch_date >= start,
                Dispatch.dispatch_date <= end,
                ReferralBook.agreement_type == atype
            ).scalar() or 0
            agreement_types[atype.value] = count

        total = sum(agreement_types.values())

        data = {
            "data": [{"agreement_type": k, "dispatches": v} for k, v in agreement_types.items()],
            "summary": {
                "total_dispatches": total,
                "agreement_types_active": len([v for v in agreement_types.values() if v > 0]),
            },
            "filters": {"start_date": start, "end_date": end, "officer_restricted": require_officer_access},
            "generated_at": datetime.now(),
            "report_name": "Agreement Type Performance (Officer Only)",
        }

        if format == "pdf":
            return self._render_pdf("agreement_type_performance.html", data).getvalue()
        else:
            headers = ["Agreement Type", "Dispatches", "Percentage"]
            rows = [
                [d["agreement_type"], d["dispatches"], f"{d['dispatches']/total*100:.1f}%" if total else "0%"]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Agreement Types", data["report_name"]).getvalue()

    def generate_foreperson_by_name_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        require_officer_access: bool = True,
    ) -> bytes:
        """Foreperson by-name report: Rule 13 anti-collusion enforcement (Officer+ only)."""
        start = start_date or (date.today() - timedelta(days=30))
        end = end_date or date.today()

        by_name_dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start,
            Dispatch.dispatch_date <= end,
            Dispatch.dispatch_method == "BY_NAME"
        ).all()

        # Group by foreperson/employer
        foreperson_stats = {}
        for dispatch in by_name_dispatches:
            if dispatch.labor_request and dispatch.labor_request.employer:
                key = dispatch.labor_request.employer.name
                foreperson_stats[key] = foreperson_stats.get(key, 0) + 1

        sorted_stats = sorted(foreperson_stats.items(), key=lambda x: x[1], reverse=True)[:15]

        data = {
            "data": [{"employer": emp, "by_name_count": count} for emp, count in sorted_stats],
            "summary": {
                "total_by_name_dispatches": len(by_name_dispatches),
                "employers_using_by_name": len(foreperson_stats),
            },
            "filters": {"start_date": start, "end_date": end, "officer_restricted": require_officer_access},
            "generated_at": datetime.now(),
            "report_name": "Foreperson By-Name Report (Officer Only)",
        }

        if format == "pdf":
            return self._render_pdf("foreperson_by_name.html", data).getvalue()
        else:
            headers = ["Employer", "By-Name Count"]
            rows = [[d["employer"], d["by_name_count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "By-Name Dispatch", data["report_name"]).getvalue()

    def generate_blackout_period_tracking_report(
        self,
        format: str = "pdf",
        as_of_date: Optional[date] = None,
        require_officer_access: bool = True,
    ) -> bytes:
        """Blackout period tracking: Rule 12 quit/discharge enforcement (Officer+ only)."""
        from src.models.blackout_period import BlackoutPeriod

        as_of = as_of_date or date.today()

        # Active blackout periods
        active_blackouts = self.db.query(BlackoutPeriod).filter(
            BlackoutPeriod.start_date <= as_of,
            BlackoutPeriod.end_date >= as_of
        ).all()

        reason_counts = {}
        for blackout in active_blackouts:
            reason = blackout.reason or "Unknown"
            reason_counts[reason] = reason_counts.get(reason, 0) + 1

        total_active = len(active_blackouts)

        # Upcoming expirations (within 7 days)
        upcoming = self.db.query(BlackoutPeriod).filter(
            BlackoutPeriod.end_date > as_of,
            BlackoutPeriod.end_date <= as_of + timedelta(days=7)
        ).count()

        data = {
            "data": [{"reason": k, "count": v} for k, v in sorted(
                reason_counts.items(), key=lambda x: x[1], reverse=True
            )],
            "summary": {
                "total_active_blackouts": total_active,
                "upcoming_expirations": upcoming,
                "reason_types": len(reason_counts),
            },
            "filters": {"as_of_date": as_of, "officer_restricted": require_officer_access},
            "generated_at": datetime.now(),
            "report_name": "Blackout Period Tracking (Officer Only)",
        }

        if format == "pdf":
            return self._render_pdf("blackout_period_tracking.html", data).getvalue()
        else:
            headers = ["Blackout Reason", "Count"]
            rows = [[d["reason"], d["count"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Blackouts", data["report_name"]).getvalue()

    # === PHASE 3: FORECASTING, INTELLIGENCE & ADMINISTRATIVE (10 Reports) ===
    # Week 40/41 P3 Reports — Low-priority but valuable long-term insights

    # P3-A: FORECASTING (3 reports)

    def generate_workforce_projection_report(
        self,
        format: str = "pdf",
        days_forward: int = 90,
    ) -> bytes:
        """Project 30/60/90 day queue levels. Gracefully handle insufficient data."""
        end_date = date.today()
        start_date = end_date - timedelta(days=90)

        # Query historical dispatch velocity
        historical_dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start_date,
            Dispatch.dispatch_date <= end_date
        ).all()

        # Check if we have sufficient data (90+ days)
        if len(historical_dispatches) < 90:
            data = {
                "data": [],
                "summary": {
                    "message": "Insufficient historical data",
                    "details": f"Only {len(historical_dispatches)} dispatch records available. System requires 90+ days of data for accurate projections."
                },
                "filters": {"days_forward": days_forward},
                "generated_at": datetime.now(),
                "report_name": "Workforce Projection (Insufficient Data)",
            }
            if format == "pdf":
                return self._render_pdf("workforce_projection.html", data).getvalue()
            else:
                headers = ["Period", "Projected Queue Level"]
                rows = [["Data Insufficient", "—"]]
                return self._render_excel(headers, rows, "Projection", data["report_name"]).getvalue()

        # Calculate average daily dispatches
        days_with_data = (end_date - start_date).days
        avg_daily_dispatches = len(historical_dispatches) / max(days_with_data, 1)

        # Get current registrations by book
        current_registrations = self.db.query(
            ReferralBook.name,
            func.count(BookRegistration.id).label("count")
        ).join(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.ACTIVE
        ).group_by(ReferralBook.name).all()

        # Project 30/60/90 days
        projections = []
        for period_days in [30, 60, 90]:
            if period_days <= days_forward:
                projected_outflow = avg_daily_dispatches * period_days
                projections.append({
                    "period": f"{period_days}-Day",
                    "projected_outflow": round(projected_outflow, 0),
                    "confidence": "High" if days_with_data >= 90 else "Medium"
                })

        data = {
            "data": projections,
            "current_books": [
                {"name": name, "active_count": count}
                for name, count in current_registrations
            ],
            "summary": {
                "avg_daily_dispatches": round(avg_daily_dispatches, 1),
                "historical_period_days": days_with_data,
                "data_quality": "Complete" if days_with_data >= 90 else "Partial"
            },
            "filters": {"days_forward": days_forward},
            "generated_at": datetime.now(),
            "report_name": "Workforce Projection Report",
        }

        if format == "pdf":
            return self._render_pdf("workforce_projection.html", data).getvalue()
        else:
            headers = ["Period", "Projected Outflow", "Confidence"]
            rows = [[d["period"], d["projected_outflow"], d["confidence"]] for d in projections]
            return self._render_excel(headers, rows, "Projections", data["report_name"]).getvalue()

    def generate_dispatch_forecast_report(
        self,
        format: str = "pdf",
        forecast_month: Optional[date] = None,
    ) -> bytes:
        """Next month dispatch volume by classification. Gracefully handle insufficient data."""
        if not forecast_month:
            forecast_month = date.today()

        # Use past 6 months for forecast training
        training_start = forecast_month - timedelta(days=180)
        training_end = forecast_month - timedelta(days=1)

        historical = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= training_start,
            Dispatch.dispatch_date <= training_end
        ).all()

        # Check sufficient data (6+ months × 20+ dispatches)
        if len(historical) < 120:
            data = {
                "data": [],
                "summary": {
                    "message": "Insufficient historical data",
                    "details": f"Only {len(historical)} dispatch records in past 6 months. System requires 120+ records for volume forecasting."
                },
                "filters": {"forecast_month": forecast_month},
                "generated_at": datetime.now(),
                "report_name": "Dispatch Forecast (Insufficient Data)",
            }
            if format == "pdf":
                return self._render_pdf("dispatch_forecast.html", data).getvalue()
            else:
                headers = ["Classification", "Forecasted Dispatches"]
                rows = [["Data Insufficient", "—"]]
                return self._render_excel(headers, rows, "Forecast", data["report_name"]).getvalue()

        # Group by classification (via book)
        classification_counts = {}
        for dispatch in historical:
            if dispatch.book_registration and dispatch.book_registration.referral_book:
                book_name = dispatch.book_registration.referral_book.name
                classification_counts[book_name] = classification_counts.get(book_name, 0) + 1

        # Average per month
        forecast_data = [
            {
                "classification": book,
                "average_monthly_volume": round(count / 6, 1),
                "forecast_volume": round(count / 6, 0)
            }
            for book, count in sorted(classification_counts.items(), key=lambda x: x[1], reverse=True)
        ]

        data = {
            "data": forecast_data,
            "summary": {
                "total_historical_dispatches": len(historical),
                "training_months": 6,
                "forecast_month": forecast_month.strftime("%B %Y"),
                "data_quality": "Complete" if len(historical) >= 120 else "Partial"
            },
            "filters": {"forecast_month": forecast_month},
            "generated_at": datetime.now(),
            "report_name": "Dispatch Forecast Report",
        }

        if format == "pdf":
            return self._render_pdf("dispatch_forecast.html", data).getvalue()
        else:
            headers = ["Classification", "Avg Monthly", "Forecast"]
            rows = [[d["classification"], d["average_monthly_volume"], d["forecast_volume"]] for d in forecast_data]
            return self._render_excel(headers, rows, "Forecast", data["report_name"]).getvalue()

    def generate_book_demand_forecast_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Per-book projected demand (labor requests forecast). Gracefully handle insufficient data."""
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=60)

        requests = self.db.query(LaborRequest).filter(
            LaborRequest.request_date >= start_date,
            LaborRequest.request_date <= end_date
        ).all()

        # Check if we have sufficient data
        if len(requests) < 20:
            data = {
                "data": [],
                "summary": {
                    "message": "Insufficient historical data",
                    "details": f"Only {len(requests)} labor requests in past 60 days. System requires 20+ requests for demand forecasting."
                },
                "filters": {"start_date": start_date, "end_date": end_date},
                "generated_at": datetime.now(),
                "report_name": "Book Demand Forecast (Insufficient Data)",
            }
            if format == "pdf":
                return self._render_pdf("book_demand_forecast.html", data).getvalue()
            else:
                headers = ["Book", "Forecasted Requests"]
                rows = [["Data Insufficient", "—"]]
                return self._render_excel(headers, rows, "Forecast", data["report_name"]).getvalue()

        # Group by book
        book_demand = {}
        total_requested = 0
        for req in requests:
            book_name = req.referral_book.name if req.referral_book else "Unknown"
            book_demand[book_name] = book_demand.get(book_name, 0) + req.workers_requested
            total_requested += req.workers_requested

        forecast_data = [
            {
                "book": book,
                "historical_demand": count,
                "demand_percentage": round(count / total_requested * 100, 1) if total_requested else 0,
            }
            for book, count in sorted(book_demand.items(), key=lambda x: x[1], reverse=True)
        ]

        data = {
            "data": forecast_data,
            "summary": {
                "total_requests": len(requests),
                "total_workers_demanded": total_requested,
                "period_days": (end_date - start_date).days,
                "data_quality": "Complete" if len(requests) >= 20 else "Partial"
            },
            "filters": {"start_date": start_date, "end_date": end_date},
            "generated_at": datetime.now(),
            "report_name": "Book Demand Forecast Report",
        }

        if format == "pdf":
            return self._render_pdf("book_demand_forecast.html", data).getvalue()
        else:
            headers = ["Book", "Historical Demand", "Percentage"]
            rows = [[d["book"], d["historical_demand"], f"{d['demand_percentage']}%"] for d in forecast_data]
            return self._render_excel(headers, rows, "Demand", data["report_name"]).getvalue()

    # P3-B: INTELLIGENCE (4 reports)

    def generate_member_availability_index_report(
        self,
        format: str = "pdf",
        as_of_date: Optional[date] = None,
    ) -> bytes:
        """Member availability index: Book fill capability score."""
        as_of = as_of_date or date.today()

        # Count active registrations by book
        book_stats = self.db.query(
            ReferralBook.name,
            func.count(BookRegistration.id).label("active_members")
        ).join(BookRegistration).filter(
            BookRegistration.status == RegistrationStatus.ACTIVE,
            BookRegistration.book_id == ReferralBook.id
        ).group_by(ReferralBook.id, ReferralBook.name).all()

        # Calculate recent dispatch velocity per book
        thirty_days_ago = as_of - timedelta(days=30)
        recent_dispatches = self.db.query(
            ReferralBook.name,
            func.count(Dispatch.id).label("dispatch_count")
        ).join(BookRegistration).join(Dispatch).filter(
            Dispatch.dispatch_date >= thirty_days_ago
        ).group_by(ReferralBook.id, ReferralBook.name).all()

        dispatch_map = {name: count for name, count in recent_dispatches}

        # Calculate index (active_members / recent_monthly_dispatches)
        availability_data = []
        for book_name, active_count in book_stats:
            monthly_velocity = dispatch_map.get(book_name, 0)
            # Avoid division by zero; higher index = better availability
            index = round((active_count / max(monthly_velocity, 1)), 1) if active_count > 0 else 0
            availability_data.append({
                "book": book_name,
                "active_members": active_count,
                "monthly_dispatches": monthly_velocity,
                "availability_index": index,
                "fill_capability": "Excellent" if index >= 3 else "Good" if index >= 1.5 else "Tight"
            })

        data = {
            "data": sorted(availability_data, key=lambda x: x["availability_index"], reverse=True),
            "summary": {
                "total_active_members": sum(x["active_members"] for x in availability_data),
                "avg_index": round(sum(x["availability_index"] for x in availability_data) / max(len(availability_data), 1), 2),
                "period": "Last 30 days"
            },
            "filters": {"as_of_date": as_of},
            "generated_at": datetime.now(),
            "report_name": "Member Availability Index Report",
        }

        if format == "pdf":
            return self._render_pdf("member_availability_index.html", data).getvalue()
        else:
            headers = ["Book", "Active Members", "Monthly Dispatches", "Availability Index", "Fill Capability"]
            rows = [
                [d["book"], d["active_members"], d["monthly_dispatches"], d["availability_index"], d["fill_capability"]]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Availability", data["report_name"]).getvalue()

    def generate_employer_loyalty_score_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        require_officer_access: bool = True,
    ) -> bytes:
        """Employer loyalty composite ranking (Officer+ only): repeat usage, fill rate, retention."""
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=180)

        dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= start_date,
            Dispatch.dispatch_date <= end_date
        ).all()

        # Aggregate by employer
        employer_stats = {}
        for dispatch in dispatches:
            if dispatch.labor_request and dispatch.labor_request.employer:
                emp_id = dispatch.labor_request.employer.id
                emp_name = dispatch.labor_request.employer.name

                if emp_id not in employer_stats:
                    employer_stats[emp_id] = {
                        "name": emp_name,
                        "total_dispatches": 0,
                        "filled_requests": 0,
                        "repeat_workers": set(),
                    }

                employer_stats[emp_id]["total_dispatches"] += 1
                if dispatch.status == DispatchStatus.FILLED:
                    employer_stats[emp_id]["filled_requests"] += 1
                if dispatch.member_id:
                    employer_stats[emp_id]["repeat_workers"].add(dispatch.member_id)

        # Calculate loyalty score (composite: repeat requests × fill rate × worker retention)
        loyalty_data = []
        for emp_id, stats in employer_stats.items():
            repeat_score = min(stats["total_dispatches"] / 10, 1.0)  # Normalize to 0-1
            fill_rate = stats["filled_requests"] / max(stats["total_dispatches"], 1)
            retention = len(stats["repeat_workers"]) / max(stats["total_dispatches"], 1)

            composite_score = round((repeat_score + fill_rate + retention) / 3 * 100, 1)

            loyalty_data.append({
                "employer": stats["name"],
                "total_requests": stats["total_dispatches"],
                "fill_rate": round(fill_rate * 100, 1),
                "repeat_workers": len(stats["repeat_workers"]),
                "loyalty_score": composite_score,
                "tier": "Gold" if composite_score >= 80 else "Silver" if composite_score >= 60 else "Bronze"
            })

        data = {
            "data": sorted(loyalty_data, key=lambda x: x["loyalty_score"], reverse=True)[:20],
            "summary": {
                "total_employers": len(employer_stats),
                "avg_loyalty_score": round(sum(x["loyalty_score"] for x in loyalty_data) / max(len(loyalty_data), 1), 1),
                "period": f"{start_date} to {end_date}"
            },
            "filters": {
                "start_date": start_date,
                "end_date": end_date,
                "officer_restricted": require_officer_access
            },
            "generated_at": datetime.now(),
            "report_name": "Employer Loyalty Score Report (Officer Only)",
        }

        if format == "pdf":
            return self._render_pdf("employer_loyalty_score.html", data).getvalue()
        else:
            headers = ["Employer", "Total Requests", "Fill Rate %", "Repeat Workers", "Loyalty Score", "Tier"]
            rows = [
                [d["employer"], d["total_requests"], d["fill_rate"], d["repeat_workers"], d["loyalty_score"], d["tier"]]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Loyalty", data["report_name"]).getvalue()

    def generate_member_journey_report(
        self,
        format: str = "pdf",
        member_id: int = None,
        require_officer_access: bool = True,
    ) -> bytes:
        """Individual member lifecycle: registrations, dispatches, check marks, exemptions (Officer+ only, requires member_id)."""
        if not member_id:
            data = {
                "data": [],
                "summary": {"message": "No member specified", "details": "member_id parameter required"},
                "filters": {"member_id": None},
                "generated_at": datetime.now(),
                "report_name": "Member Journey (No Member Selected)",
            }
            if format == "pdf":
                return self._render_pdf("member_journey.html", data).getvalue()
            else:
                headers = ["Parameter", "Value"]
                rows = [["member_id", "Required"]]
                return self._render_excel(headers, rows, "Journey", data["report_name"]).getvalue()

        member = self.db.query(Member).filter(Member.id == member_id).first()
        if not member:
            data = {
                "data": [],
                "summary": {"message": "Member not found", "details": f"Member ID {member_id} does not exist"},
                "filters": {"member_id": member_id},
                "generated_at": datetime.now(),
                "report_name": "Member Journey (Not Found)",
            }
            if format == "pdf":
                return self._render_pdf("member_journey.html", data).getvalue()
            else:
                headers = ["Status", "Message"]
                rows = [["Error", f"Member {member_id} not found"]]
                return self._render_excel(headers, rows, "Journey", data["report_name"]).getvalue()

        # Collect member's referral history
        registrations = self.db.query(BookRegistration).filter(
            BookRegistration.member_id == member_id
        ).order_by(BookRegistration.created_at).all()

        dispatches = self.db.query(Dispatch).filter(
            Dispatch.member_id == member_id
        ).order_by(Dispatch.dispatch_date).all()

        journey_timeline = []
        for reg in registrations:
            journey_timeline.append({
                "date": reg.created_at.date() if reg.created_at else None,
                "event": "Registered",
                "book": reg.referral_book.name if reg.referral_book else "Unknown",
                "detail": f"Status: {reg.status.value if hasattr(reg.status, 'value') else reg.status}"
            })

        for dispatch in dispatches:
            journey_timeline.append({
                "date": dispatch.dispatch_date,
                "event": "Dispatched",
                "book": dispatch.book_registration.referral_book.name if dispatch.book_registration and dispatch.book_registration.referral_book else "Unknown",
                "detail": f"Status: {dispatch.status.value if hasattr(dispatch.status, 'value') else dispatch.status}"
            })

        data = {
            "data": sorted(journey_timeline, key=lambda x: x["date"] or date.min),
            "member": {
                "id": member.id,
                "name": f"{member.first_name} {member.last_name}",
                "email": member.email
            },
            "summary": {
                "total_registrations": len(registrations),
                "total_dispatches": len(dispatches),
                "active_on_books": sum(1 for r in registrations if r.status == RegistrationStatus.ACTIVE),
            },
            "filters": {"member_id": member_id, "officer_restricted": require_officer_access},
            "generated_at": datetime.now(),
            "report_name": f"Member Journey Report - {member.first_name} {member.last_name}",
        }

        if format == "pdf":
            return self._render_pdf("member_journey.html", data).getvalue()
        else:
            headers = ["Date", "Event", "Book", "Detail"]
            rows = [[d["date"], d["event"], d["book"], d["detail"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Journey", data["report_name"]).getvalue()

    def generate_comparative_book_performance_report(
        self,
        format: str = "pdf",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
    ) -> bytes:
        """Normalized cross-book comparison: velocity, fill rate, queue health."""
        if not end_date:
            end_date = date.today()
        if not start_date:
            start_date = end_date - timedelta(days=90)

        books = self.db.query(ReferralBook).all()
        comparison_data = []

        for book in books:
            # Active registrations
            active_regs = self.db.query(func.count(BookRegistration.id)).filter(
                BookRegistration.referral_book_id == book.id,
                BookRegistration.status == RegistrationStatus.ACTIVE
            ).scalar() or 0

            # Dispatches in period
            dispatches = self.db.query(func.count(Dispatch.id)).filter(
                Dispatch.dispatch_date >= start_date,
                Dispatch.dispatch_date <= end_date,
                Dispatch.book_registration_id.in_(
                    self.db.query(BookRegistration.id).filter(
                        BookRegistration.referral_book_id == book.id
                    )
                )
            ).scalar() or 0

            # Unfilled requests
            unfilled = self.db.query(func.count(LaborRequest.id)).filter(
                LaborRequest.referral_book_id == book.id,
                LaborRequest.request_date >= start_date,
                LaborRequest.request_date <= end_date,
                LaborRequest.workers_dispatched == 0
            ).scalar() or 0

            # Normalize metrics (0-100 scale)
            velocity_score = min((dispatches / max(active_regs, 1)) * 20, 100) if active_regs > 0 else 0
            fill_score = 100 - min((unfilled / max(dispatches + unfilled, 1)) * 100, 100)

            overall_score = round((velocity_score + fill_score) / 2, 1)

            comparison_data.append({
                "book": book.name,
                "active_members": active_regs,
                "dispatches": dispatches,
                "unfilled_requests": unfilled,
                "velocity_score": round(velocity_score, 1),
                "fill_score": round(fill_score, 1),
                "overall_performance": overall_score,
                "performance_tier": "Excellent" if overall_score >= 80 else "Good" if overall_score >= 60 else "Fair"
            })

        data = {
            "data": sorted(comparison_data, key=lambda x: x["overall_performance"], reverse=True),
            "summary": {
                "total_books": len(books),
                "avg_performance": round(sum(x["overall_performance"] for x in comparison_data) / max(len(comparison_data), 1), 1),
                "period": f"{start_date} to {end_date}",
                "total_dispatches": sum(x["dispatches"] for x in comparison_data)
            },
            "filters": {"start_date": start_date, "end_date": end_date},
            "generated_at": datetime.now(),
            "report_name": "Comparative Book Performance Report",
        }

        if format == "pdf":
            return self._render_pdf("comparative_book_performance.html", data).getvalue()
        else:
            headers = ["Book", "Active Members", "Dispatches", "Unfilled Requests", "Velocity Score", "Fill Score", "Overall Performance", "Tier"]
            rows = [
                [d["book"], d["active_members"], d["dispatches"], d["unfilled_requests"],
                 d["velocity_score"], d["fill_score"], d["overall_performance"], d["performance_tier"]]
                for d in data["data"]
            ]
            return self._render_excel(headers, rows, "Performance", data["report_name"]).getvalue()

    # P3-C: ADMINISTRATIVE (3 reports)

    def generate_custom_export_report(
        self,
        format: str = "excel",
        entity_type: str = "members",
        filters: Optional[dict] = None,
    ) -> bytes:
        """Ad-hoc data dump: flexible entity_type parameter. Excel only."""
        filters = filters or {}

        if entity_type == "members":
            query = self.db.query(Member)
            if filters.get("status"):
                query = query.filter(Member.status == filters["status"])

            members = query.all()
            headers = ["ID", "Name", "Email", "Status", "Classification", "Current Employer"]
            rows = [
                [
                    m.id,
                    f"{m.first_name} {m.last_name}",
                    m.email,
                    m.status.value if hasattr(m.status, 'value') else m.status,
                    m.classification.value if m.classification and hasattr(m.classification, 'value') else m.classification,
                    m.current_employer.name if m.current_employer else "—"
                ]
                for m in members
            ]

        elif entity_type == "registrations":
            query = self.db.query(BookRegistration)
            if filters.get("status"):
                query = query.filter(BookRegistration.status == filters["status"])
            if filters.get("book_id"):
                query = query.filter(BookRegistration.referral_book_id == filters["book_id"])

            regs = query.all()
            headers = ["Member ID", "Member Name", "Book", "Status", "APN", "Registered Date"]
            rows = [
                [
                    r.member_id,
                    f"{r.member.first_name} {r.member.last_name}" if r.member else "—",
                    r.referral_book.name if r.referral_book else "—",
                    r.status.value if hasattr(r.status, 'value') else r.status,
                    r.applicant_priority_number or "—",
                    r.created_at.strftime("%m/%d/%Y") if r.created_at else "—"
                ]
                for r in regs
            ]

        elif entity_type == "dispatches":
            query = self.db.query(Dispatch)
            if filters.get("start_date"):
                query = query.filter(Dispatch.dispatch_date >= filters["start_date"])
            if filters.get("end_date"):
                query = query.filter(Dispatch.dispatch_date <= filters["end_date"])

            dispatches = query.all()
            headers = ["Date", "Member", "Book", "Employer", "Status", "Dispatch Method"]
            rows = [
                [
                    d.dispatch_date.strftime("%m/%d/%Y") if d.dispatch_date else "—",
                    f"{d.member.first_name} {d.member.last_name}" if d.member else "—",
                    d.book_registration.referral_book.name if d.book_registration and d.book_registration.referral_book else "—",
                    d.labor_request.employer.name if d.labor_request and d.labor_request.employer else "—",
                    d.status.value if hasattr(d.status, 'value') else d.status,
                    d.dispatch_method if d.dispatch_method else "—"
                ]
                for d in dispatches
            ]

        else:
            headers = ["Error"]
            rows = [[f"Unknown entity_type: {entity_type}"]]

        data = {
            "report_name": f"Custom Export - {entity_type.title()}",
            "entity_type": entity_type,
            "record_count": len(rows),
            "filters": filters
        }

        return self._render_excel(headers, rows, entity_type.title(), data["report_name"]).getvalue()

    def generate_annual_summary_report(
        self,
        format: str = "pdf",
        year: Optional[int] = None,
    ) -> bytes:
        """Year-in-review: dispatch volume, registrations, major events."""
        if not year:
            year = date.today().year

        year_start = date(year, 1, 1)
        year_end = date(year, 12, 31)

        # Annual dispatches
        annual_dispatches = self.db.query(Dispatch).filter(
            Dispatch.dispatch_date >= year_start,
            Dispatch.dispatch_date <= year_end
        ).all()

        # New registrations
        annual_registrations = self.db.query(BookRegistration).filter(
            BookRegistration.created_at >= datetime.combine(year_start, time.min),
            BookRegistration.created_at <= datetime.combine(year_end, time.max)
        ).all()

        # Monthly breakdown
        monthly_data = {}
        for month in range(1, 13):
            month_start = date(year, month, 1)
            if month == 12:
                month_end = date(year, 12, 31)
            else:
                month_end = date(year, month + 1, 1) - timedelta(days=1)

            month_count = sum(1 for d in annual_dispatches if month_start <= d.dispatch_date <= month_end)
            monthly_data[month] = month_count

        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

        data = {
            "data": [{"month": month_names[i], "dispatches": monthly_data.get(i + 1, 0)} for i in range(12)],
            "summary": {
                "year": year,
                "total_dispatches": len(annual_dispatches),
                "total_new_registrations": len(annual_registrations),
                "avg_monthly_dispatches": round(len(annual_dispatches) / 12, 1),
                "avg_filled_rate": round(
                    sum(1 for d in annual_dispatches if d.status == DispatchStatus.FILLED) / max(len(annual_dispatches), 1) * 100, 1
                ) if annual_dispatches else 0
            },
            "filters": {"year": year},
            "generated_at": datetime.now(),
            "report_name": f"Annual Summary Report - {year}",
        }

        if format == "pdf":
            return self._render_pdf("annual_summary.html", data).getvalue()
        else:
            headers = ["Month", "Dispatches"]
            rows = [[d["month"], d["dispatches"]] for d in data["data"]]
            return self._render_excel(headers, rows, "Annual Summary", data["report_name"]).getvalue()

    def generate_data_quality_report(
        self,
        format: str = "pdf",
        require_admin_access: bool = True,
    ) -> bytes:
        """Data hygiene audit: missing fields, orphaned records, inconsistencies (Admin only)."""
        issues = []

        # Check for registrations with missing member references
        orphaned_registrations = self.db.query(func.count(BookRegistration.id)).filter(
            BookRegistration.member_id.is_(None)
        ).scalar() or 0
        if orphaned_registrations > 0:
            issues.append({
                "category": "Orphaned Records",
                "issue": "Registrations without member",
                "count": orphaned_registrations,
                "severity": "High"
            })

        # Check for dispatches with missing labor requests
        orphaned_dispatches = self.db.query(func.count(Dispatch.id)).filter(
            Dispatch.labor_request_id.is_(None)
        ).scalar() or 0
        if orphaned_dispatches > 0:
            issues.append({
                "category": "Orphaned Records",
                "issue": "Dispatches without labor request",
                "count": orphaned_dispatches,
                "severity": "High"
            })

        # Check for members without contact info
        no_email = self.db.query(func.count(Member.id)).filter(
            Member.email.is_(None)
        ).scalar() or 0
        if no_email > 0:
            issues.append({
                "category": "Missing Data",
                "issue": "Members without email",
                "count": no_email,
                "severity": "Medium"
            })

        # Check for labor requests without books
        no_book = self.db.query(func.count(LaborRequest.id)).filter(
            LaborRequest.referral_book_id.is_(None)
        ).scalar() or 0
        if no_book > 0:
            issues.append({
                "category": "Missing Data",
                "issue": "Labor requests without book",
                "count": no_book,
                "severity": "High"
            })

        # Data completeness stats
        total_members = self.db.query(func.count(Member.id)).scalar() or 0
        total_registrations = self.db.query(func.count(BookRegistration.id)).scalar() or 0
        total_dispatches = self.db.query(func.count(Dispatch.id)).scalar() or 0

        data = {
            "data": issues,
            "summary": {
                "total_issues": len(issues),
                "high_severity": sum(1 for i in issues if i["severity"] == "High"),
                "medium_severity": sum(1 for i in issues if i["severity"] == "Medium"),
                "total_members": total_members,
                "total_registrations": total_registrations,
                "total_dispatches": total_dispatches,
                "data_quality_score": max(0, 100 - (len(issues) * 10))  # Simplified scoring
            },
            "filters": {"admin_restricted": require_admin_access},
            "generated_at": datetime.now(),
            "report_name": "Data Quality Audit Report (Admin Only)",
        }

        if format == "pdf":
            return self._render_pdf("data_quality_report.html", data).getvalue()
        else:
            headers = ["Category", "Issue", "Count", "Severity"]
            rows = [[i["category"], i["issue"], i["count"], i["severity"]] for i in issues]
            return self._render_excel(headers, rows, "Data Quality", data["report_name"]).getvalue()
