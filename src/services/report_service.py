"""Report generation service."""

import io
from datetime import datetime
from typing import Optional, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Lazy import for WeasyPrint (requires system libraries)
_weasyprint_available = None


def _get_weasyprint():
    """Lazy load WeasyPrint - only imported when PDF generation is needed."""
    global _weasyprint_available
    if _weasyprint_available is None:
        try:
            from weasyprint import HTML, CSS
            _weasyprint_available = (HTML, CSS)
        except (ImportError, OSError):
            _weasyprint_available = False
    return _weasyprint_available


class ReportService:
    """Service for generating PDF and Excel reports."""

    # Default CSS for PDF reports
    DEFAULT_CSS = """
        @page {
            size: letter;
            margin: 1in;
            @top-center {
                content: "IP2A Database Report";
                font-size: 10pt;
                color: #666;
            }
            @bottom-center {
                content: "Page " counter(page) " of " counter(pages);
                font-size: 10pt;
                color: #666;
            }
        }

        body {
            font-family: 'Helvetica Neue', Arial, sans-serif;
            font-size: 11pt;
            line-height: 1.4;
            color: #333;
        }

        h1 {
            color: #1a365d;
            font-size: 24pt;
            margin-bottom: 0.5em;
            border-bottom: 2px solid #1a365d;
            padding-bottom: 0.25em;
        }

        h2 {
            color: #2c5282;
            font-size: 16pt;
            margin-top: 1em;
            margin-bottom: 0.5em;
        }

        h3 {
            color: #2d3748;
            font-size: 13pt;
            margin-top: 0.75em;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            margin: 1em 0;
            font-size: 10pt;
        }

        th {
            background-color: #edf2f7;
            border: 1px solid #cbd5e0;
            padding: 8px 12px;
            text-align: left;
            font-weight: 600;
        }

        td {
            border: 1px solid #e2e8f0;
            padding: 8px 12px;
        }

        tr:nth-child(even) {
            background-color: #f7fafc;
        }

        .header {
            display: flex;
            justify-content: space-between;
            margin-bottom: 1.5em;
        }

        .meta {
            font-size: 10pt;
            color: #718096;
        }

        .stat-box {
            background-color: #ebf8ff;
            border: 1px solid #90cdf4;
            border-radius: 4px;
            padding: 12px;
            margin: 0.5em 0;
        }

        .stat-label {
            font-size: 10pt;
            color: #4a5568;
        }

        .stat-value {
            font-size: 18pt;
            font-weight: 600;
            color: #2b6cb0;
        }

        .badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 9pt;
            font-weight: 500;
        }

        .badge-success { background-color: #c6f6d5; color: #22543d; }
        .badge-warning { background-color: #fefcbf; color: #744210; }
        .badge-error { background-color: #fed7d7; color: #822727; }
        .badge-info { background-color: #bee3f8; color: #2a4365; }

        .certificate {
            text-align: center;
            padding: 2em;
            border: 3px double #1a365d;
            margin: 1em;
        }

        .certificate h1 {
            border: none;
            font-size: 28pt;
        }

        .certificate .name {
            font-size: 24pt;
            color: #2d3748;
            margin: 1em 0;
        }

        .footer {
            margin-top: 2em;
            padding-top: 1em;
            border-top: 1px solid #e2e8f0;
            font-size: 9pt;
            color: #718096;
        }
    """

    @staticmethod
    def generate_pdf(
        html_content: str,
        css: Optional[str] = None,
        base_url: Optional[str] = None,
    ) -> bytes:
        """
        Generate PDF from HTML content.

        Args:
            html_content: HTML string to convert
            css: Optional custom CSS (uses default if not provided)
            base_url: Base URL for resolving relative paths

        Returns:
            PDF as bytes

        Raises:
            RuntimeError: If WeasyPrint is not available
        """
        weasyprint = _get_weasyprint()
        if not weasyprint:
            raise RuntimeError(
                "WeasyPrint is not available. Install system dependencies: "
                "libpango-1.0-0 libpangocairo-1.0-0 libgdk-pixbuf2.0-0"
            )

        HTML, CSS = weasyprint

        if css is None:
            css = ReportService.DEFAULT_CSS

        html = HTML(string=html_content, base_url=base_url)
        stylesheet = CSS(string=css)

        pdf_buffer = io.BytesIO()
        html.write_pdf(pdf_buffer, stylesheets=[stylesheet])
        pdf_buffer.seek(0)

        return pdf_buffer.getvalue()

    @staticmethod
    def generate_excel(
        data: list[dict[str, Any]],
        columns: list[dict[str, str]],
        sheet_name: str = "Report",
        title: Optional[str] = None,
    ) -> bytes:
        """
        Generate Excel file from data.

        Args:
            data: List of dictionaries with row data
            columns: List of column definitions [{"key": "field", "header": "Display Name"}]
            sheet_name: Name of the worksheet
            title: Optional title row

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row_num = 1

        # Title row (optional)
        if title:
            ws.cell(row=row_num, column=1, value=title)
            ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(columns))
            row_num += 2

        # Header row
        for col_num, col_def in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=col_def["header"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        row_num += 1

        # Data rows
        for row_data in data:
            for col_num, col_def in enumerate(columns, 1):
                value = row_data.get(col_def["key"], "")
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
            row_num += 1

        # Auto-adjust column widths
        for col_num, col_def in enumerate(columns, 1):
            max_length = len(col_def["header"])
            for row in range(1, row_num):
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Add metadata footer
        ws.cell(row=row_num + 1, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=row_num + 1, column=1).font = Font(italic=True, color="666666")

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    @staticmethod
    def format_currency(amount) -> str:
        """Format amount as currency string."""
        if amount is None:
            return "$0.00"
        return f"${float(amount):,.2f}"

    @staticmethod
    def format_date(dt, format_str: str = "%B %d, %Y") -> str:
        """Format datetime as string."""
        if dt is None:
            return "—"
        return dt.strftime(format_str)

    @staticmethod
    def format_phone(phone: Optional[str]) -> str:
        """Format phone number."""
        if not phone:
            return "—"
        # Simple formatting for 10-digit numbers
        digits = "".join(filter(str.isdigit, phone))
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        return phone